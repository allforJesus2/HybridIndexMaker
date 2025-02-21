import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook


class ExcelDataPullApp:
    def __init__(self, root, excel_path=None):
        self.root = root
        self.root.title("Excel Search Application")
        self.headers = []
        self.entry_widgets = {}
        self.unique_values = {}  # Store unique values for each column
        self.excel_path = excel_path

        # Main container
        main_container = ttk.Frame(root, padding="10")
        main_container.pack(expand=True, fill="both")

        # File selection frame
        file_frame = ttk.Frame(main_container)
        file_frame.pack(fill="x", pady=(0, 10))

        select_file_btn = ttk.Button(file_frame, text="Select Excel File", command=self.select_file)
        select_file_btn.pack(side="left", padx=5)

        self.file_label = ttk.Label(file_frame, text="No file selected")
        self.file_label.pack(side="left")

        # Create container for input and output frames
        content_container = ttk.Frame(main_container)
        content_container.pack(expand=True, fill="both")

        # Input Frame with scrollbar
        input_frame = ttk.LabelFrame(content_container, text="Search Criteria", padding="5")
        input_frame.pack(side="left", fill="both", expand=True, padx=5)

        self.input_canvas = tk.Canvas(input_frame)
        scrollbar = ttk.Scrollbar(input_frame, orient="vertical", command=self.input_canvas.yview)
        self.scrollable_frame = ttk.Frame(self.input_canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.input_canvas.configure(scrollregion=self.input_canvas.bbox("all"))
        )

        self.input_canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.input_canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        self.input_canvas.pack(side="left", fill="both", expand=True)

        # Output Frame
        output_frame = ttk.LabelFrame(content_container, text="Output Settings", padding="5")
        output_frame.pack(side="left", fill="both", expand=True, padx=5)

        ttk.Label(output_frame, text="Available Columns:").pack(anchor="w")
        self.available_listbox = tk.Listbox(output_frame, height=6)
        self.available_listbox.pack(fill="x", pady=(0, 5))

        # Button frame for add/remove
        btn_frame = ttk.Frame(output_frame)
        btn_frame.pack(fill="x", pady=5)
        ttk.Button(btn_frame, text="Add →", command=self.add_column).pack(side="left", expand=True, padx=2)
        ttk.Button(btn_frame, text="← Remove", command=self.remove_column).pack(side="left", expand=True, padx=2)

        ttk.Label(output_frame, text="Selected Columns (Order):").pack(anchor="w")
        selected_frame = ttk.Frame(output_frame)
        selected_frame.pack(fill="x", pady=(0, 5))

        self.selected_listbox = tk.Listbox(selected_frame, height=6)
        self.selected_listbox.pack(side="left", fill="x", expand=True)

        # Order buttons frame
        order_btn_frame = ttk.Frame(selected_frame)
        order_btn_frame.pack(side="left", padx=2)
        ttk.Button(order_btn_frame, text="▲", command=self.move_up).pack(pady=2)
        ttk.Button(order_btn_frame, text="▼", command=self.move_down).pack(pady=2)

        # Separator frame
        sep_frame = ttk.Frame(output_frame)
        sep_frame.pack(fill="x", pady=(0, 10))
        ttk.Label(sep_frame, text="Separator:").pack(side="left")
        self.separator_entry = ttk.Entry(sep_frame, width=5)
        self.separator_entry.pack(side="left", padx=5)
        self.separator_entry.insert(0, "")

        ttk.Button(output_frame, text="Pull Data", command=self.pull_data).pack(fill="x", pady=(0, 10))
        ttk.Label(output_frame, text="Results:").pack(anchor="w")
        self.results_text = tk.Text(output_frame, height=15, width=40)
        self.results_text.pack(fill="both", expand=True)

    def get_unique_values(self, sheet, header_index):
        """Get unique values from a column"""
        unique_values = set()
        for row in list(sheet.rows)[1:]:  # Skip header row
            cell_value = str(row[header_index].value or "").strip()
            if cell_value:
                unique_values.add(cell_value)
        return sorted(list(unique_values))

    def select_file(self):
        self.excel_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if self.excel_path:
            self.file_label.config(text=self.excel_path)
            try:
                wb = load_workbook(self.excel_path, read_only=True)
                sheet = wb.active
                self.headers = [cell.value for cell in next(sheet.rows)]

                # Get unique values for each column
                self.unique_values = {
                    header: self.get_unique_values(sheet, idx)
                    for idx, header in enumerate(self.headers)
                }

                wb.close()

                # Clear existing widgets
                for widget in self.scrollable_frame.winfo_children():
                    widget.destroy()
                self.entry_widgets.clear()

                # Create input fields with dropdowns
                for header in self.headers:
                    frame = ttk.Frame(self.scrollable_frame)
                    frame.pack(fill="x", pady=2)

                    ttk.Label(frame, text=f"{header}:").pack(side="left")

                    entry = ttk.Entry(frame)
                    entry.pack(side="left", padx=(5, 5), fill="x", expand=True)
                    self.entry_widgets[header] = entry

                    # Add dropdown button
                    dropdown_btn = ttk.Button(
                        frame,
                        text="▼",
                        width=3,
                        command=lambda h=header, e=entry: self.show_dropdown(h, e)
                    )
                    dropdown_btn.pack(side="left")

                # Update available columns listbox
                self.available_listbox.delete(0, tk.END)
                self.selected_listbox.delete(0, tk.END)
                for header in self.headers:
                    self.available_listbox.insert(tk.END, header)

            except Exception as e:
                messagebox.showerror("Error", f"Failed to load Excel file: {str(e)}")

    def show_dropdown(self, header, entry_widget):
        """Show dropdown with unique values"""
        dropdown = tk.Toplevel(self.root)
        dropdown.title(f"Select {header}")
        dropdown.transient(self.root)

        # Position dropdown near the button
        x = entry_widget.winfo_rootx()
        y = entry_widget.winfo_rooty() + entry_widget.winfo_height()
        dropdown.geometry(f"+{x}+{y}")

        # Create listbox with values
        listbox = tk.Listbox(dropdown, width=200, height=200)
        listbox.pack(fill="both", expand=True)

        # Add scrollbar
        scrollbar = ttk.Scrollbar(listbox)
        scrollbar.pack(side="right", fill="y")
        listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=listbox.yview)

        # Populate listbox
        for value in self.unique_values[header]:
            listbox.insert(tk.END, value)

        def select_value(event=None):
            selection = listbox.curselection()
            if selection:
                entry_widget.delete(0, tk.END)
                entry_widget.insert(0, listbox.get(selection))
                dropdown.destroy()

        listbox.bind('<Double-Button-1>', select_value)
        listbox.bind('<Return>', select_value)

    # Rest of the methods remain the same
    def move_up(self):
        idx = self.selected_listbox.curselection()
        if idx and idx[0] > 0:
            text = self.selected_listbox.get(idx)
            self.selected_listbox.delete(idx)
            self.selected_listbox.insert(idx[0] - 1, text)
            self.selected_listbox.selection_set(idx[0] - 1)

    def move_down(self):
        idx = self.selected_listbox.curselection()
        if idx and idx[0] < self.selected_listbox.size() - 1:
            text = self.selected_listbox.get(idx)
            self.selected_listbox.delete(idx)
            self.selected_listbox.insert(idx[0] + 1, text)
            self.selected_listbox.selection_set(idx[0] + 1)

    def add_column(self):
        selection = self.available_listbox.curselection()
        if selection:
            self.selected_listbox.insert(tk.END, self.available_listbox.get(selection))
            self.available_listbox.delete(selection)

    def remove_column(self):
        selection = self.selected_listbox.curselection()
        if selection:
            self.available_listbox.insert(tk.END, self.selected_listbox.get(selection))
            self.selected_listbox.delete(selection)

    def pull_data(self):
        if not self.excel_path:
            messagebox.showwarning("Warning", "Please select an Excel file first.")
            return

        try:
            criteria = {
                header: entry.get().strip()
                for header, entry in self.entry_widgets.items()
                if entry.get().strip()
            }

            output_columns = list(self.selected_listbox.get(0, tk.END))
            if not output_columns:
                messagebox.showwarning("Warning", "Please select at least one output column.")
                return

            separator = self.separator_entry.get()

            wb = load_workbook(self.excel_path, read_only=True)
            sheet = wb.active
            rows = list(sheet.rows)
            headers = [cell.value for cell in rows[0]]
            header_indices = {header: idx for idx, header in enumerate(headers)}

            self.results_text.delete(1.0, tk.END)

            for row in rows[1:]:
                matches_criteria = True
                for header, search_value in criteria.items():
                    cell_value = str(row[header_indices[header]].value or "").strip()
                    if search_value.lower() not in cell_value.lower():
                        matches_criteria = False
                        break

                if matches_criteria:
                    output_values = [str(row[header_indices[col]].value or "") for col in output_columns]
                    self.results_text.insert(tk.END, f"{separator.join(output_values)}\n")

            wb.close()

            if self.results_text.get(1.0, tk.END).strip() == "":
                self.results_text.insert(tk.END, "No matching results found.")

        except Exception as e:
            messagebox.showerror("Error", f"Error processing Excel file: {str(e)}")


def main():
    root = tk.Tk()
    app = ExcelDataPullApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()