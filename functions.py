
import fitz
from PyPDF2 import PdfMerger
import copy
import re
import hashlib
import time
import pickle
import openpyxl
import os
import torch
from new_fns import *
from utilities.easyocr_mosaic import HardOCR
import tkinter as tk
from tkinter import ttk
import cv2
from PIL import Image, ImageTk
from utilities.extend_line_vertices import *

# Standalone Functions

# region Utilities

def generate_color(string):
    # Generate a hash value for the input string using MD5
    hash_value = hashlib.md5(string.encode('utf-8')).hexdigest()

    # Take the first six characters of the hash value
    hex_code = hash_value[:6]

    # Convert the hex code to RGB values
    r, g, b = int(hex_code[:2], 16), int(hex_code[2:4], 16), int(hex_code[4:], 16)

    # Return the RGB values as a tuple
    return (r, g, b)

def calculate_center(box):
    """Calculate the center of a box."""
    x_min, y_min, x_max, y_max = box.tolist()
    center_x = (x_min + x_max) / 2
    center_y = (y_min + y_max) / 2
    return np.array([center_x, center_y])

def calculate_distance(center1, center2):
    """Calculate the Euclidean distance between two centers."""
    return np.linalg.norm(center1 - center2)

def point_in_box(x, y, box):
    xmin, ymin, xmax, ymax = box
    if xmin <= x <= xmax and ymin <= y <= ymax:
        return True
    return False

def overlap_bbox(obox, dbox, extension, include_inside):
    #this fn works as follows
    #if include inside is true, we ONLY check for complete enclosed items
    #if it's not we will ignore completely enclosed items
    #so if its not completely enlcosed but there is overlap we return true else we return false


    # box 1 is a text box, box2 is an object box
    # Extract coordinates of box1
    x1_min, y1_min, x1_max, y1_max = obox

    # Extract coordinates of box2
    x2_min, y2_min, x2_max, y2_max = dbox

    # Extend box2 by the given extension
    x2_min_ext = max(x2_min - extension, 0)
    y2_min_ext = max(y2_min - extension, 0)
    x2_max_ext = x2_max + extension
    y2_max_ext = y2_max + extension
    if include_inside:
        # Check for overlap. includes completely enclosed boxes
        if (x1_min <= x2_max_ext and x1_max >= x2_min_ext and y1_min <= y2_max_ext and y1_max >= y2_min_ext):
            return True
        else:
            return False
    else:
        # Check if box1 is completely inside un-extended box2
        if x2_min <= x1_min and x2_max >= x1_max and y2_min <= y1_min and y2_max >= y1_max:
            return False
        else:
            # Check for overlap.
            if (x1_min <= x2_max_ext and x1_max >= x2_min_ext and y1_min <= y2_max_ext and y1_max >= y2_min_ext):
                return True
            else:
                return False

def save_easyocr_results_pickle(results, filename='easyocr_results.pkl'):
    with open(filename, 'wb') as f:
        pickle.dump(results, f)

def load_easyocr_results_pickle(filename='easyocr_results.pkl'):
    try:
        with open(filename, 'rb') as f:
            return pickle.load(f)
    except FileNotFoundError:
        print(f"{filename} not found.")
        return []

def pil_to_cv2(pil_image):
    """
    Convert a PIL image to a cv2 image.
    :param pil_image: PIL image object
    :return: cv2 image object
    """
    # Convert PIL image to numpy array
    numpy_image = np.array(pil_image)

    # If the image has a transparency channel (RGBA)
    if numpy_image.shape[-1] == 4:
        # Convert RGBA to RGB by dropping the transparency channel
        numpy_image = numpy_image[:, :, :3]

    # Convert numpy array to cv2 image
    cv2_image = cv2.cvtColor(numpy_image, cv2.COLOR_RGB2BGR)

    return cv2_image


def show_image_in_tkinter(filled_img, title="Image View"):
    # Create a new Toplevel window
    debug_window = tk.Toplevel()
    debug_window.title(title)

    # Set maximum window dimensions
    max_width = 1200
    max_height = 900

    # Convert BGR to RGB
    rgb_img = cv2.cvtColor(filled_img, cv2.COLOR_BGR2RGB)

    # Convert to PIL Image
    pil_img = Image.fromarray(rgb_img)

    # Get image dimensions
    img_width, img_height = pil_img.size

    # Convert to PhotoImage
    photo = ImageTk.PhotoImage(image=pil_img)

    # Create a frame to hold the canvas and scrollbars
    frame = tk.Frame(debug_window)
    frame.pack(expand=True, fill='both')

    # Create canvas and scrollbars
    canvas = tk.Canvas(frame, width=min(img_width, max_width), height=min(img_height, max_height))
    h_scrollbar = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=canvas.xview)
    v_scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=canvas.yview)

    # Configure scrollbars
    canvas.configure(xscrollcommand=h_scrollbar.set, yscrollcommand=v_scrollbar.set)

    # Pack layout (changed from grid to match HoughLinesApp)
    h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
    v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    # Create the image directly on the canvas (no Label needed)
    canvas.create_image(0, 0, image=photo, anchor=tk.NW)

    # Configure canvas scrolling region
    canvas.configure(scrollregion=(0, 0, img_width, img_height))

    # Store photo reference to prevent garbage collection
    canvas.photo = photo

    # Add mousewheel scrolling functions
    def _on_mousewheel_y(event):
        if event.num == 4 or event.delta > 0:
            canvas.yview_scroll(-1, "units")
        else:
            canvas.yview_scroll(1, "units")

    def _on_mousewheel_x(event):
        if event.num == 4 or event.delta > 0:
            canvas.xview_scroll(-1, "units")
        else:
            canvas.xview_scroll(1, "units")

    # Bind mousewheel events to canvas
    canvas.bind('<MouseWheel>', _on_mousewheel_y)  # Windows
    canvas.bind('<Button-4>', _on_mousewheel_y)  # Linux
    canvas.bind('<Button-5>', _on_mousewheel_y)  # Linux
    canvas.bind('<Shift-MouseWheel>', _on_mousewheel_x)  # Windows
    canvas.bind('<Shift-Button-4>', _on_mousewheel_x)  # Linux
    canvas.bind('<Shift-Button-5>', _on_mousewheel_x)  # Linux

    # Make the window stay on top (optional)
    debug_window.lift()

    return debug_window  # Return window reference to prevent garbage collection

# endregion


# region File Operations

def compile_excels(input_folder, output_folder=None, prefix='page', timestamp=True, recursive=False):
    """
    Compile multiple Excel files into a single Excel file.

    Args:
    input_folder (str): Path to the folder containing input Excel files.
    output_folder (str, optional): Path to save the output Excel file. If None, uses input_folder.
    prefix (str, optional): Prefix of Excel files to compile. Defaults to 'page'.
    timestamp (bool, optional): Whether to include a timestamp in the output filename. Defaults to True.
    recursive (bool, optional): Whether to search for Excel files recursively in the input folder. Defaults to False.

    Returns:
    str: Path to the compiled Excel file.
    """
    if output_folder is None:
        output_folder = input_folder

    if timestamp:
        timestamp_str = f"_{int(time.time())}"
    else:
        timestamp_str = ""

    output_filename = os.path.join(output_folder, f"compiled_output{timestamp_str}.xlsx")

    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active
    first_file = True

    if recursive:
        for root, dirs, files in os.walk(input_folder):
            for file_name in files:
                if file_name.endswith(".xlsx") and file_name.startswith(prefix):
                    file_path = os.path.join(root, file_name)
                    workbook = openpyxl.load_workbook(file_path)
                    sheet = workbook.active

                    if first_file:
                        for row in sheet.iter_rows(min_row=1, max_row=1):
                            for cell in row:
                                output_sheet[cell.coordinate].value = cell.value
                        first_file = False

                    for row in sheet.iter_rows(min_row=2):
                        output_sheet.append([cell.value for cell in row])
    else:
        for file_name in os.listdir(input_folder):
            if file_name.endswith(".xlsx") and file_name.startswith(prefix):
                file_path = os.path.join(input_folder, file_name)
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active

                if first_file:
                    for row in sheet.iter_rows(min_row=1, max_row=1):
                        for cell in row:
                            output_sheet[cell.coordinate].value = cell.value
                    first_file = False

                for row in sheet.iter_rows(min_row=2):
                    output_sheet.append([cell.value for cell in row])

    output_workbook.save(output_filename)
    print('excels compiled to: ', output_filename)
    return output_filename

def merge_pdf(directory_path):
    # Set the directory path
    #directory_path = r"C:\Users\dcaoili\OneDrive - Samuel Engineering\Documents\PIDS - WORKING\missouri cobalt 22024\New folder (2)\2024-04-17 Area 620"

    # Create a PDF merger object
    pdf_merger = PdfMerger()

    # Loop through all the files in the directory
    for filename in os.listdir(directory_path):
        if filename.endswith('.pdf'):
            # Open the file and append it to the PDF merger object
            print(f"opening {filename}")
            with open(os.path.join(directory_path, filename), 'rb') as pdf_file:
                pdf_merger.append(pdf_file)

    # Write the merged PDF to a file
    with open(os.path.join(directory_path, 'merged.pdf'), 'wb') as output_file:
        pdf_merger.write(output_file)


def create_progress_window(total_pages):
    progress_window = tk.Toplevel()
    progress_window.title("Converting PDF")
    progress_window.geometry("300x150")

    # Center the window
    window_width = 300
    window_height = 150
    screen_width = progress_window.winfo_screenwidth()
    screen_height = progress_window.winfo_screenheight()
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2
    progress_window.geometry(f"{window_width}x{window_height}+{x}+{y}")

    # Add labels and progress bar
    label = ttk.Label(progress_window, text="Converting PDF to images...")
    label.pack(pady=10)

    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(progress_window, variable=progress_var, maximum=total_pages)
    progress_bar.pack(pady=10, padx=20, fill=tk.X)

    page_label = ttk.Label(progress_window, text="Page: 0/0")
    page_label.pack(pady=10)

    return progress_window, progress_var, page_label


def pdf2png(pdf_file, target_width):
    doc = fitz.open(pdf_file)
    total_pages = len(doc)

    # Create progress window
    progress_window, progress_var, page_label = create_progress_window(total_pages)

    # Calculate DPI
    page = doc[0]
    page_width = page.rect.width
    dpi = int(target_width * 72 / page_width)

    # Create output folder
    folder_name = os.path.join(os.path.dirname(pdf_file),
                               os.path.splitext(os.path.basename(pdf_file))[0] + "_images")
    os.makedirs(folder_name, exist_ok=True)

    try:
        # Convert pages
        for page_num, page in enumerate(doc):
            # Update progress
            progress_var.set(page_num + 1)
            page_label.config(text=f"Page: {page_num + 1}/{total_pages}")
            progress_window.update()

            # Convert page
            pix = page.get_pixmap(matrix=fitz.Matrix(dpi / 72, dpi / 72))
            png_file = os.path.join(folder_name, f"page_{page_num + 1}.png")
            pix.save(png_file)

        progress_window.destroy()
        return folder_name

    except Exception as e:
        progress_window.destroy()
        raise e
    finally:
        doc.close()
# endregion


# region Text Processing

def merge_common_substrings(str1, str2):
    # Split the strings into lists of words
    words1 = str1.split()
    words2 = str2.split()

    merged_words = []
    i, j = 0, 0

    while i < len(words1) and j < len(words2):
        if words1[i] == words2[j]:
            merged_words.append(words1[i])
            i += 1
            j += 1
        else:
            merged_word = ""
            if i < len(words1):
                merged_word += words1[i]
            if j < len(words2):
                merged_word += "/" + words2[j]
            merged_words.append(merged_word)
            i += 1
            j += 1

    # Append any remaining words from either string
    merged_words.extend(words1[i:])
    merged_words.extend(["/" + word for word in words2[j:]])
    last_word = merged_words[-1:][0]
    print('last word', last_word)
    joined_words = " ".join(merged_words)
    if not any(char.isdigit() for char in last_word):
        joined_words = joined_words+"S"
    return joined_words

def condense_hyphen_string(s):
    if not s:
        return ""
    words = s.split()
    result = []

    for word in words:
        if '/' in word:
            parts = word.split('/')
            prefixes = [part.split('-')[:-1] for part in parts]
            suffixes = [part.split('-')[-1] for part in parts]

            if all(prefix == prefixes[0] for prefix in prefixes):
                condensed = '-'.join(prefixes[0] + ['/'.join(suffixes)])
                result.append(condensed)
            else:
                result.append(word)
        else:
            result.append(word)

    return ' '.join(result)
# endregion


# region Image Processing


def draw_detection_boxes(img, labels, boxes, scores, size=5, minscore=.5):
    img = copy.copy(img)
    # plot_pic(img,labels,boxes,scores)
    # Define some colors for the boxes and labels

    for label, box, score in zip(labels, boxes, scores):
        if score > minscore:
            # Extract the coordinates of the box
            my_list = box
            my_list = [int(x) for x in my_list]
            x1, y1, x2, y2 = my_list

            # Draw a rectangle around the box
            cv2.rectangle(img, (x1, y1), (x2, y2), (255, 0, 0), thickness=2)
            # Add a label above the box
            cv2.putText(img, label + ":0." + str(int(float(score * 1000))), (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, .7,
                        (255, 0, 0), thickness=2)

    # Display the image with the boxes and labels overlayed using Matplotlib
    # fig, ax = plt.subplots(figsize=(size, size))
    # ax.imshow(img)
    # plt.show()
    return img

def remove_objects_from_image(img, inst_data, ocr_results):
    # Create a copy of the original image
    masked_img = img.copy()

    # Process object results
    for data in inst_data:
        # Convert box coordinates to integers
        x1, y1, x2, y2 = map(int, data['box'])

        # Draw a white filled rectangle to cover the object
        cv2.rectangle(masked_img, (x1, y1), (x2, y2), (255, 255, 255), -1)

    # Process OCR results
    for result in ocr_results:
        box = result[0]
        x1, y1, x2, y2 = int(box[0][0]), int(box[0][1]), int(box[2][0]), int(box[2][1])

        # Draw a white filled rectangle to cover the OCR text
        cv2.rectangle(masked_img, (x1, y1), (x2, y2), (255, 255, 255), -1)

    return masked_img

def plot_ocr_results(img, results):
    img_ocr_results = img.copy()
    for result in results:
        # Get the recognized text
        text = result[1]
        box = result[0]
        x1, y1, x2, y2 = int(box[0][0]), int(box[0][1]), int(box[2][0]), int(box[2][1])

        pt1 = (x1, y1)
        pt2 = (x2, y2)
        # draw boxes for viewing results
        cv2.rectangle(img_ocr_results, pt1, pt2, (0, 0, 255), 2)
        # Add the recognized text on top of the box
        cv2.putText(img_ocr_results, text, (x1, y1 - 10), cv2.FONT_HERSHEY_TRIPLEX, .7, (0, 0, 255), 2)

    # plot_pic2(img_ocr_results,20)
    return img_ocr_results



# endregion


# region OCR and Text Analysis

def get_text_in_ocr_results(re_pattern, results):
    pattern = re.compile(re_pattern)
    text_list = []
    for result in results:
        # we need a list because we might want to overwrite
        text = result[1]
        if pattern.match(text):
            text_list.append(result)
    # returning a list of results that match the text
    return text_list

def get_comment(ocr_results, dbox, box_expand, include_inside=False):
    # dbox is a detecto box
    # obox is an easyOCR box
    comment = ''
    object_box = dbox.tolist()
    object_box = [int(x) for x in object_box]


    for item in ocr_results:
        #we cant do for obox, word, score in ocr because ocr_results changes lenght i.e. color or a word is appened
        obox, word, confidence = item


        text_box = [obox[0][0], obox[0][1], obox[2][0], obox[2][1]]

        if overlap_bbox(text_box, object_box, box_expand, include_inside):
            comment += f" {word}"
    return comment

# endregion


# region Line Processing


def process_line_data(img, ocr_results, re_line,
                      hough_params=None,
                      canny_params=None,
                      extension_params=None,
                      paint_line_thickness=5,
                      line_join_threshold=20,
                      line_box_scale=1.2,
                      erosion_kernel=10,
                      erosion_iterations=2,
                      binary_threshold=200,
                      line_img_scale=1.0,
                      debug_line=True,
                      remove_significant_lines_only=True,
                      detecto_boxes=None,
                      clean_img=True,
                      simple=True,
                      remove_text_before=False,
                      text_min_score=0.5,
                      white_out_color=(255,255,255)
                      ):
    """Process line-related data from an image and OCR results.

    Returns:
        dict: Keys are line text, values are dict containing 'image' and 'color'
    """
    if not ocr_results:
        return

    if simple:
        ocr_line_txt = get_text_in_ocr_results(re_line, ocr_results)
        return ocr_line_txt[0][1] if ocr_line_txt else ''

    line_data = {}
    if hough_params and canny_params and re_line:

        if clean_img:
            img_cleaned = remove_objects_and_text_from_img(img, ocr_results, detecto_boxes,
                                                           color=white_out_color, text_min_score=text_min_score)
        else:
            img_cleaned = img

        # Original image processing code remains the same until the flood fill part
        gray = cv2.cvtColor(img_cleaned, cv2.COLOR_BGR2GRAY)
        edges = cv2.Canny(gray, canny_params['low_threshold'], canny_params['high_threshold'],
                          apertureSize=canny_params['aperture_size'])
        hough_lines = cv2.HoughLinesP(edges, hough_params['rho'], hough_params['theta'],
                                hough_params['threshold'],
                                minLineLength=hough_params['min_line_length'],
                                maxLineGap=hough_params['max_line_gap'])

        if extension_params:
            mt = extension_params['merge_threshold']
            la = extension_params['look_ahead']
            mn = extension_params['max_neighbors']
            if mt != 0 and la != 0:
                hough_lines = extend_line_vertices_optimized(hough_lines, mt, la, mn)


        if debug_line:
            show_houghlines(img_cleaned, hough_lines)

        significant_line_groups = {}
        all_line_groups = group_lines(hough_lines, max_distance=line_join_threshold) # joins lines by endpoint to create groups
        ocr_line_txt = get_text_in_ocr_results(re_line, ocr_results)
        ocr_line_txt = scale_ocr_box_additive(ocr_line_txt, line_box_scale)

        sig_line_group_list = []
        for line_group in all_line_groups:
            for box, text, confidence in ocr_line_txt:
                if box_intersects_line(box, line_group):
                    significant_line_groups.setdefault(text, []).append(line_group)
                    sig_line_group_list.append(line_group)

        line_boxes = {}
        # Add all OCR boxes to line_boxes and ensure each text has an entry in significant_line_groups
        for box, text, confidence in ocr_line_txt:
            line_boxes.setdefault(text, []).append(box)
            if text not in significant_line_groups:
                significant_line_groups[text] = []

        if remove_text_before:
            img = remove_text_from_img(img, ocr_results, color=white_out_color, score_thresh=text_min_score)

        if remove_significant_lines_only:
            img_line_groups_removed = white_out_line_groups(img, sig_line_group_list, paint_line_thickness)
        else:
            img_line_groups_removed = white_out_line_groups(img, all_line_groups, paint_line_thickness)

        img_lines_removed_gray = cv2.cvtColor(img_line_groups_removed, cv2.COLOR_BGR2GRAY)
        _, thresholded_image = cv2.threshold(img_lines_removed_gray, binary_threshold, 255, cv2.THRESH_BINARY)
        thresholded_3channel = cv2.cvtColor(thresholded_image, cv2.COLOR_GRAY2BGR)




        kernel = np.ones((erosion_kernel, erosion_kernel), np.uint8)
        img_line_groups_removed_eroded = cv2.erode(thresholded_3channel, kernel, iterations=erosion_iterations)

        # Scale the image before flood fill if line_img_scale is not 1.0
        if line_img_scale != 1.0:
            height, width = img_line_groups_removed_eroded.shape[:2]
            new_height = int(height * line_img_scale)
            new_width = int(width * line_img_scale)
            img_line_groups_removed_eroded = cv2.resize(
                img_line_groups_removed_eroded,
                (new_width, new_height),
                interpolation=cv2.INTER_NEAREST
            )


        for line_text in significant_line_groups:
            color = generate_color(line_text)

            # Scale line groups and boxes if necessary
            if line_img_scale != 1.0:
                scaled_line_groups = [
                    scale_line_group(line_group, line_img_scale)
                    for line_group in significant_line_groups[line_text]
                ]
                scaled_boxes = [
                    scale_box(box, line_img_scale)
                    for box in line_boxes[line_text]
                ]
            else:
                scaled_line_groups = significant_line_groups[line_text]
                scaled_boxes = line_boxes[line_text]

            # Process the image with scaled components
            processed_img = place_lines_on_image_and_flood_fill(
                img_line_groups_removed_eroded,
                scaled_line_groups,
                color,
                scaled_boxes,
                line_thickness=max(1, int(paint_line_thickness * line_img_scale)),
                debug=debug_line
            )

            line_data[line_text] = {
                'image': processed_img,
                'color': color
            }

    return line_data

def process_line_data_v2(img, ocr_results, re_line,
                      hough_params=None,
                      canny_params=None,
                      extension_params=None,
                      paint_line_thickness=5,
                      line_join_threshold=20,
                      line_box_scale=1.2,
                      erosion_kernel=10,
                      erosion_iterations=2,
                      binary_threshold=200,
                      line_img_scale=1.0,
                      debug_line=True,
                      remove_significant_lines_only=True,
                      detecto_boxes=None,
                      clean_img=True,
                      simple=True,
                      remove_text_before=False,
                      text_min_score=0.5,
                      white_out_color=(255,255,255)
                      ):
    """Process line-related data from an image and OCR results.

    Returns:
        dict: Keys are line text, values are dict containing 'image' and 'color'
    """
    if not ocr_results:
        return

    if simple:
        ocr_line_txt = get_text_in_ocr_results(re_line, ocr_results)
        return ocr_line_txt[0][1] if ocr_line_txt else ''

    line_data = {}
    if not (hough_params and canny_params and re_line):
        return line_data

    if clean_img:
        img_cleaned = remove_objects_and_text_from_img(img, ocr_results, detecto_boxes,
                                                       color=white_out_color, text_min_score=text_min_score)
    else:
        img_cleaned = img

    # Original image processing code remains the same until the flood fill part
    gray = cv2.cvtColor(img_cleaned, cv2.COLOR_BGR2GRAY)
    edges = cv2.Canny(gray, canny_params['low_threshold'], canny_params['high_threshold'],
                      apertureSize=canny_params['aperture_size'])
    hough_lines = cv2.HoughLinesP(edges, hough_params['rho'], hough_params['theta'],
                            hough_params['threshold'],
                            minLineLength=hough_params['min_line_length'],
                            maxLineGap=hough_params['max_line_gap'])

    if debug_line:
        show_houghlines(img_cleaned, hough_lines)

    # group lines whose endpoints are close
    all_line_groups = group_lines(hough_lines, max_distance=line_join_threshold) # joins lines by endpoint to create groups

    # get and scale up ocr line id text results
    ocr_line_txt = get_text_in_ocr_results(re_line, ocr_results)
    ocr_line_txt = scale_ocr_box_additive(ocr_line_txt, line_box_scale)

    # remove all the lines and process the image so we can do a good floodfill
    if remove_text_before:
        img = remove_text_from_img(img, ocr_results, color=white_out_color, score_thresh=text_min_score)
    img_line_groups_removed = white_out_line_groups(img, all_line_groups, paint_line_thickness)
    img_lines_removed_gray = cv2.cvtColor(img_line_groups_removed, cv2.COLOR_BGR2GRAY)
    _, thresholded_image = cv2.threshold(img_lines_removed_gray, binary_threshold, 255, cv2.THRESH_BINARY)
    thresholded_3channel = cv2.cvtColor(thresholded_image, cv2.COLOR_GRAY2BGR)
    kernel = np.ones((erosion_kernel, erosion_kernel), np.uint8)
    img_line_groups_removed_eroded = cv2.erode(thresholded_3channel, kernel, iterations=erosion_iterations)
    # Scale the image before flood fill if line_img_scale is not 1.0
    if line_img_scale != 1.0:
        height, width = img_line_groups_removed_eroded.shape[:2]
        new_height = int(height * line_img_scale)
        new_width = int(width * line_img_scale)
        img_line_groups_removed_eroded = cv2.resize(
            img_line_groups_removed_eroded,
            (new_width, new_height),
            interpolation=cv2.INTER_NEAREST
        )

    # make colors
    line_color_data = {}
    for box, line_text, confidence in ocr_line_txt:
        color = generate_color(line_text)
        line_color_data[line_text]=color

    pics = []
    for box, text, confidence in ocr_line_txt:

        scaled_box = scale_box(box, line_img_scale) if line_img_scale != 1.0 else box

        for line_group in all_line_groups:

            scaled_line_group = scale_line_group(line_group, line_img_scale) if line_img_scale != 1.0 else line_group

            # grab the color
            color = line_color_data[text]

            # Process the image with scaled components
            processed_img = place_lines_on_image_and_flood_fill(
                img_line_groups_removed_eroded,
                scaled_line_group,
                color,
                scaled_box,
                line_thickness=max(1, int(paint_line_thickness * line_img_scale)),
                debug=debug_line
            )

            pics.append(processed_img)

    return line_color_data, pics


def get_lines_from_box(box, line_data, img_scale=1.0):
    lines = []
    for line_text, data in line_data.items():
        mode_color = region_mode_color(box, data['image'], img_scale)
        expected_color = data['color']
        if mode_color == expected_color:
            lines.append(line_text)

    return sorted(lines)

def place_lines_on_image_and_flood_fill(img_line_groups_removed, line_group, line_color, line_box, line_thickness=5,
                                        debug=True,
                                        ):
    img = img_line_groups_removed.copy()

    # Draw lines, handling nested structure
    for line_set in line_group:
        for line in line_set:
            x1, y1, x2, y2 = line  # Each line_set is [[x1,y1,x2,y2], ...]
            cv2.line(img, (int(x1), int(y1)), (int(x2), int(y2)), (0, 0, 0), line_thickness)


    seed_pixel_boxes = [(box, line_color) for box in line_box]
    filled_img = multi_region_flood_fill(seed_pixel_boxes, img, 1.0)

    if debug:
        print('line group ', line_group)
        # Show and wait
        show_image_in_tkinter(filled_img, "Filled Image")

    return filled_img


def multi_region_flood_fill(seed_pixel_boxes, line_img, expand_scale):
    # Create a new image to perform the flood fill on
    img = line_img.copy()

    # Get the size of the image
    height, width, channels = img.shape

    # Initialize the pixel stacks with the outer boundary pixels in each box
    pixel_stacks = []
    for i, seed_box in enumerate(seed_pixel_boxes):
        box, color = seed_box

        x1, y1, x2, y2 = box[0][0], box[0][1], box[2][0], box[2][1]
        expand_h = int(abs(x1-x2)*(expand_scale-1)/2)
        expand_v = int(abs(y1-y2)*(expand_scale-1)/2)
        x1, y1, x2, y2 = x1-expand_h, y1-expand_v, x2+expand_h, y2+expand_v
        # Create a new list for each pixel stack
        pixel_stack = []

        # Draw and fill the rectangle using cv2.rectangle
        cv2.rectangle(img, (x1, y1), (x2, y2), color, thickness=cv2.FILLED)

        # Iterate over the top and bottom rows of the box and append their pixels
        for x in range(x1, x2 + 1):
            pixel_stack.append((x, y1, color))
            pixel_stack.append((x, y2, color))

        # Iterate over the left and right columns of the box (excluding corners) and append their pixels
        for y in range(y1, y2 + 1):
            pixel_stack.append((x1, y, color))
            pixel_stack.append((x2, y, color))

        # Append the pixel stack for this box to the list of pixel stacks
        pixel_stacks.append(pixel_stack)

    # While there is still an element in any of the lists:
    # An empty list returns false
    while pixel_stacks:
        # Enumerate gives us an index n
        for pixel_stack in pixel_stacks:
            # print("chekcing stack "+str(n))
            start_size = len(pixel_stack)
            # i goes from the bottom to the top
            for i in range(start_size):
                pixel = pixel_stack.pop(0)
                x, y, color = pixel
                # Get neighbors
                neighbors = [(x + 1, y), (x - 1, y), (x, y + 1), (x, y - 1)]
                for neighbor in neighbors:
                    nx, ny = neighbor
                    # Check if neighbor is within image bounds
                    if (0 <= nx < width) and (0 <= ny < height):
                        # Get the color of the neighboring pixel
                        neighbor_color = tuple(img[ny, nx])
                        # If the neighboring pixel is black
                        if neighbor_color == (0, 0, 0):
                            # Add pixel to pixel list
                            # print("append")
                            pixel_stack.append((nx, ny, color))
                            # Set color on filled image
                            img[ny, nx] = color

        # Remove empty sublists
        pixel_stacks = [x for x in pixel_stacks if x]

    # Return the filled image
    return img

def region_mode_color(box, img, img_scale):
    x_min, y_min, x_max, y_max = box.tolist()

    # Scale the coordinates to match the scaled image
    scaled_coords = [
        int(x_min * img_scale),
        int(y_min * img_scale),
        int(x_max * img_scale),
        int(y_max * img_scale)
    ]

    # Make sure coordinates don't exceed image dimensions
    height, width = img.shape[:2]
    x_min = max(0, min(scaled_coords[0], width - 1))
    y_min = max(0, min(scaled_coords[1], height - 1))
    x_max = max(0, min(scaled_coords[2], width - 1))
    y_max = max(0, min(scaled_coords[3], height - 1))

    roi = img[y_min:y_max, x_min:x_max]

    # Create a mask for non-black and non-white pixels
    non_black_white_mask = np.all((roi != [0, 0, 0]) & (roi != [255, 255, 255]), axis=2)

    # Extract non-black, non-white pixel colors
    non_black_white_pixels = roi[non_black_white_mask]

    # Find the most popular color among non-black, non-white pixels
    unique_colors, counts = np.unique(non_black_white_pixels, axis=0, return_counts=True)
    try:
        most_popular_color = unique_colors[np.argmax(counts)]
    except:
        return (0, 0, 0)

    return tuple(most_popular_color)

def box_intersects_line(box, line_group, scale=1.0):
    """
    Checks if any line in the line group intersects with the given box.

    Args:
        box: List of corner points from OCR result [[x1,y1], [x2,y1], [x2,y2], [x1,y2]]
        line_group: List of lines where each line is [x1, y1, x2, y2]
        scale: Float value to scale the box
    """
    # Extract coordinates from OCR box format
    box_x1, box_y1 = box[0]
    box_x2, box_y2 = box[2]

    # Convert to numeric type if needed
    box_x1 = float(box_x1)
    box_x2 = float(box_x2)
    box_y1 = float(box_y1)
    box_y2 = float(box_y2)

    # Ensure box coordinates are in correct order
    box_x1, box_x2 = min(box_x1, box_x2), max(box_x1, box_x2)
    box_y1, box_y2 = min(box_y1, box_y2), max(box_y1, box_y2)

    # Calculate box center
    center_x = (box_x1 + box_x2) / 2
    center_y = (box_y1 + box_y2) / 2

    # Calculate box dimensions and scaling
    width = box_x2 - box_x1
    height = box_y2 - box_y1
    scaled_width = width * scale
    scaled_height = height * scale

    # Calculate new box coordinates maintaining the center point
    box_x1 = center_x - (scaled_width / 2)
    box_x2 = center_x + (scaled_width / 2)
    box_y1 = center_y - (scaled_height / 2)
    box_y2 = center_y + (scaled_height / 2)

    def ccw(A, B, C):
        return (C[1] - A[1]) * (B[0] - A[0]) > (B[1] - A[1]) * (C[0] - A[0])

    def line_segments_intersect(line1_start, line1_end, line2_start, line2_end):
        return ccw(line1_start, line2_start, line2_end) != ccw(line1_end, line2_start, line2_end) and \
            ccw(line1_start, line1_end, line2_start) != ccw(line1_start, line1_end, line2_end)

    def point_in_box(point):
        x, y = point
        return box_x1 <= x <= box_x2 and box_y1 <= y <= box_y2

    for line in line_group:
        x1, y1, x2, y2 = line
        line_start = (x1, y1)
        line_end = (x2, y2)

        if point_in_box(line_start) or point_in_box(line_end):
            return True

        box_edges = [
            ((box_x1, box_y1), (box_x2, box_y1)),
            ((box_x2, box_y1), (box_x2, box_y2)),
            ((box_x2, box_y2), (box_x1, box_y2)),
            ((box_x1, box_y2), (box_x1, box_y1))
        ]

        for edge_start, edge_end in box_edges:
            if line_segments_intersect(line_start, line_end, edge_start, edge_end):
                return True

    return False

def remove_objects_and_text_from_img(img, ocr_results, detecto_boxes, color=(255, 255, 255), text_min_score=0.5):
    img = remove_text_from_img(img, ocr_results, color, text_min_score)
    img = remove_objects_from_img(img, detecto_boxes, color)
    return img

def remove_text_from_img(img, ocr_results, color=(255, 255, 255), score_thresh=0.5):
    img = img.copy()
    # img is a cv2 img
    for result in ocr_results:
        box = result[0]
        score = result[2]
        if score < score_thresh:
            continue

        # Extract coordinates from OCR box format
        box_x1, box_y1 = box[0]
        box_x2, box_y2 = box[2]
        # White out the text region
        cv2.rectangle(img, (int(box_x1), int(box_y1)), (int(box_x2), int(box_y2)), color, -1)

    return img

def remove_objects_from_img(img, detecto_boxes, color=(255, 255, 255)):
    img = img.copy()
    # img is a cv2 img
    for box in detecto_boxes:
        box_x1, box_y1, box_x2, box_y2 = map(int, box)
        # White out the object region
        cv2.rectangle(img, (box_x1, box_y1), (box_x2, box_y2), color, -1)

    return img

def show_houghlines(img, lines):

    debug_img = img.copy()

    # Draw detected lines in green
    for line in lines:
        x1, y1, x2, y2 = line[0]
        cv2.line(debug_img, (x1, y1), (x2, y2), (0, 255, 0), 1)

        # Draw blue X markers at endpoints
        marker_size = 4
        # First endpoint
        cv2.line(debug_img,
                 (x1 - marker_size, y1 - marker_size),
                 (x1 + marker_size, y1 + marker_size),
                 (255, 0, 0), 1)
        cv2.line(debug_img,
                 (x1 - marker_size, y1 + marker_size),
                 (x1 + marker_size, y1 - marker_size),
                 (255, 0, 0), 1)

        # Second endpoint
        cv2.line(debug_img,
                 (x2 - marker_size, y2 - marker_size),
                 (x2 + marker_size, y2 + marker_size),
                 (255, 0, 0), 1)
        cv2.line(debug_img,
                 (x2 - marker_size, y2 + marker_size),
                 (x2 + marker_size, y2 - marker_size),
                 (255, 0, 0), 1)

    show_image_in_tkinter(debug_img, "Detected Lines Overlay")

# endregion


# region Instrument Processing

def return_inst_data(prediction_data, img,
                     reader, instrument_reader_settings, reader_settings,
                     expand=1.0,
                     radius=180,
                     inst_labels=None,
                     other_labels=None,
                     min_scores=None,
                     offset=None,
                     comment_box_expand=30,
                     tag_label_groups=None,
                     capture_ocr=True,
                     reader_sub_img_size=1300,
                     reader_stride=1250,
                     filter_ocr_threshold=0.9,
                     line_params=None
                     ):
    all_data = []
    group_inst = []
    group_other = []
    got_one = False
    detecto_boxes = []
    pred_boxes = []

    for label, box, score, visual_elements in prediction_data:
        try:
            if score < min_scores[label]:
                continue
        except Exception as e:
            print(e, '. Maybe try setting minscores --> settings >> set object minscores')

        if label in inst_labels:
            group_inst.append((label, box, score, visual_elements))
            detecto_boxes.append(box)
        if label in other_labels:
            group_other.append((label, box, score, visual_elements))

        if offset is not None:
            offset_tensor = torch.tensor([offset[0], offset[1], offset[0], offset[1]])
            offset_box = box + offset_tensor
            pred_boxes.append(offset_box)
        else:
            pred_boxes.append(box)

    if not group_inst:
        return

    local_ocr_results = None
    if capture_ocr:
        print('getting local ocr')
        local_ocr_results = HardOCR(img, reader, reader_settings, sub_img_size=reader_sub_img_size,
                                    stride=reader_stride)
        print('unfiltered local ocr results:\n', local_ocr_results)
        local_ocr_results = filter_ocr_results(local_ocr_results, pred_boxes, overlap_threshold=filter_ocr_threshold)
        print('filtered local ocr results:\n', local_ocr_results)

    line_data = process_line_data(
        img,
        local_ocr_results,
        **line_params.__dict__,
        detecto_boxes=detecto_boxes,
    )# line_color_data, pics

    for label, box, score, visual_elements in group_inst:
        if line_params.simple: # if simple mode
            lines = line_data
        elif line_data: # complex mode
            lines = get_lines_from_box(box, line_data, img_scale=line_params.line_img_scale)
            if lines and len(lines) == 1:
                lines = lines[0]
        else:
            lines = ''

        tag, tag_no = '', ''
        x_min, y_min, x_max, y_max = map(int, box.tolist())
        inst_center = calculate_center(box)

        x_expand = int((x_max - x_min) * (expand - 1) / 2)
        y_expand = int((y_max - y_min) * (expand - 1) / 2)

        crop_img = img[(y_min - y_expand):(y_max + y_expand), (x_min - x_expand):(x_max + x_expand)]
        try:
            results = reader.readtext(crop_img, **instrument_reader_settings)
            if results:
                if not got_one:
                    filename = 'temp/instrument_capture.png'
                    cv2.imwrite(filename, crop_img)
                    got_one = True

                tag = results[0][1]
                tag_no = ' '.join([box[1] for box in results[1:]])
        except Exception as e:
            print('error in instrument OCR:', e)
            continue

        valid_types = ['']
        if tag_label_groups and tag:
            valid_types = get_valid_types_for_tag(tag, tag_label_groups)

        inst_type = find_closest_other(inst_center, group_other, label, radius, valid_types, tag_label_groups)

        comment = ''
        if offset is not None:
            offset_tensor = torch.tensor([offset[0], offset[1], offset[0], offset[1]])
            offset_box = box + offset_tensor
        else:
            offset_box = box

        if local_ocr_results:
            print('getting local ocr comment')
            comment = get_comment(local_ocr_results, box, comment_box_expand)

        data = {
            'tag': tag,
            'tag_no': "'" + tag_no,
            'score': score,
            'box': offset_box,
            'label': label,
            'type': inst_type,
            'comment': comment,
            'line': lines,
            'visual_elements': visual_elements
        }
        all_data.append(data)

    return all_data

def get_valid_types_for_tag(tag, tag_label_groups):
    # For each key in tag_label_groups, split it into individual tags
    # and check if our tag starts with any of them
    for tag_group, valid_types in tag_label_groups.items():
        tag_prefixes = tag_group.split()
        if tag in tag_prefixes:
            return valid_types
    return None  # Return None if no matching group found

def find_closest_other(inst_center, group_other, current_label, radius, valid_types=[''], tag_label_groups=None):
    """
    Find the closest item in group_other that doesn't have the same label and is in valid_types if specified.

    Args:
        inst_center: Center coordinates of the current instrument
        group_other: List of (label, box, score) tuples for other objects
        current_label: Label of the current instrument to exclude
        radius: Maximum distance to consider
        valid_types: Optional list of valid type labels to consider
    """
    min_distance = radius
    closest_label = ''


    for label, box, score, _ in group_other:

        # Skip items with the same label as the current instrument
        if label == current_label:
            continue

        if tag_label_groups:
            # Skip if we have valid_types and this label isn't in it
            if valid_types is None:
                continue
            print('label ', label)
            print('valid_types ', valid_types)
            if label not in valid_types:
                continue

        other_center = calculate_center(box)
        distance = calculate_distance(inst_center, other_center)
        if distance < min_distance:
            min_distance = distance
            closest_label = label

    return closest_label

def filter_ocr_results(ocr_results, pred_boxes, overlap_threshold=.9):
    filtered_results = []

    for ocr_result in ocr_results:
        ocr_box = ocr_result[0]
        is_contained = False

        for pred_box in pred_boxes:
            overlap = calculate_overlap(ocr_box, pred_box)
            if overlap >= overlap_threshold:
                is_contained = True
                break

        if not is_contained:
            filtered_results.append(ocr_result)

    return filtered_results

def calculate_overlap(ocr_box, pred_box):
        # Convert OCR box format to x1,y1,x2,y2
        ocr_x1, ocr_y1 = ocr_box[0]
        ocr_x2, ocr_y2 = ocr_box[2]

        # Get prediction box coordinates
        pred_x1, pred_y1, pred_x2, pred_y2 = pred_box.tolist()

        # Calculate intersection coordinates
        x_left = max(ocr_x1, pred_x1)
        y_top = max(ocr_y1, pred_y1)
        x_right = min(ocr_x2, pred_x2)
        y_bottom = min(ocr_y2, pred_y2)

        # Check if there is an intersection
        if x_right < x_left or y_bottom < y_top:
            return 0.0

        # Calculate areas
        intersection_area = (x_right - x_left) * (y_bottom - y_top)
        ocr_area = (ocr_x2 - ocr_x1) * (ocr_y2 - ocr_y1)

        # Calculate overlap ratio relative to OCR box size
        overlap_ratio = intersection_area / ocr_area

        return overlap_ratio

# endregion

