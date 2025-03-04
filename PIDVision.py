import os
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
import torch
print(f"PyTorch version: {torch.__version__}")
print(f"CUDA available: {torch.cuda.is_available()}")
if torch.cuda.is_available():
    print(f"CUDA device: {torch.cuda.get_device_name(0)}")
from functions import *

from console_redirect import *
from detecto.core import Model
from detecto_gui.detecto_gui import ObjectDetectionApp
from utilities.find_an_instrument import FindAnInstrumentApp
from utilities.group_manager import GroupManager
from utilities.image_editor import ImageEditor
from utilities.minscore_edit import SliderApp
from model_predict_mosaic import *
from utilities.multi_window_dictionary_maker import DictionaryBuilder
from utilities.ocr_results_viewer import OCRViewer
from utilities.reader_settings import SetReaderSettings
from utilities.houghlines import HoughLinesApp
from utilities.xlsx_data_puller import ExcelDataPullApp
from tkinter import filedialog
from auto_scroll import AutoScrollbar
import easyocr
import json
import openpyxl
import os
import re
import threading
import tkinter as tk
import tkinter.messagebox
import tkinter.simpledialog
import xlwings as xw
from tkinter import ttk, messagebox
from licensing_system import LicenseManager
from utilities.line_processing_params import LineProcessingParams
from progress_window import ProgressWindow
from optimized_image_canvas import OptimizedImageCanvas
from splash_screen import SplashScreen
import pickle
import datetime
from text_corrections import TextCorrections, TextCorrectionsEditor
from config import Config

class CaptureBox:
    def __init__(self, box_type, coordinates, data=None):
        self.box_type = box_type  # 'instrument_group', 'service_in', etc
        self.coordinates = coordinates  # (x1,y1,x2,y2)
        self.data = data  # Associated data (instrument data from return_inst_data, text, etc)
        self.visual_elements = None  # Canvas elements (boxes, text) - will be recreated on page load

class CaptureGroup:
    def __init__(self):
        self.boxes = []  # List of CaptureBox objects
        self.pid = None
        self.line = None
        self.service_in = None
        self.service_out = None
        self.equipment = None
        self.comment = None
        self.file = None

    def add_box(self, box):
        self.boxes.append(box)

    def clear(self):
        self.boxes = []
        self.pid = None
        self.line = None
        self.service_in = None
        self.service_out = None
        self.equipment = None
        self.comment = None
        self.file = None

class PIDVisionApp:

    # region Initialization and Setup

    def __init__(self, root):
        # Initialize root window
        self.root = root

        # Initialize license manager
        self.license_manager = LicenseManager()

        # Track boxes that have been saved to page_data
        self.saved_boxes = set()

        '''
        # Check license before showing splash screen
        is_licensed, license_status = self.license_manager.check_license()
        if not is_licensed:
            # Instead of immediately shutting down, offer chance to enter a license key
            if self.prompt_for_license_key():
                # License was successfully activated, continue initialization
                pass
            else:
                # User cancelled or failed to enter valid license
                root.destroy()
                return
        '''
        # Hide the main window initially
        self.root.withdraw()
        # Create and show splash screen

        self.splash = SplashScreen(root, r"logo-big.png")

        # Initialize configuration
        self.config = Config()

        # Start initialization in a separate thread
        init_thread = threading.Thread(target=self.initialize_app)
        init_thread.start()

        # Check if initialization is complete
        self.check_initialization(init_thread)

        # Add tracking for annotations and their visibility state
        self.annotations = []
        self.annotations_visible = self.config.annotations_visible

        # Initialize text corrections
        self.text_corrections = TextCorrections(self.folder_path if hasattr(self, 'folder_path') else '')

    def prompt_for_license_key(self):
        """
        Display a dialog to input license key when trial has expired.
        Returns True if a valid license was activated, False otherwise.
        """
        # Create a custom dialog with better formatting
        activation_dialog = tk.Toplevel(self.root)
        activation_dialog.title("License Expired")
        activation_dialog.geometry("500x250")
        activation_dialog.resizable(False, False)
        
        # Make it modal
        activation_dialog.transient(self.root)
        activation_dialog.grab_set()
        
        # Add instructions
        tk.Label(activation_dialog, 
                 text="Your trial period has expired.",
                 font=("Arial", 12, "bold")).pack(pady=(15, 5))
        
        tk.Label(activation_dialog, 
                 text="Please enter your license key to continue using PIDVision:",
                 font=("Arial", 11)).pack(pady=(5, 10))
        
        # Add entry field with more space
        license_entry = tk.Entry(activation_dialog, width=50)
        license_entry.pack(pady=10, padx=20)
        license_entry.focus_set()  # Set focus to the entry field
        
        # Add status label for feedback
        status_label = tk.Label(activation_dialog, text="", fg="red")
        status_label.pack(pady=5)
        
        # Result variable to return from function
        result = [False]
        
        def do_activate():
            license_key = license_entry.get().strip()
            if license_key:
                success, message = self.license_manager.activate_license(license_key)
                if success:
                    result[0] = True
                    activation_dialog.destroy()
                else:
                    status_label.config(text=message, fg="red")
            else:
                status_label.config(text="Please enter a license key", fg="red")
        
        # Add buttons for activation and cancel
        button_frame = tk.Frame(activation_dialog)
        button_frame.pack(pady=15)
        
        tk.Button(button_frame, text="Activate", command=do_activate, width=10).pack(side=tk.LEFT, padx=10)
        tk.Button(button_frame, text="Exit", command=activation_dialog.destroy, width=10).pack(side=tk.LEFT, padx=10)
        
        # Add info about purchase
        purchase_label = tk.Label(
            activation_dialog, 
            text="Don't have a license? Purchase one from our website.",
            font=("Arial", 9), fg="blue", cursor="hand2"
        )
        purchase_label.pack(pady=10)
        
        # Add binding to open purchase page
        purchase_label.bind("<Button-1>", lambda e: self.open_purchase_page())
        
        # Bind Enter key to activation
        activation_dialog.bind("<Return>", lambda e: do_activate())
        
        # Center the dialog
        activation_dialog.update_idletasks()
        width = activation_dialog.winfo_width()
        height = activation_dialog.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        activation_dialog.geometry(f"+{x}+{y}")
        
        # Wait for dialog to close
        self.root.wait_window(activation_dialog)
        
        return result[0]

    def initialize_app(self):
        # Create menu bar
        self.create_menu_bar()

        # Initialize console redirects
        self.console_popup = ConsolePopup(self.root)

        # Initialize canvas and scrollbars
        self.create_canvas_and_scrollbars()

        # Initialize image attributes
        self.initialize_image_attributes()

        # Initialize other attributes
        self.initialize_other_attributes()

        # Initialize instrument and equipment models
        #self.initialize_models()

        # Bind key shortcuts to the respective commands
        self.bind_key_shortcuts()
        # Create data window
        self.create_data_window()
        # Initialize data display
        self.update_data_display()

        # Initialize captured data
        self.capture_mode = 'pid'

    def check_initialization(self, init_thread):
        """Check if initialization is complete and show main window"""
        if init_thread.is_alive():
            # Check again in 100ms
            self.root.after(100, lambda: self.check_initialization(init_thread))
        else:
            # Initialization complete, show main window
            self.show_main_window()

    def show_main_window(self):
        """Show the main window and destroy splash screen"""
        self.splash.destroy()
        self.root.deiconify()
        self.root.title("PIDVision.AI")

    def create_menu_bar(self):
        # Create the main menu bar
        self.menu_bar = tk.Menu(self.root)

        # File Menu
        self.file_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.file_menu.add_command(label="Load Project Folder", command=self.load_project_folder)
        self.file_menu.add_command(label="Open Project Folder", command=self.open_project_folder)
        self.file_menu.add_command(label="Open Current Image", command=self.open_image_path)
        self.file_menu.add_command(label="Open Index", command=self.open_workbook)
        self.file_menu.add_command(label="Save Workbook", command=self.save_workbook)
        self.file_menu.add_separator()
        self.file_menu.add_command(label="Save Page Data", command=self.save_page_data)
        self.file_menu.add_command(label="Load Page Data", command=self.load_page_data)
        self.menu_bar.add_cascade(label="File", menu=self.file_menu)

        # View Menu
        self.view_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.view_menu.add_command(label="Next Page", command=self.next_image)
        self.view_menu.add_command(label="Previous Page", command=self.previous_image)
        self.view_menu.add_command(label="Go to Page...", command=self.open_go_to_page)
        self.view_menu.add_command(label="Open Console", command=self.open_console)
        self.view_menu.add_command(label="Open Page Results", command=self.open_page_results)
        self.view_menu.add_command(label="Select PID", command=self.go_to_pid)
        self.menu_bar.add_cascade(label="View", menu=self.view_menu)

        # Capture Menu
        self.capture_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.capture_menu.add_command(label="PID", command=lambda: self.set_capture('pid'))
        self.capture_menu.add_command(label="Instrument Group", command=lambda: self.set_capture('instruments'))
        self.capture_menu.add_command(label="Line", command=lambda: self.set_capture('line'))
        self.capture_menu.add_command(label="Equipment", command=lambda: self.set_capture('equipment'))
        self.capture_menu.add_command(label="Service In", command=lambda: self.set_capture('service_in'))
        self.capture_menu.add_command(label="Service Out", command=lambda: self.set_capture('service_out'))
        self.capture_menu.add_command(label="Comment", command=lambda: self.set_capture('comment'))
        self.capture_menu.add_separator()
        self.capture_menu.add_command(label="Clear Instrument Group", command=self.clear_instrument_group)
        self.capture_menu.add_command(label="Swap Services", command=self.swap_services)
        self.capture_menu.add_command(label="Vote on Tag Numbers", command=self.vote)
        self.menu_bar.add_cascade(label="Capture", menu=self.capture_menu)

        # Data Menu
        self.data_menu = tk.Menu(self.menu_bar, tearoff=0)

        # Write Mode submenu
        self.write_mode_menu = tk.Menu(self.data_menu, tearoff=0)
        self.write_mode_menu.add_command(label="Live Write (xlwings)", command=lambda: self.set_write_mode('xlwings'))
        self.write_mode_menu.add_command(label="Quick Write (openpyxl)",
                                         command=lambda: self.set_write_mode('openpyxl'))
        self.data_menu.add_cascade(label="Write Mode", menu=self.write_mode_menu)

        self.data_menu.add_command(label="Append Data to Index", command=self.append_data_to_excel)

        # Generate Reports submenu
        self.reports_menu = tk.Menu(self.data_menu, tearoff=0)
        self.reports_menu.add_command(label="Instrument Count (All Pages)", command=self.generate_instrument_count)
        self.reports_menu.add_command(label="Instrument Count (Single Page)", command=self.one_instrument_count)
        self.reports_menu.add_command(label="Compile Instrument Counts", command=self.compile_excels)
        self.reports_menu.add_command(label="Filename PID List", command=self.make_pid_page_xlsx)
        self.reports_menu.add_command(label="Get OCR Results", command=self.get_ocr)
        self.reports_menu.add_command(label="Generate Page Data Report", command=self.generate_page_data_report)
        self.data_menu.add_cascade(label="Generate Reports", menu=self.reports_menu)

        self.menu_bar.add_cascade(label="Data", menu=self.data_menu)

        # Tools Menu
        self.tools_menu = tk.Menu(self.menu_bar, tearoff=0)

        # Applications submenu
        self.apps_menu = tk.Menu(self.tools_menu, tearoff=0)
        self.apps_menu.add_command(label="Image Editor", command=self.open_image_editor)
        self.apps_menu.add_command(label="Find Instrument", command=self.open_FAIA)
        self.apps_menu.add_command(label="OCR Results Viewer", command=self.open_ocr_results_viewer)
        self.apps_menu.add_command(label="Train Detection Model", command=self.open_detecto_gui)
        self.tools_menu.add_cascade(label="Applications", menu=self.apps_menu)

        # PDF Tools submenu
        self.pdf_menu = tk.Menu(self.tools_menu, tearoff=0)
        self.pdf_menu.add_command(label="Create Images from PDF", command=self.create_images_from_pdf)
        self.pdf_menu.add_command(label="Merge PDFs", command=self.merge_pdfs)
        self.tools_menu.add_cascade(label="PDF Tools", menu=self.pdf_menu)

        # Model Management submenu
        self.model_menu = tk.Menu(self.tools_menu, tearoff=0)
        self.model_menu.add_command(label="Load Object Detection Model", command=self.load_pretrained_model)
        self.model_menu.add_command(label="Run Model Test", command=self.test_model_mosaic)

        self.tools_menu.add_cascade(label="Model Management", menu=self.model_menu)

        self.menu_bar.add_cascade(label="Tools", menu=self.tools_menu)

        # Settings Menu
        self.settings_menu = tk.Menu(self.menu_bar, tearoff=0)

        # Reader Settings submenu
        self.reader_settings_menu = tk.Menu(self.settings_menu, tearoff=0)
        self.reader_settings_menu.add_command(label="Instrument Reader", command=self.open_instrument_reader_settings)
        self.reader_settings_menu.add_command(label="General Reader", command=self.open_general_reader_settings)
        self.reader_settings_menu.add_command(label="Reader sub image size and stride", command=self.set_reader_size_and_stride)
        self.reader_settings_menu.add_command(label="Do local OCR", command=self.set_local_ocr)
        self.reader_settings_menu.add_command(label="Comment filtering threshold", command=self.set_ocr_object_interference_threshold)

        self.settings_menu.add_cascade(label="Reader Settings", menu=self.reader_settings_menu)

        # Detection Settings submenu
        self.detection_settings_menu = tk.Menu(self.settings_menu, tearoff=0)
        self.detection_settings_menu.add_command(label="Object Min Scores", command=self.set_object_scores)
        self.detection_settings_menu.add_command(label="NMS Threshold", command=self.set_nms_threshold)
        self.detection_settings_menu.add_command(label="Object Box Expand", command=self.set_object_box_expand)
        self.settings_menu.add_cascade(label="Detection Settings", menu=self.detection_settings_menu)

        # Group Settings submenu
        self.group_settings_menu = tk.Menu(self.settings_menu, tearoff=0)
        self.group_settings_menu.add_command(label="Tag Prefix Groups", command=self.set_tag_label_groups)
        self.group_settings_menu.add_command(label="Instrument Groups", command=self.categorize_labels)
        self.group_settings_menu.add_command(label="Group Association Radius", command=self.set_association_radius)
        self.settings_menu.add_cascade(label="Group Settings", menu=self.group_settings_menu)

        self.settings_menu.add_command(label="Comment Box Settings", command=self.set_comment_box_expand)
        self.settings_menu.add_separator()

        self.settings_menu.add_command(label="Set Line Regular Expression", command=self.set_re_line)
        self.settings_menu.add_command(label="Set HoughLineP and Canny Params", command=self.get_hough_canny_params)
        self.settings_menu.add_command(label="Set Line paint settings", command=self.set_line_params)

        self.settings_menu.add_command(label="Save Settings", command=self.save_attributes)
        self.menu_bar.add_cascade(label="Settings", menu=self.settings_menu)

        # Help Menu
        self.help_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.help_menu.add_command(label="Keybindings", command=self.show_keybindings)
        self.help_menu.add_command(label="Object Detection Labels", command=self.show_labels)
        self.menu_bar.add_cascade(label="Help", menu=self.help_menu)

        # Configure the root window to use the menu bar
        self.root.config(menu=self.menu_bar)

        # Add after your Help menu
        self.license_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.license_menu.add_command(label="Activate License", command=self.activate_license)
        self.license_menu.add_command(label="Check License Status", command=self.check_license_status)
        self.license_menu.add_command(label="Purchase License", command=self.open_purchase_page)

        self.menu_bar.add_cascade(label="License", menu=self.license_menu)
        
        # Add to settings menu
        self.settings_menu.add_command(label="Text Corrections", command=self.open_text_corrections)
        
    def create_canvas_and_scrollbars(self):
        # Vertical and horizontal scrollbars for canvas
        vbar = AutoScrollbar(self.root, orient='vertical')
        hbar = AutoScrollbar(self.root, orient='horizontal')
        vbar.grid(row=0, column=1, sticky='ns')
        hbar.grid(row=1, column=0, sticky='we')

        # Create canvas and put image on it
        self.canvas = tk.Canvas(self.root, highlightthickness=0,
                                xscrollcommand=hbar.set, yscrollcommand=vbar.set)
        self.canvas.grid(row=0, column=0, sticky='nswe')
        self.canvas.update()  # wait till canvas is created
        # Add the optimized canvas handler
        self.optimized_canvas = OptimizedImageCanvas(self.canvas)

        vbar.configure(command=self.scroll_y)  # bind scrollbars to the canvas
        hbar.configure(command=self.scroll_x)

        # Make the canvas expandable
        self.root.rowconfigure(0, weight=1)
        self.root.columnconfigure(0, weight=1)

        self.crop_rectangle = None
        self.image_container = None

    def initialize_image_attributes(self):
        # Initialize image attributes
        self.image_list = []
        self.image_path = ''
        self.current_image_index = self.config.current_image_index
        self.original_image = None
        self.imscale = 1.0  # scale for the canvas image
        self.delta = 1.3  # zoom magnitude
        self.cv2img = None

    def initialize_models(self):
        print('loading models from ', self.config.model_inst_path)
        try:
            # load instrument recognition model
            self.load_pretrained_model(self.config.model_inst_path)
        except Exception as e:
            print('Error', e)
            print('load model failed. you will likely have to load the model from command')

    def initialize_other_attributes(self):
        # Copy configuration values to instance variables for compatibility
        self.model_inst_path = self.config.model_inst_path
        self.ocr_results = None
        self.do_local_ocr = self.config.do_local_ocr
        self.filter_ocr_threshold = self.config.filter_ocr_threshold
        self.comment_box_expand = self.config.comment_box_expand
        self.line = None
        self.re_line = self.config.re_line
        self.service_in = None
        self.service_out = None
        self.equipment = None
        self.pid = None
        self.pid_coords = self.config.pid_coords
        self.comment = None
        self.persistent_boxes = []
        self.persistent_texts = []
        self.correct_fn = None
        self.correct_fn_path = None
        self.mouse_pressed = False
        self.start_x = 0
        self.start_y = 0
        self.current_selection_box = None
        self.current_text = None
        self.wb = None
        self.sheet = None
        self.capture_mode = 'pid'
        self.page_data = {}  # Dictionary to store CaptureGroup objects for each page
        self.current_capture_group = CaptureGroup()  # Current working capture group
        self.capture_actions = {
            'pid': self.capture_pid,
            'instruments': self.capture_instruments,
            'line': self.capture_line,
            'equipment': self.capture_equipment,
            'service_in': self.capture_service_in,
            'service_out': self.capture_service_out,
            'comment': self.capture_comment
        }
        self.whole_page_ocr_results = None
        
        # Initialize EasyOCR reader
        print("Initializing EasyOCR reader...")
        self.reader = easyocr.Reader(['en'], gpu=True)
        
        # Copy more configuration values
        self.reader_stride = self.config.reader_stride
        self.reader_sub_img_size = self.config.reader_sub_img_size
        self.pred_square_size = self.config.pred_square_size
        self.pred_stride = self.config.pred_stride
        self.detection_labels = self.config.detection_labels
        self.tag_label_groups = self.config.tag_label_groups
        self.group_inst = self.config.group_inst
        self.group_other = self.config.group_other
        self.object_box_expand = self.config.object_box_expand
        self.min_scores = self.config.min_scores
        self.association_radius = self.config.association_radius
        self.default_min_detection_score = self.config.default_min_detection_score
        self.nms_threshold = self.config.nms_threshold
        self.inst_data = []
        self.active_inst_box_count = 0
        self.write_mode = self.config.write_mode
        self.equipment_defined = None
        self.crop_start = None
        self.crop_end = None
        self.workbook_path = None
        self.instrument_box_mapping = {}

        # Line detection parameters
        self.canny_params = self.config.canny_params
        self.hough_params = self.config.hough_params
        self.extension_params = self.config.extension_params
        self.paint_line_thickness = self.config.paint_line_thickness
        self.line_join_threshold = self.config.line_join_threshold
        self.line_box_scale = self.config.line_box_scale
        self.line_img_erosion = self.config.line_img_erosion
        self.line_erosion_iterations = self.config.line_erosion_iterations
        self.line_img_binary_threshold = self.config.line_img_binary_threshold
        self.line_img_scale = self.config.line_img_scale
        self.simple_line_mode = self.config.simple_line_mode
        self.debug_line = self.config.debug_line
        self.remove_significant_lines_only = self.config.remove_significant_lines_only
        self.remove_text_before = self.config.remove_text_before
        self.text_min_score = self.config.text_min_score
        self.white_out_color = self.config.white_out_color
        
        # Reader settings
        self.instrument_reader_settings = self.config.instrument_reader_settings
        self.reader_settings = self.config.reader_settings

    def bind_events_to_canvas(self):
        # Bind events to the canvas
        self.canvas.bind('<Configure>', self.show_image)  # canvas is resized
        self.canvas.bind('<MouseWheel>', self.wheel)  # with Windows and MacOS, but not Linux
        self.canvas.bind('<Button-5>', self.wheel)  # only with Linux, wheel scroll down
        self.canvas.bind('<Button-4>', self.wheel)  # only with Linux, wheel scroll up
        self.canvas.bind('<ButtonPress-1>', self.handle_left_click)
        self.canvas.bind('<B1-Motion>', self.update_crop)
        self.canvas.bind('<ButtonRelease-1>', self.end_crop)
        
        # Right click handling - combine panning and editing
        self.canvas.bind('<ButtonPress-3>', self.handle_right_click)
        self.canvas.bind('<B3-Motion>', self.handle_right_motion)
        self.canvas.bind('<ButtonRelease-3>', self.handle_right_release)
        
        # Add instance variables to track right-click state
        self.right_click_start = None
        self.right_click_moved = False

        
    def display_group_data(self, capture_group, box):
        """Display data for the selected capture group"""
        # Update current data with group data
        self.pid = capture_group.pid
        self.line = capture_group.line
        self.service_in = capture_group.service_in
        self.service_out = capture_group.service_out
        self.equipment = capture_group.equipment
        self.comment = capture_group.comment
        
        # Collect instrument data from ALL boxes in the group
        self.inst_data = []
        for box in capture_group.boxes:
            if box.box_type == 'instrument_group' and box.data:
                self.inst_data.extend(box.data.copy())

        self.update_data_display()

    def bind_key_shortcuts(self):
        # Bind key shortcuts to the respective commands
        self.root.bind('n', lambda event: self.next_image())
        self.root.bind('b', lambda event: self.previous_image())
        self.root.bind('p', lambda event: self.set_capture('pid'))
        self.root.bind('a', lambda event: self.set_capture('instruments'))
        self.root.bind('f', lambda event: self.set_capture('line'))
        self.root.bind('e', lambda event: self.set_capture('equipment'))
        self.root.bind('z', lambda event: self.set_capture('service_in'))
        self.root.bind('x', lambda event: self.set_capture('service_out'))
        self.root.bind('w', lambda event: self.append_data_to_excel())
        self.root.bind('c', lambda event: self.clear_instrument_group())
        self.root.bind('v', lambda event: self.vote())
        self.root.bind('s', lambda event: self.swap_services())
        self.root.bind('g', lambda event: self.set_capture('comment'))
        self.root.bind('t', lambda event: self.save_current_group())
        self.root.bind('<Return>', lambda event: self.set_comment())

        # Bind key shortcuts to the respective commands (uppercase)
        self.root.bind('N', lambda event: self.next_image())
        self.root.bind('B', lambda event: self.previous_image())
        self.root.bind('P', lambda event: self.set_capture('pid'))
        self.root.bind('A', lambda event: self.set_capture('instruments'))
        self.root.bind('F', lambda event: self.set_capture('line'))
        self.root.bind('E', lambda event: self.set_capture('equipment'))
        self.root.bind('Z', lambda event: self.set_capture('service_in'))
        self.root.bind('X', lambda event: self.set_capture('service_out'))
        self.root.bind('W', lambda event: self.append_data_to_excel())
        self.root.bind('C', lambda event: self.clear_instrument_group())
        self.root.bind('V', lambda event: self.vote())
        self.root.bind('S', lambda event: self.swap_services())
        self.root.bind('G', lambda event: self.set_capture('comment'))
        self.root.bind('T', lambda event: self.save_current_group())

        # Initialize shift key binding
        self.shift_held = False
        self.root.bind('<KeyPress-Shift_L>', self.shift_pressed)
        self.root.bind('<KeyRelease-Shift_L>', self.shift_released)

        self.ctrl_held = False
        self.root.bind('<KeyPress-Control_L>', self.ctrl_pressed)
        self.root.bind('<KeyRelease-Control_L>', self.ctrl_released)

        # Add binding for 'g' to toggle annotations
        self.root.bind('g', lambda event: self.toggle_annotations())
        self.root.bind('G', lambda event: self.toggle_annotations())

    def save_attributes(self):
        """Save class attributes to configuration"""
        # Update config with current values
        self.config.current_image_index = self.current_image_index
        self.config.pid_coords = self.pid_coords
        self.config.model_inst_path = self.model_inst_path
        self.config.do_local_ocr = self.do_local_ocr
        self.config.filter_ocr_threshold = self.filter_ocr_threshold
        self.config.comment_box_expand = self.comment_box_expand
        self.config.re_line = self.re_line
        self.config.reader_stride = self.reader_stride
        self.config.reader_sub_img_size = self.reader_sub_img_size
        self.config.pred_square_size = self.pred_square_size
        self.config.pred_stride = self.pred_stride
        self.config.detection_labels = self.detection_labels
        self.config.tag_label_groups = self.tag_label_groups
        self.config.group_inst = self.group_inst
        self.config.group_other = self.group_other
        self.config.object_box_expand = self.object_box_expand
        self.config.min_scores = self.min_scores
        self.config.association_radius = self.association_radius
        self.config.default_min_detection_score = self.default_min_detection_score
        self.config.nms_threshold = self.nms_threshold
        self.config.write_mode = self.write_mode
        self.config.canny_params = self.canny_params
        self.config.hough_params = self.hough_params
        self.config.extension_params = self.extension_params
        self.config.paint_line_thickness = self.paint_line_thickness
        self.config.line_join_threshold = self.line_join_threshold
        self.config.line_box_scale = self.line_box_scale
        self.config.line_img_erosion = self.line_img_erosion
        self.config.line_erosion_iterations = self.line_erosion_iterations
        self.config.line_img_binary_threshold = self.line_img_binary_threshold
        self.config.line_img_scale = self.line_img_scale
        self.config.simple_line_mode = self.simple_line_mode
        self.config.debug_line = self.debug_line
        self.config.remove_significant_lines_only = self.remove_significant_lines_only
        self.config.remove_text_before = self.remove_text_before
        self.config.text_min_score = self.text_min_score
        self.config.white_out_color = self.white_out_color
        self.config.instrument_reader_settings = self.instrument_reader_settings
        self.config.reader_settings = self.reader_settings
        self.config.annotations_visible = self.annotations_visible
        
        # Save to file
        self.config.save(self.folder_path)

    def load_attributes(self):
        """Load class attributes from configuration"""
        # Load from file
        self.config.load(self.folder_path)
        
        # Initialize attributes from config
        self.initialize_other_attributes()

    def create_capture_text(self):
        # Remove the previous capture text if it exists
        if hasattr(self, 'capture_text') and self.capture_text:
            self.canvas.delete(self.capture_text)

        # Create the capture text
        self.capture_text = self.canvas.create_text(0, 0, text=self.capture_mode, font=("Arial", 8), fill="orange")

        # Bind the mouse motion event to update the capture text position
        self.canvas.bind("<Motion>", self.update_capture_text)

    # endregion

    # region Image Loading and Navigation

    def load_image(self):
        if self.image_list:
            if hasattr(self, 'optimized_canvas'):
                self.optimized_canvas.clear_cache()

            self.image_path = self.image_list[self.current_image_index]
            print(self.image_path)
            name, ext = os.path.splitext(self.image_path)
            self.results_folder = name + "_results"
            if not os.path.exists(self.results_folder):
                os.makedirs(self.results_folder)

            self.original_image = Image.open(self.image_path)
            self.cv2img = pil_to_cv2(self.original_image)

            self.load_ocr()

            self.width, self.height = self.original_image.size

            # Put image into container rectangle and use it to set proper coordinates to the image
            self.image_container = self.canvas.create_rectangle(0, 0, self.width, self.height, width=0)

            # Bind events to the canvas
            self.bind_events_to_canvas()

            # Get the current window size
            window_width = self.canvas.winfo_width()
            window_height = self.canvas.winfo_height()

            # Calculate the scaling factor
            width_ratio = window_width / self.width
            height_ratio = window_height / self.height
            scale_factor = min(width_ratio, height_ratio)

            # Calculate new dimensions
            new_width = int(self.width * scale_factor)
            new_height = int(self.height * scale_factor)

            # Clear the previous image from the canvas
            self.canvas.delete("all")

            # Set the initial scale
            self.imscale = scale_factor

            # Update the scroll region
            self.canvas.config(scrollregion=(0, 0, new_width, new_height))

            # Create a new image container
            self.image_container = self.canvas.create_rectangle(0, 0, new_width, new_height, width=0)

            # Show the image
            self.show_image()

            # Center the image
            self.center_image()

            # Create capture text
            self.create_capture_text()

            # Recreate boxes from page data if they exist
            self.recreate_boxes()

    def center_image(self):
        # Get the current window size
        window_width = self.canvas.winfo_width()
        window_height = self.canvas.winfo_height()

        # Get the current image size
        bbox = self.canvas.bbox(self.image_container)
        image_width = bbox[2] - bbox[0]
        image_height = bbox[3] - bbox[1]

        # Calculate offsets
        offset_x = 0  # (window_width - image_width) / 2
        offset_y = 0  # (window_height - image_height) / 2

        # Move the image
        self.canvas.move(self.image_container, offset_x, offset_y)
        self.show_image()

    def go_to_page(self, page_index):
        if self.image_list:
            if page_index < 0:
                self.current_image_index = 0  # Set to the first image
            elif page_index >= len(self.image_list):
                self.current_image_index = len(self.image_list) - 1  # Set to the last image
            else:
                self.current_image_index = page_index

            self.load_image()

            if self.pid_coords:
                # Crop the image using the scaled coordinates
                print(self.pid_coords)

                cropped_image = self.original_image.crop(self.pid_coords)

                cropped_image = pil_to_cv2(cropped_image)
                # this is weird but we need to set the self.cropped coords so that we dont let the last used coord overwrite pid
                self.cropped_x1, self.cropped_y1, self.cropped_x2, self.cropped_y2 = self.pid_coords
                self.capture_pid(cropped_image)

                self.update_data_display()

            self.load_ocr()

    def go_to_pid(self):
        path = os.path.join(self.folder_path, "pid_page_1.xlsx")
        print(path)
        workbook = openpyxl.load_workbook(os.path.join(self.folder_path, "pid_page_1.xlsx"), read_only=True)
        worksheet = workbook['pid_page']

        # Create mapping of PID to page index
        pid_to_page = {}
        for row in worksheet.rows:
            if row[0].value and row[1].value:  # Check if both filename and PID exist
                filename = row[0].value
                pid = row[1].value
                # Extract page number from filename (e.g., 'page_1.png' -> 1)
                try:
                    page_num = int(re.search(r'page_(\d+)', filename).group(1)) - 1  # Convert to 0-based index
                    pid_to_page[pid] = page_num
                except (AttributeError, ValueError):
                    continue

        if not pid_to_page:
            messagebox.showerror("Error", "No valid PID mappings found")
            return

        # Create dialog for PID selection
        dialog = tk.Toplevel()
        dialog.title("Select PID")
        dialog.geometry("500x200")  # Width x Height

        # Create combobox with PIDs
        selected_pid = tk.StringVar()
        combo = ttk.Combobox(dialog, textvariable=selected_pid)
        combo['values'] = list(pid_to_page.keys())
        combo.pack(padx=20, pady=20, fill=tk.X, expand=True)

        def on_ok():
            if selected_pid.get() in pid_to_page:
                self.go_to_page(pid_to_page[selected_pid.get()])
            dialog.destroy()

        tk.Button(dialog, text="OK", command=on_ok).pack(pady=10)
        dialog.transient(self.root)  # Make dialog modal
        dialog.grab_set()
        dialog.wait_window()

    def next_image(self):
        self.go_to_page(self.current_image_index + 1)

    def previous_image(self):
        self.go_to_page(self.current_image_index - 1)

    def open_go_to_page(self):

        page_index = tk.simpledialog.askinteger("Go to Page", "Enter the page index:")

        if page_index is not None:
            self.go_to_page(page_index - 1)  # Adjust the index since it's 0-based
            self.clear_boxes()

    def load_project_folder(self, given_folder=None):
        # Define a custom sorting key function
        def natural_sort_key(s):
            return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', s)]

        if not given_folder:
            self.folder_path = filedialog.askdirectory()
        else:
            self.folder_path = given_folder

        if self.folder_path:
            # Initialize config with the project folder
            self.config = Config(self.folder_path)
            
            self.workbook_path = os.path.join(self.folder_path, 'index.xlsx')

            self.image_list = [os.path.join(self.folder_path, f) for f in os.listdir(self.folder_path) if
                               f.endswith(('.png', '.jpg', '.jpeg', '.gif'))]
            # Sort the collected files using the natural sort key function
            self.image_list.sort(key=natural_sort_key)
        
            # Initialize attributes from config
            self.initialize_other_attributes()
            
            print('self.current_image_index', self.current_image_index)

            if self.current_image_index:
                self.go_to_page(self.current_image_index)
            else:
                self.go_to_page(1)
            # Update text corrections path and load corrections
            self.text_corrections = TextCorrections(self.folder_path)

            self.initialize_models()

            if not self.group_inst or not self.tag_label_groups:
                if tk.messagebox.askyesno("Initialize Settings", "Initialize capture settings?"):
                    # Open settings dialogs in sequence
                    self.set_object_scores()
                    self.categorize_labels()
                    self.set_tag_label_groups()



    def create_images_from_pdf(self):
        # Ask for DPI value
        width = tk.simpledialog.askinteger("DPI", "Enter the image width:", initialvalue=5000)

        # Open file dialog to select PDF file
        root = tk.Tk()
        root.withdraw()  # Hide the root window
        pdf_file = filedialog.askopenfilename(title="Select PDF File", filetypes=[("PDF Files", "*.pdf")])

        if pdf_file and width:
            # Call the pdf2png function with the selected PDF file and DPI value
            images_folder = pdf2png(pdf_file, width)
        else:
            return

        if tk.messagebox.askyesno("Open the Project?", "Open the Project?"):
            self.load_project_folder(images_folder)

    # endregion

    # region Canvas and Image Manipulation

    def scroll_y(self, *args, **kwargs):
        ''' Scroll canvas vertically and redraw the image '''
        self.canvas.yview(*args, **kwargs)  # scroll vertically
        self.show_image()  # redraw the image

    def scroll_x(self, *args, **kwargs):
        ''' Scroll canvas horizontally and redraw the image '''
        self.canvas.xview(*args, **kwargs)  # scroll horizontally
        self.show_image()  # redraw the image

    def move_from(self, event):
        ''' Remember previous coordinates for scrolling with the mouse '''
        self.canvas.scan_mark(event.x, event.y)

    def move_to(self, event):
        """Drag (move) canvas to the new position"""
        if event.state & 0x0100:  # Right mouse button
            self.right_click_moved = True
        self.canvas.scan_dragto(event.x, event.y, gain=1)
        self.show_image()  # redraw the image

    def wheel(self, event):
        ''' Zoom with mouse wheel '''
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        bbox = self.canvas.bbox(self.image_container)  # get image area
        if bbox[0] < x < bbox[2] and bbox[1] < y < bbox[3]:
            pass  # Ok! Inside the image
        else:
            return  # zoom only inside image area
        scale = 1.0
        # Respond to Linux (event.num) or Windows (event.delta) wheel event
        if event.num == 5 or event.delta == -120:  # scroll down
            i = min(self.width, self.height)
            if int(i * self.imscale) < 30: return  # image is less than 30 pixels
            self.imscale /= self.delta
            scale /= self.delta
        if event.num == 4 or event.delta == 120:  # scroll up
            i = min(self.canvas.winfo_width(), self.canvas.winfo_height())
            if i < self.imscale: return  # 1 pixel is bigger than the visible area
            self.imscale *= self.delta
            scale *= self.delta
        self.canvas.scale('all', x, y, scale, scale)  # rescale all canvas objects
        # print('scale',scale)
        self.show_image()

    def show_image(self, event=None):
        bbox1 = self.canvas.bbox(self.image_container)
        if not bbox1:
            return

        # Remove 1 pixel shift at the sides of the bbox1
        bbox1 = (bbox1[0] + 1, bbox1[1] + 1, bbox1[2] - 1, bbox1[3] - 1)
        bbox2 = (self.canvas.canvasx(0),
                 self.canvas.canvasy(0),
                 self.canvas.canvasx(self.canvas.winfo_width()),
                 self.canvas.canvasy(self.canvas.winfo_height()))

        # Configure scroll region
        bbox = [min(bbox1[0], bbox2[0]), min(bbox1[1], bbox2[1]),
                max(bbox1[2], bbox2[2]), max(bbox1[3], bbox2[3])]
        self.canvas.configure(scrollregion=bbox)

        # Use optimized canvas to show the image
        self.optimized_canvas.show_image(self.original_image, self.imscale, bbox1, bbox2)

        # Bring crop rectangle to front if it exists
        if self.crop_rectangle:
            self.canvas.tag_raise(self.crop_rectangle)

    def update_capture_text(self, event):
        # Update the text's position to follow the cursor
        # x, y = event.x, event.y
        x, y = self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)
        self.canvas.coords(self.capture_text, x - 8, y - 8)

    # endregion

    # region Cropping and Selection

    def start_crop(self, event):
        self.crop_start = (self.canvas.canvasx(event.x), self.canvas.canvasy(event.y))
        if self.crop_rectangle and self.crop_rectangle not in self.persistent_boxes:
            self.canvas.delete(self.crop_rectangle)

    def update_crop(self, event):
        if self.crop_start:
            x, y = self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)
            if self.crop_rectangle and self.crop_rectangle not in self.persistent_boxes:
                self.canvas.delete(self.crop_rectangle)
            if hasattr(self, 'dimension_text'):
                self.canvas.delete(self.dimension_text)

            # Calculate dimensions in image coordinates
            start_x, start_y = self.canvas_to_image(self.crop_start[0], self.crop_start[1])
            end_x, end_y = self.canvas_to_image(x, y)
            width = abs(end_x - start_x)
            height = abs(end_y - start_y)

            self.crop_rectangle = self.canvas.create_rectangle(
                self.crop_start[0],
                self.crop_start[1],
                x, y,
                outline='orange',
                fill='orange',
                stipple='gray12'  # Creates a checkerboard pattern for transparency effect
            )

            # Create dimension text below cursor
            self.dimension_text = self.canvas.create_text(
                x, y + 25,  # Position text 20 pixels below cursor
                text=f'{width}, {height}',
                fill='red',
                font=('Arial', 8)
            )

    def handle_left_click(self, event):
        """Handle left click events - start crop"""
        x, y = self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)
        self.start_crop(event)
        # Store original click position for checking if mouse moved
        self.original_click_pos = (x, y)

    def end_crop(self, event):
        if self.crop_start:
            self.crop_end = (self.canvas.canvasx(event.x), self.canvas.canvasy(event.y))
            
            # Check if mouse has moved significantly
            if hasattr(self, 'original_click_pos'):
                current_pos = (self.canvas.canvasx(event.x), self.canvas.canvasy(event.y))
                distance = ((current_pos[0] - self.original_click_pos[0])**2 + 
                          (current_pos[1] - self.original_click_pos[1])**2)**0.5
                
                if distance < 5:  # Threshold for considering it a click vs drag
                    # Convert click coordinates to image coordinates for group detection
                    img_x, img_y = self.canvas_to_image(self.original_click_pos[0], self.original_click_pos[1])
                    
                    # Check if click is inside any existing group box
                    if self.image_path in self.page_data:
                        for capture_group in self.page_data[self.image_path]:
                            for box in capture_group.boxes:
                                if box.box_type == 'instrument_group':
                                    x1, y1, x2, y2 = box.coordinates
                                    if x1 <= img_x <= x2 and y1 <= img_y <= y2:
                                        # Click is inside this group box, display its data
                                        self.display_group_data(capture_group, box)
                                        # Clean up
                                        if self.crop_rectangle and self.crop_rectangle not in self.persistent_boxes:
                                            self.canvas.delete(self.crop_rectangle)
                                        if hasattr(self, 'dimension_text'):
                                            self.canvas.delete(self.dimension_text)
                                        self.crop_start = None
                                        self.crop_end = None
                                        return
            
            # If we get here, either the mouse moved or no group was clicked
            self.perform_crop()
            if hasattr(self, 'dimension_text'):
                self.canvas.delete(self.dimension_text)

    def perform_crop(self):
        if self.crop_start and self.crop_end:
            # Convert canvas coordinates to image coordinates
            x1c, y1c = self.canvas_to_image(self.crop_start[0], self.crop_start[1])
            x2c, y2c = self.canvas_to_image(self.crop_end[0], self.crop_end[1])
            # Ensure x1 < x2 and y1 < y2
            x1 = min(x1c, x2c)
            x2 = max(x1c, x2c)

            y1 = min(y1c, y2c)
            y2 = max(y1c, y2c)

            self.cropped_x1 = x1
            self.cropped_y1 = y1
            self.cropped_x2 = x2
            self.cropped_y2 = y2

            # Crop the image
            cropped_image = self.original_image.crop((x1, y1, x2, y2))
            self.cropped_image = pil_to_cv2(cropped_image)
            # Save or display the cropped image
            cropped_image.save("temp/ocr_capture.png")
            print("Image cropped and saved to temp/ocr_capture.png")

            # Perform the action based on self.capture
            if self.capture_mode in self.capture_actions:
                self.capture_actions[self.capture_mode](self.cropped_image)

                # a little goofy but this is to paste the last line captured in the line box
                if self.capture_mode == 'instruments' and self.inst_data:
                    line = self.inst_data[-1].get('line')
                    if line:
                        self.line = line

                self.update_data_display()
            else:
                print(f"Invalid capture action: {self.capture_mode}")

            # Clear the crop rectangle
            if self.crop_rectangle not in self.persistent_boxes:
                self.canvas.delete(self.crop_rectangle)
            self.crop_start = None
            self.crop_end = None

    def canvas_to_image(self, canvas_x, canvas_y):
        # Convert canvas coordinates to image coordinates
        bbox = self.canvas.bbox(self.image_container)
        image_x = int((canvas_x - bbox[0]) / self.imscale)
        image_y = int((canvas_y - bbox[1]) / self.imscale)
        return image_x, image_y

    def clear_boxes(self):

        for box in self.persistent_boxes:
            self.canvas.delete(box)  # Remove the previous box

        # not sure if this is necessary as zip clears stuff
        self.persistent_boxes = []
        self.annotations = []  # Clear annotations list

    # endregion

    # region Data Capture

    def capture_pid(self, cropped_image):
        print('Perform actions for capturing PID')
        try:
            # Perform actions for capturing line
            result = self.reader.readtext(cropped_image, **self.reader_settings)
            if result:
                # Apply text corrections to each OCR result
                corrected_text = [self.text_corrections.apply_corrections(box[1]) for box in result]
                self.pid = ' '.join(corrected_text)
                self.pid_coords = (self.cropped_x1, self.cropped_y1, self.cropped_x2, self.cropped_y2)
            else:
                self.pid = ''
                print('no result')
        except Exception as e:
            print('error capturing pid \n', e)

    def canvas_coords_from_image(self, x, y):
        """Convert image coordinates to canvas coordinates, accounting for scaling and offset"""
        bbox = self.canvas.bbox(self.image_container)
        canvas_x = bbox[0] + (x * self.imscale)
        canvas_y = bbox[1] + (y * self.imscale)
        return canvas_x, canvas_y

    def draw_detection_box(self, box, color='red'):
        """Draw a detection box on the canvas, accounting for scaling and offset"""
        x1, y1, x2, y2 = [float(coord) for coord in box]

        # Convert image coordinates to canvas coordinates
        # If the box coordinates already include the offset (from saved data), don't add it again
        if hasattr(self, 'recreating_boxes') and self.recreating_boxes:
            canvas_x1, canvas_y1 = self.canvas_coords_from_image(x1, y1)
            canvas_x2, canvas_y2 = self.canvas_coords_from_image(x2, y2)
        else:
            canvas_x1, canvas_y1 = self.canvas_coords_from_image(x1 + self.cropped_x1, y1 + self.cropped_y1)
            canvas_x2, canvas_y2 = self.canvas_coords_from_image(x2 + self.cropped_x1, y2 + self.cropped_y1)

        # Create rectangle on canvas
        return self.canvas.create_rectangle(
            canvas_x1, canvas_y1, canvas_x2, canvas_y2,
            outline=color,
            width=2
        )

    def capture_instruments(self, cropped_image):
        # Lazy load model
        if not hasattr(self, 'model_inst') or self.model_inst is None:
            print("Loading object detection model...")
            self.model_inst = Model.load(self.model_inst_path, self.detection_labels)
        
        offset = (self.cropped_x1, self.cropped_y1)
        labels, boxes, scores = model_predict_on_mosaic(
            cropped_image,
            self.model_inst,
            threshold=self.nms_threshold,
            square_size=self.pred_square_size,
            stride=self.pred_stride
        )

        if not labels:
            return

        # Process visual elements
        visual_elements_list, new_active_boxes = self.process_visual_elements(
            labels, boxes, scores, offset
        )

        # Update persistent boxes and box count
        self.persistent_boxes.extend(new_active_boxes)
        self.active_inst_box_count += len(new_active_boxes)

        # Create prediction data with visual elements
        inst_prediction_data = zip(labels, boxes, scores, visual_elements_list)

        # Create LineProcessingParams instance
        line_params = LineProcessingParams(
            re_line=self.re_line,  # Added this
            simple=self.simple_line_mode,  # Changed from simple_mode to simple
            debug_line=self.debug_line,
            remove_significant_lines_only=self.remove_significant_lines_only,
            paint_line_thickness=self.paint_line_thickness,
            line_join_threshold=self.line_join_threshold,
            line_box_scale=self.line_box_scale,
            erosion_kernel=self.line_img_erosion,  # Changed from line_img_erosion to erosion_kernel
            erosion_iterations=self.line_erosion_iterations,
            binary_threshold=self.line_img_binary_threshold,
            line_img_scale=self.line_img_scale,
            hough_params = self.hough_params,
            canny_params = self.canny_params,
            extension_params = self.extension_params,
            remove_text_before = self.remove_text_before,
            text_min_score = self.text_min_score,
            white_out_color = self.white_out_color
        )

        # Process instrument data
        inst_data = return_inst_data(
            inst_prediction_data,
            cropped_image,
            self.reader,
            instrument_reader_settings=self.instrument_reader_settings,
            reader_settings=self.reader_settings,
            radius=self.association_radius,
            min_scores=self.min_scores,
            expand=self.object_box_expand,
            offset=offset,
            comment_box_expand=self.comment_box_expand,
            inst_labels=self.group_inst,
            other_labels=self.group_other,
            tag_label_groups=self.tag_label_groups,
            capture_ocr=self.do_local_ocr,
            filter_ocr_threshold=self.filter_ocr_threshold,
            line_params=line_params  # Pass the LineProcessingParams instance
        )

        # Print and store instrument data
        if inst_data:
            for data in inst_data:
                print(data)

        self.inst_data.extend(inst_data)

    def capture_line(self, cropped_image):
        self.process_captured_text(cropped_image, 'line')

    def capture_equipment(self, cropped_image):
        self.process_captured_text(cropped_image, 'equipment')

    def capture_service_in(self, cropped_image):
        self.process_captured_text(cropped_image, 'service_in')

    def capture_service_out(self, cropped_image):
        self.process_captured_text(cropped_image, 'service_out')

    def capture_comment(self, cropped_image):
        print('Perform actions for capturing a comment')
        # Perform actions for capturing line
        result = self.reader.readtext(cropped_image, **self.reader_settings)
        if result:
            # Apply text corrections to each OCR result
            corrected_text = [self.text_corrections.apply_corrections(box[1]) for box in result]
            self.comment = ' '.join(corrected_text)
        else:
            self.comment = ''

    def process_captured_text(self, cropped_image, target_attribute):
        """
        Unified method for processing captured text from images.

        Args:
            cropped_image: The image to process
            target_attribute: String name of attribute to update ('line', 'equipment', 'service_in', 'service_out')
        """
        # Handle rotation for line captures only
        if target_attribute == 'line':
            height, width = cropped_image.shape[:2]
            if height > width:
                cropped_image = cv2.rotate(cropped_image, cv2.ROTATE_90_CLOCKWISE)

        # Perform OCR
        result = self.reader.readtext(cropped_image, **self.reader_settings)

        if not result:
            setattr(self, target_attribute, '')
            return

        # Apply text corrections to each OCR result
        corrected_text = [self.text_corrections.apply_corrections(box[1]) for box in result]
        just_text = ' '.join(corrected_text)
        current_value = getattr(self, target_attribute, '')  # Default to empty string if attribute doesn't exist

        # Process text based on modifier keys
        if self.ctrl_held:
            new_text = merge_common_substrings(current_value, just_text)
        elif self.shift_held:
            new_text = f"{current_value} {just_text}".strip()
        else:
            new_text = just_text

        # Update the target attribute
        setattr(self, target_attribute, new_text)

        # Clear other attributes based on the capture type
        if target_attribute in ('line', 'equipment'):
            other_attr = 'equipment' if target_attribute == 'line' else 'line'
            setattr(self, other_attr, None)
        elif target_attribute in ('service_in', 'service_out'):
            self.equipment = None

    def set_capture(self, capture_type):
        self.capture_mode = capture_type
        self.canvas.itemconfig(self.capture_text, text=self.capture_mode)

    # endregion

    # region Data Management

    def create_data_window(self):
        self.data_window = tk.Toplevel(self.root)
        self.data_window.title("Captured Data")

        # Get screen dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        # Account for Windows taskbar (typically 40 pixels)
        taskbar_height = 40
        available_height = screen_height - taskbar_height

        # Set data window width and position
        data_window_width = 300  # Increased from 250 to accommodate headers
        data_window_x = screen_width - data_window_width

        # Position window accounting for taskbar
        self.data_window.geometry(f'{data_window_width}x{available_height}+{data_window_x}+0')

        # Create main frame with padding that fills the window
        self.data_frame = ttk.Frame(self.data_window)
        self.data_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Create upper frame for entry fields
        upper_frame = ttk.Frame(self.data_frame)
        upper_frame.pack(fill=tk.X, expand=False)

        # Entry fields in upper frame
        ttk.Label(upper_frame, text="PID:").pack(anchor='w')
        self.pid_entry = ttk.Entry(upper_frame)
        self.pid_entry.pack(fill=tk.X, padx=5, pady=(0, 10))


        # Create page navigation frame
        page_frame = ttk.Frame(upper_frame)
        page_frame.pack(anchor='w', pady=(0, 10))

        # Back button
        back_button = ttk.Button(page_frame, text="←", width=2, command=self.previous_image)
        back_button.pack(side=tk.LEFT, padx=(0,2))

        ttk.Label(page_frame, text="Page: ").pack(side=tk.LEFT)

        # Create page entry with validation
        self.page_entry = ttk.Entry(page_frame, width=6)
        self.page_entry.pack(side=tk.LEFT, padx=2)

        # Label for total pages
        self.total_pages_label = ttk.Label(page_frame, text=f" of {len(self.image_list)}")
        self.total_pages_label.pack(side=tk.LEFT)

        # Forward button
        forward_button = ttk.Button(page_frame, text="→", width=2, command=self.next_image)
        forward_button.pack(side=tk.LEFT, padx=(2,0))

        # Bind Enter key to page navigation
        self.page_entry.bind('<Return>', self.navigate_to_page)


        # Other entry fields
        field_configs = [
            ("Line:", "line_entry"),
            ("Service In:", "service_in_entry"),
            ("Service Out:", "service_out_entry"),
            ("Equipment:", "equipment_entry"),
            ("Comment:", "comment_entry")
        ]

        for label_text, entry_name in field_configs:
            ttk.Label(upper_frame, text=label_text).pack(anchor='w')
            entry = ttk.Entry(upper_frame)
            entry.pack(fill=tk.X, padx=5, pady=(0, 10))
            setattr(self, entry_name, entry)

        # Create lower frame for tree
        lower_frame = ttk.Frame(self.data_frame)
        lower_frame.pack(fill=tk.BOTH, expand=True)

        # Instrument Data Tree label
        ttk.Label(lower_frame, text="Instrument Data:").pack(anchor='w')

        # Create tree frame to contain tree and its scrollbar
        tree_frame = ttk.Frame(lower_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=(5, 0))

        # Create Treeview with updated columns
        self.inst_tree = ttk.Treeview(tree_frame, columns=('Tag', 'Tag No', 'Type', 'Line'),
                                      show='headings')

        # Configure column widths and headings
        self.inst_tree.column('Tag', width=80, minwidth=80)
        self.inst_tree.column('Tag No', width=100, minwidth=100)
        self.inst_tree.column('Type', width=100, minwidth=100)
        self.inst_tree.column('Line', width=120, minwidth=120)  # Add new column

        self.inst_tree.heading('Tag', text='Tag')
        self.inst_tree.heading('Tag No', text='Tag No')
        self.inst_tree.heading('Type', text='Type')
        self.inst_tree.heading('Line', text='Line')  # Add new heading

        # Add vertical scrollbar for tree
        tree_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL,
                                       command=self.inst_tree.yview)
        self.inst_tree.configure(yscrollcommand=tree_scrollbar.set)

        # Pack tree and its scrollbar
        self.inst_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Bind events
        for entry in [self.pid_entry, self.line_entry, self.service_in_entry,
                      self.service_out_entry, self.equipment_entry, self.comment_entry]:
            entry.bind('<FocusOut>', lambda e: self.get_data_from_window())

        # Double-click editing for tree
        self.inst_tree.bind('<Double-1>', self.edit_instrument)

        # Initialize mapping dictionary
        self.instrument_box_mapping = {}

        # Refresh tree bindings
        self.refresh_tree_bindings()

        # Window management
        self.data_window.protocol("WM_DELETE_WINDOW", self.update_data_display)
        self.root.bind("<FocusIn>", lambda event: self.data_window.lift())

        # Position main window
        root_window_width = screen_width - data_window_width
        self.root.geometry(f'{root_window_width}x{available_height}+0+0')

    def navigate_to_page(self, event):
        """Handle page navigation from entry widget"""
        try:
            # Get page number from entry
            page_num = int(self.page_entry.get())

            # Validate page number
            if 1 <= page_num <= len(self.image_list):
                # Go to page (subtract 1 since page_num is 1-based but go_to_page expects 0-based index)
                self.go_to_page(page_num - 1)
            else:
                tk.messagebox.showwarning("Invalid Page",
                                          f"Please enter a page number between 1 and {len(self.image_list)}")

        except ValueError:
            tk.messagebox.showwarning("Invalid Input",
                                      "Please enter a valid page number")

        # Restore focus to entry widget
        self.page_entry.focus()

    def update_data_display(self):
        # Clear existing tree items and box mappings
        for item in self.inst_tree.get_children():
            self.inst_tree.delete(item)
        self.instrument_box_mapping.clear()

        # Update entries with current values
        self.pid_entry.delete(0, tk.END)
        self.pid_entry.insert(0, str(self.pid) if self.pid else "")

        # In the update_data_display method, replace the page label update with:
        self.page_entry.delete(0, tk.END)
        self.page_entry.insert(0, str(self.current_image_index + 1))
        self.total_pages_label.config(text=f" of {len(self.image_list)}")


        self.line_entry.delete(0, tk.END)
        self.line_entry.insert(0, str(self.line) if self.line else "")

        self.service_in_entry.delete(0, tk.END)
        self.service_in_entry.insert(0, condense_hyphen_string(self.service_in) if self.service_in else "")

        self.service_out_entry.delete(0, tk.END)
        self.service_out_entry.insert(0, condense_hyphen_string(self.service_out) if self.service_out else "")

        self.equipment_entry.delete(0, tk.END)
        self.equipment_entry.insert(0, str(self.equipment) if self.equipment else "")

        self.comment_entry.delete(0, tk.END)
        self.comment_entry.insert(0, str(self.comment) if self.comment else "")

        # Update instrument tree and mapping
        for i, data in enumerate(self.inst_data):
            tree_item = self.inst_tree.insert('', tk.END, values=(
                data.get('tag', ''),
                data.get('tag_no', ''),
                data.get('type', ''),
                data.get('line', '')  # Add line data

            ))

            # Store both the detection box and label text IDs in the mapping
            if 'visual_elements' in data:
                self.instrument_box_mapping[tree_item] = data['visual_elements']

    def get_data_from_window(self):
        """Retrieve data from the window widgets and update instrument data"""
        self.pid = self.pid_entry.get()
        self.line = self.line_entry.get()
        self.service_in = self.service_in_entry.get()
        self.service_out = self.service_out_entry.get()
        self.equipment = self.equipment_entry.get()
        self.comment = self.comment_entry.get()

        # Update existing instrument data while preserving structure
        updated_data = []
        for item in self.inst_tree.get_children():
            values = self.inst_tree.item(item)['values']
            inst_dict = {
                'tag': values[0] if values[0] else '',
                'tag_no': values[1] if values[1] else '',
                'type': values[2] if values[2] else ''
            }

            # Get the visual elements for this tree item
            visual_elements = self.instrument_box_mapping.get(item)

            # Find matching data using visual elements as unique identifier
            if self.inst_data and visual_elements:
                existing = next((d for d in self.inst_data
                                 if d.get('visual_elements') == visual_elements), {})
                inst_dict.update({k: v for k, v in existing.items()
                                  if k not in ['tag', 'tag_no', 'type']})

            updated_data.append(inst_dict)

        print('updated data from data window: \n', updated_data)
        self.inst_data = updated_data

    def edit_instrument(self, event):
        """Handle double-click editing of instrument data"""
        item = self.inst_tree.selection()[0]
        values = self.inst_tree.item(item)['values']

        # Get the current data for this instrument
        current_inst_data = None
        visual_elements = self.instrument_box_mapping.get(item)
        if visual_elements and self.inst_data:
            current_inst_data = next((d for d in self.inst_data
                                      if d.get('visual_elements') == visual_elements), None)

        # Create popup window for editing
        edit_window = tk.Toplevel(self.data_window)
        edit_window.title("Edit Instrument")
        edit_window.bind('<Return>', lambda e: save_changes())

        # Create and pack a frame for better layout
        frame = ttk.Frame(edit_window, padding="10")
        frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Create entry fields
        ttk.Label(frame, text="Tag:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        tag_entry = ttk.Entry(frame)
        tag_entry.insert(0, values[0])
        tag_entry.grid(row=0, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))

        ttk.Label(frame, text="Tag No:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        tag_no_entry = ttk.Entry(frame)
        tag_no_entry.insert(0, values[1])
        tag_no_entry.grid(row=1, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))

        ttk.Label(frame, text="Type:").grid(row=2, column=0, padx=5, pady=5, sticky=tk.W)
        type_entry = ttk.Entry(frame)
        type_entry.insert(0, values[2])
        type_entry.grid(row=2, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))

        ttk.Label(frame, text="Line:").grid(row=3, column=0, padx=5, pady=5, sticky=tk.W)
        line_entry = ttk.Entry(frame)
        if len(values) > 3:  # Check if line value exists in the tree
            line_entry.insert(0, values[3])
        elif current_inst_data and 'line' in current_inst_data:
            line_entry.insert(0, current_inst_data['line'])
        line_entry.grid(row=3, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))

        ttk.Label(frame, text="Comment:").grid(row=4, column=0, padx=5, pady=5, sticky=tk.W)
        comment_entry = ttk.Entry(frame, width=40)
        if current_inst_data and 'comment' in current_inst_data:
            comment_entry.insert(0, current_inst_data['comment'])
        comment_entry.grid(row=4, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))

        def save_changes():
            # Update tree view with all fields including line
            self.inst_tree.item(item, values=(
                tag_entry.get(),
                tag_no_entry.get(),
                type_entry.get(),
                line_entry.get()
            ))

            # Update instrument data
            if current_inst_data:
                current_inst_data['tag'] = tag_entry.get()
                current_inst_data['tag_no'] = tag_no_entry.get()
                current_inst_data['type'] = type_entry.get()
                current_inst_data['line'] = line_entry.get()
                current_inst_data['comment'] = comment_entry.get()

            edit_window.destroy()

        # Create a frame for buttons
        button_frame = ttk.Frame(frame)
        button_frame.grid(row=5, column=0, columnspan=2, pady=10)

        ttk.Button(button_frame, text="Save", command=save_changes).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=edit_window.destroy).pack(side=tk.LEFT, padx=5)

        # Configure grid weights
        frame.columnconfigure(1, weight=1)

        # Make dialog modal
        edit_window.transient(self.data_window)
        edit_window.grab_set()

        # Center the window
        edit_window.update_idletasks()
        width = edit_window.winfo_width()
        height = edit_window.winfo_height()
        x = self.data_window.winfo_x() - (self.data_window.winfo_width() // 2) - (width // 2)
        y = self.data_window.winfo_y() + (self.data_window.winfo_height() // 2) - (height // 2)
        edit_window.geometry(f'+{x}+{y}')
    
    def append_data(self, excel_type='xlwings'):
        """
        Append data to Excel file using either xlwings or openpyxl, optimized for bulk writing.
        """
        self.get_data_from_window()

        # Initialize workbook and worksheet
        if excel_type == 'xlwings':
            if not os.path.exists(self.workbook_path):
                self.wb = xw.Book()
                self.wb.save(self.workbook_path)
            wb = xw.Book(self.workbook_path)
            if 'Instrument Index' not in wb.sheet_names:
                wb.sheets.add(name='Instrument Index')
            ws = wb.sheets['Instrument Index']

            # Get last row
            last_row = ws.range('A1').expand('down').last_cell.row if ws.range('A1').value is not None else 0

            # Prepare data as a list of lists for bulk writing
            all_data = []
            headers = None

            for data in self.inst_data:
                processed_data = self._prepare_data_dict(data)
                if not headers:
                    headers = list(processed_data.keys())
                    if last_row == 0:  # If sheet is empty
                        all_data.append(headers)

                row_data = []
                for header in headers:
                    value = processed_data[header]

                    row_data.append(str(value) if value is not None else '')
                all_data.append(row_data)

            # Write all data at once
            if all_data:
                range_str = f'A{last_row + 1}'
                ws.range(range_str).value = all_data

        else:  # openpyxl
            if not os.path.exists(self.workbook_path):
                self.wb = openpyxl.Workbook()
            elif not self.wb:
                self.wb = openpyxl.load_workbook(self.workbook_path)
            if 'Instrument Index' not in self.wb.sheetnames:
                self.ws = self.wb.create_sheet('Instrument Index')
            else:
                self.ws = self.wb['Instrument Index']
            ws = self.ws

            # Get last row
            last_row = len(list(ws.rows))

            # Prepare data
            all_data = []
            headers = None

            for data in self.inst_data:
                processed_data = self._prepare_data_dict(data)
                if not headers:
                    headers = list(processed_data.keys())
                    if last_row == 0:  # If sheet is empty
                        ws.append(headers)
                        last_row += 1

                row_data = []
                for header in headers:
                    value = processed_data[header]
                    # Handle multi-element tensors
                    if hasattr(value, 'numpy'):  # PyTorch tensor
                        value = value.detach().numpy()
                    if hasattr(value, 'tolist'):  # NumPy array
                        value = value.tolist()
                    if isinstance(value, list):
                        value = ', '.join(str(v) for v in value)
                    row_data.append(str(value) if value is not None else '')
                all_data.append(row_data)

            # Write all data at once
            if all_data:
                for row_data in all_data:
                    ws.append(row_data)

            self.wb.save(self.workbook_path)

        self.turn_boxes_blue()

    def _prepare_data_dict(self, data):
        """
        Prepare a single data dictionary with all necessary processing.
        """
        processed_data = data.copy()

        # Add default fields
        processed_data.update({
            'PID': self.pid,
            'SERVICE': '',
            'LINE/EQUIP': '',
            'INPUT COMMENT': self.comment if self.comment else '',
            'FILE': self.image_path,
        })

        # Process service information
        si = condense_hyphen_string(self.service_in)
        so = condense_hyphen_string(self.service_out)

        if self.service_in and self.service_out:
            processed_data[
                'SERVICE'] = f"{si} TO {so}" if self.service_in != self.service_out else f"{si} RECIRCULATION"
        elif self.service_in:
            processed_data['SERVICE'] = f"{si} OUTLET"
        elif self.service_out:
            processed_data['SERVICE'] = f"TO {so}"

        # Process line/equipment information
        if self.line:
            processed_data['LINE/EQUIP'] = self.line
        elif self.equipment:
            words = self.equipment.split(' ')
            processed_data['LINE/EQUIP'] = words[0]
            processed_data['SERVICE'] = ' '.join(words[1:])

        return processed_data

    def append_data_to_excel(self):
        try:
            self.append_data(self.write_mode)
        except Exception as e:
            tk.messagebox.showerror(e)
            print(f"error {e}")

    def create_excel_header(self, worksheet, data):
        """
        Create a header row in the given worksheet based on the keys in the data dictionary
        and format entire columns as text.
        :param worksheet: The worksheet to add the header to (either xlwings or openpyxl worksheet)
        :param data: A dictionary containing the data structure
        :return: None
        """
        header = list(data.keys())
        print(header)

        if isinstance(worksheet, xw.main.Sheet):
            # xlwings worksheet
            for i, column_header in enumerate(header, start=1):
                cell = worksheet.range((1, i))
                cell.value = column_header
                # Format entire column as text
                # Get the last row in the worksheet
                last_row = worksheet.cells.last_cell.row
                # Format the entire column from row 1 to last row
            column_range = worksheet.range((1, 2), (last_row, 2))
            column_range.number_format = '@'

        elif isinstance(worksheet, openpyxl.worksheet.worksheet.Worksheet):
            # openpyxl worksheet
            for col, column_header in enumerate(header, start=1):
                worksheet.cell(row=1, column=col).value = column_header
            # Format entire column as text
            column_letter = openpyxl.utils.get_column_letter(2)
            for cell in worksheet[column_letter]:
                cell.number_format = '@'
        else:
            raise ValueError("Unsupported worksheet type")

    def populate_excel_row(self, worksheet, data, row):
        """
        Populate a row in the given worksheet with the provided data.

        :param worksheet: The worksheet to populate (either xlwings or openpyxl worksheet)
        :param data: A dictionary containing the data to be written
        :param row: The row number to populate
        :return: None
        """
        # Ensure all dictionary values are strings and empty strings for None
        print('writing data: ', data)
        
        # Exclude visual elements from the data
        processed_data = {k: str(v) if v is not None else '' for k, v in data.items() if 'visual_elements' not in k}
        print('processed data: ', processed_data)

        for col, (key, value) in enumerate(processed_data.items(), start=1):
            # Skip if value is an empty string
            if value == '':
                continue

            if isinstance(worksheet, xw.main.Sheet):
                # xlwings worksheet
                cell = worksheet.range((row, col))
                cell.value = value
            elif isinstance(worksheet, openpyxl.worksheet.worksheet.Worksheet):
                # openpyxl worksheet
                cell = worksheet.cell(row=row, column=col)
                cell.value = value
            else:
                raise ValueError("Unsupported worksheet type")


    def turn_boxes_blue(self):
        # Slice the list to get active items
        active_boxes = self.persistent_boxes[-self.active_inst_box_count:]

        # Iterate over the sliced list and change colors based on item type
        for box_id in active_boxes:
            item_type = self.canvas.type(box_id)
            if item_type == "rectangle":
                self.canvas.itemconfig(box_id, outline='#87CEEB')
            elif item_type == "text":
                self.canvas.itemconfig(box_id, fill='#87CEEB')

        self.active_inst_box_count = 0
        self.inst_data = []
        self.update_data_display()

    def vote(self):
        # Create a dictionary to store tag_no counts
        tag_counts = {}

        # Count the occurrences of each tag_no
        for data in self.inst_data:
            tag_no = data['tag_no']
            tag_counts[tag_no] = tag_counts.get(tag_no, 0) + 1

        # Find the most frequent tag_no
        most_frequent_tag_no = max(tag_counts, key=tag_counts.get)

        # Assign the most frequent tag_no to each entry
        for data in self.inst_data:
            data['tag_no'] = most_frequent_tag_no

        self.update_data_display()

    def swap_services(self):
        self.service_in, self.service_out = self.service_out, self.service_in
        self.update_data_display()

    def set_comment(self):
        self.comment = tk.simpledialog.askstring("Input", "Please enter a Comment:")
        self.update_data_display()

    def clear_instrument_group(self):
        # Only clear boxes that haven't been saved (not in saved_boxes)
        if self.inst_data:
            # Get the boxes from the last capture (active boxes)
            active_boxes = self.persistent_boxes[-self.active_inst_box_count:]
            
            # Only remove boxes that haven't been saved
            boxes_to_remove = [box for box in active_boxes if box not in self.saved_boxes]
            
            # Delete unsaved boxes from canvas
            for box in boxes_to_remove:
                self.canvas.delete(box)
                
            # Remove deleted boxes from persistent_boxes list
            self.persistent_boxes = [box for box in self.persistent_boxes if box not in boxes_to_remove]
                
        # Clear current instrument data and reset active box count
        self.inst_data = []
        self.active_inst_box_count = 0
        self.update_data_display()

    # endregion

    # region Visual Data Object Linking Instrument Data Tree

    def refresh_tree_bindings(self):
        """Setup tree view hover bindings"""
        self.inst_tree.bind('<Motion>', self.on_tree_motion)
        self.inst_tree.bind('<Leave>', self.on_tree_leave)
        self.inst_tree.bind('<Enter>', self.on_tree_enter)

    def on_tree_enter(self, event):
        """Show all boxes initially when entering tree widget"""
        self.show_all_boxes()

    def on_tree_leave(self, event):
        """Show all boxes when leaving tree widget"""
        self.show_all_boxes()

    def hide_all_boxes_except(self, tree_item):
        """Hide all instrument boxes except for the one being hovered"""
        for item_id, box_ids in self.instrument_box_mapping.items():
            box = box_ids['box']
            text = box_ids['text']
            if item_id == tree_item:
                self.canvas.itemconfig(box, state='normal')
                self.canvas.itemconfig(text, state='normal')
            else:
                self.canvas.itemconfig(box, state='hidden')
                self.canvas.itemconfig(text, state='hidden')

    def on_tree_motion(self, event):
        """Handle mouse movement over tree items"""
        item = self.inst_tree.identify_row(event.y)
        if item:
            self.highlight_selected_instrument(item)
        else:
            self.show_all_boxes()

    def highlight_selected_instrument(self, tree_item):
        """Highlight the selected instrument and hide others"""
        for item_id, visual_elements in self.instrument_box_mapping.items():
            box = visual_elements['box']
            text = visual_elements['text']
            if item_id == tree_item:
                self.canvas.itemconfig(box, state='normal')
                self.canvas.itemconfig(text, state='normal')
            else:
                self.canvas.itemconfig(box, state='hidden')
                self.canvas.itemconfig(text, state='hidden')

    def show_all_boxes(self):
        """Show all instrument boxes"""
        for visual_elements in self.instrument_box_mapping.values():
            self.canvas.itemconfig(visual_elements['box'], state='normal')
            self.canvas.itemconfig(visual_elements['text'], state='normal')

    def process_visual_elements(self, labels, boxes, scores, offset):
        visual_elements_list = []
        active_boxes = []

        # Add crop rectangle to persistent boxes
        active_boxes.append(self.crop_rectangle)

        # Process each detection
        for label, box, score in zip(labels, boxes, scores):
            # Skip if minimum score threshold not met
            if self.min_scores.get(label) is None:
                continue
            if self.min_scores.get(label) >= float(score):
                continue

            # Determine color based on instrument group
            color = 'red' if label in self.group_inst else 'blue'

            # Draw detection box
            detection_box = self.draw_detection_box(box, color)
            active_boxes.append(detection_box)

            # Create text label - handle offset differently for recreation
            if hasattr(self, 'recreating_boxes') and self.recreating_boxes:
                x1, y1 = self.canvas_coords_from_image(
                    float(box[0]),
                    float(box[1])
                )
            else:
                x1, y1 = self.canvas_coords_from_image(
                    float(box[0]) + offset[0],
                    float(box[1]) + offset[1]
                )

            text_id = self.canvas.create_text(
                x1, y1 - 5,  # Position above the box
                text=f"{label} ({float(score):.2f})",
                fill=color,
                anchor='sw',
                font=('Arial', 8)
            )
            active_boxes.append(text_id)

            # Store visual elements for this detection
            visual_elements_list.append({
                'box': detection_box,
                'text': text_id
            })

            # Print detection information
            print(f"{label}: {score:.2f} | ", end='')

        print()  # New line after printing all detections

        return visual_elements_list, active_boxes

    # endregion

    # region OCR and Model Operations

    def get_ocr(self):
        self.ocr_results = HardOCR(self.cv2img, self.reader, self.reader_settings)
        img_ocr_results = plot_ocr_results(self.cv2img, self.ocr_results)
        print(self.ocr_results)
        # save a pic
        img_save_path = os.path.join(self.results_folder, 'ocr_img.png')
        cv2.imwrite(img_save_path, img_ocr_results)

        # save actual results
        save_path = os.path.join(self.results_folder, 'ocr.pkl')
        save_easyocr_results_pickle(self.ocr_results, filename=save_path)

        os.startfile(img_save_path)

    def get_all_ocr(self):
        self.current_image_index = 0

        for i in range(len(self.image_list)):
            self.load_image()
            self.get_ocr()
            self.current_image_index += 1

    def load_ocr(self):
        file_name = os.path.join(self.results_folder, 'ocr.pkl')
        self.ocr_results = load_easyocr_results_pickle(filename=file_name)

    def load_model(self):
        # Create a module specification
        self.model_inst_path = filedialog.askopenfilename(filetypes=[('PTH Files', '*.pth')])
        print('loading model')
        self.model_inst = Model.load(self.model_inst_path, self.detection_labels)

    def load_pretrained_model(self, model_path=None):
        if model_path == None:
            model_path = filedialog.askopenfilename(title="Select pretrained model file",
                                                    filetypes=[("Model files", "*.pth")])
        if model_path:
            label_path = model_path.replace(".pth", ".txt")
            if os.path.exists(label_path):
                with open(label_path, "r") as f:
                    self.detection_labels = [x.strip() for x in f.read().split(",")]
                    print(self.detection_labels)
            else:
                tk.messagebox.showinfo("Load error", f"{label_path} with comma separated labels not found")
            print('make sure labels are loaded correctly')
            self.model_inst = Model.load(model_path, self.detection_labels)
            print("Pretrained model loaded successfully.")

    def set_write_mode(self, mode):
        if hasattr(self, 'wb') and self.wb is not None:
            try:
                if self.write_mode == 'xlwings':
                    self.wb.close()
                elif self.write_mode == 'openpyxl':
                    self.wb.close()
                self.wb = None
                self.ws = None
            except Exception as e:
                print(f"Error closing workbook: {e}")

        self.write_mode = mode

    def test_model_mosaic(self):

        test_on_capture = tk.messagebox.askyesno(title='test on capture?',
                                                 message='Do you want to test the model on the Capture?')

        if self.model_inst is None:
            tk.messagebox.showerror("Error", "Model not loaded. Load a pretrained model first.")
            return

        minscore = tk.simpledialog.askinteger("Scale percent", "Enter a minscore 1-100:", initialvalue=50) / 100

        if test_on_capture == False:
            # Ask for an image to test
            image_path = filedialog.askopenfilename(title="Select an image for testing", initialfile=self.image_path,
                                                    filetypes=(("PNG files", "*.png"), ("JPEG files", "*.jpg")))
        else:
            image_path = 'ocr_capture.png'

        if image_path:
            image = cv2.imread(image_path)

            labels, boxes, scores = model_predict_on_mosaic(image, self.model_inst, threshold=self.nms_threshold)

            # Overlay boxes and labels on the image
            img_with_boxes = draw_detection_boxes(image, labels, boxes, scores, minscore=minscore)

            # Save the image with overlaid boxes and labels
            # temp_dir = tempfile.gettempdir()
            output_image_path = "result_image.png"
            cv2.imwrite(output_image_path, img_with_boxes)

            # Open the saved image using the default image viewer
            os.startfile(output_image_path)

    # endregion

    # region Settings and Configuration

    def set_instrument_reader_settings(self, rs):
        self.instrument_reader_settings = rs
        self.save_attributes()
        print('instrument reader settings: ', rs)

    def set_reader_settings(self, rs):
        self.reader_settings = rs
        self.save_attributes()
        print('reader settings: ', rs)

    def set_reader_size_and_stride(self):
        self.reader_sub_img_size = tk.simpledialog.askinteger('OCR Sub image size',
                                                              'Provide the Sub image size for the OCR reader when doing an insturment capture: ',
                                                              initialvalue=self.reader_sub_img_size)
        self.reader_stride = tk.simpledialog.askinteger('OCR Sub image stride',
                                                        'Provide the image window stride for the OCR reader when doing an insturment capture: ',
                                                        initialvalue=self.reader_stride)

    def set_local_ocr(self):
        self.do_local_ocr = tk.messagebox.askyesno('Do local ocr on caputure?','Do local ocr on caputure?')

    def set_ocr_object_interference_threshold(self):
        self.filter_ocr_threshold = tk.simpledialog.askfloat('Enter a overlap threshold', 'Enter a threshold to remove ocr items that overlap with a detected object box. Useful for reduceing comment garbage (1.0 removes stuff completely within only and 0.1 removes items that are barely touching): ', initialvalue=self.filter_ocr_threshold)

    def open_instrument_reader_settings(self):
        # Create a new window for ObjectDetectionApp
        img_path = 'temp/instrument_capture.png'
        if os.path.exists(img_path):
            reader_settings_root = tk.Toplevel()
            SetReaderSettings(reader_settings_root, img_path, self.reader,
                              reader_settings=self.instrument_reader_settings,
                              callback=self.set_instrument_reader_settings)
        else:
            print('first capture an instrument')

    def open_general_reader_settings(self):
        # Create a new window for ObjectDetectionApp
        img_path = 'temp/ocr_capture.png'
        if os.path.exists(img_path):
            reader_settings_root = tk.Toplevel()
            SetReaderSettings(reader_settings_root, img_path, self.reader,
                              reader_settings=self.reader_settings,
                              callback=self.set_reader_settings)
        else:
            print('first capture a region')

    def set_tag_label_groups(self):
        group_window = tk.Toplevel(self.root)
        print(self.tag_label_groups)
        app = DictionaryBuilder(group_window, self.tag_label_groups, self.detection_labels)
        self.tag_label_groups = app.run()
        print(self.tag_label_groups)

    def set_nms_threshold(self):
        response = tkinter.simpledialog.askfloat(title='Non-Maximum-Supression',
                                                 prompt='Set NMS overlap threshold (smaller = fewer boxes): ',
                                                 initialvalue=self.nms_threshold)
        try:
            self.nms_threshold = float(response)
        except Exception as e:
            tk.messagebox.showinfo('set NMS fail', e)

    def set_object_box_expand(self):
        response = tkinter.simpledialog.askfloat(title='Set percent box expand',
                                                 prompt='Enter the % box expand for group inst. Use a lower value (i.e 0.8) to focus ocr on instrument contents: ',
                                                 initialvalue=self.object_box_expand)
        try:
            self.object_box_expand = float(response)
        except Exception as e:
            tk.messagebox.showinfo('set box expand fail', e)

    def set_object_scores(self):
        def set_minscores(score_dict):
            self.min_scores = score_dict

        slider_window = tk.Toplevel(self.root)

        for label in self.detection_labels:
            if self.min_scores.get(label):
                continue
            else:
                self.min_scores[label] = self.default_min_detection_score

        SliderApp(slider_window, self.min_scores, callback=set_minscores)

    def categorize_labels(self):
        def set_group_callback(x, y):
            self.group_inst, self.group_other = x, y

        manager = GroupManager(self.detection_labels, self.group_inst, self.group_other, callback=set_group_callback)
        manager.run()
        print(self.group_other)

    def set_comment_box_expand(self):
        self.comment_box_expand = tk.simpledialog.askinteger(prompt="Enter Comment Box Expand",
                                                                      title="Enter Comment Box Expand",
                                                                      initialvalue=self.comment_box_expand)

    def set_re_line(self):
        example = r'.*\"-[A-Z]{1,5}-[A-Z\d]{3,5}-.*'
        new_value = tk.simpledialog.askstring(
            "Line Regular Expression",
            f"Enter the Line Regular Expression or leave empty for no capture\nExample {example}:",
            initialvalue=self.re_line
        )
        if new_value is not None:  # Only update if not cancelled
            self.re_line = new_value

    def get_hough_canny_params(self):
        def set_canny_and_hough_params(cp, hp, ep):
            self.canny_params = vars(cp)
            self.hough_params = vars(hp)
            self.extension_params = vars(ep)
            print('canny params: ', self.canny_params)
            print('hough params: ', self.hough_params)
            print('extension params: ', self.extension_params)


        hough_app_root = tk.Toplevel()
        app = HoughLinesApp(
            hough_app_root,
            params_callback=set_canny_and_hough_params,
            image=self.cropped_image,
            canny_params=self.canny_params,
            hough_params=self.hough_params,
            extension_params=self.extension_params
        )

    def set_line_params(self):
        line_params_window = tk.Toplevel()
        line_params_window.title("Line Parameters")

        # Dictionary of parameters: {param_name: (variable_name, current_value, type_converter, optional_tuple_length)}
        parameters = {
            "Join Threshold": ("line_join_threshold", self.line_join_threshold, int),
            "Line Thickness": ("paint_line_thickness", self.paint_line_thickness, int),
            "Line OCR Box Scale": ("line_box_scale", self.line_box_scale, float),
            "Erosion Kernel": ("line_img_erosion", self.line_img_erosion, int),
            "Erosion Iterations": ("line_erosion_iterations", self.line_erosion_iterations, int),
            "Binary Threshold": ("line_img_binary_threshold", self.line_img_binary_threshold, int),
            "Image Scale": ("line_img_scale", self.line_img_scale, float),
            "Simple Line Mode": ("simple_line_mode", self.simple_line_mode, bool),
            "Debug Line": ("debug_line", self.debug_line, bool),
            "Remove Significant Lines Only": (
            "remove_significant_lines_only", self.remove_significant_lines_only, bool),
            "Remove Text Before Filling": ("remove_text_before", self.remove_text_before, bool),
            "Minimum Confidence to Remove Text": ("text_min_score", self.text_min_score, float),
            "White out Color": ("white_out_color", self.white_out_color, (tuple, 3))
            # Specify tuple length of 3 for RGB
        }

        # Container frame for better organization
        frame = tk.Frame(line_params_window, padx=10, pady=5)
        frame.pack(expand=True, fill='both')

        # Dictionary to store variable objects
        param_vars = {}

        def create_tuple_entry(parent, attr_name, current_value, tuple_length):
            """Create multiple entries for tuple values"""
            tuple_frame = tk.Frame(parent)
            tuple_frame.pack(side='right', padx=5)

            # Create an entry for each tuple element
            tuple_vars = []
            for i in range(tuple_length):
                var = tk.StringVar(value=str(current_value[i]) if i < len(current_value) else "0")
                entry = tk.Entry(tuple_frame, textvariable=var, width=5)
                entry.pack(side='left', padx=2)
                tuple_vars.append(var)

            return tuple_vars

        def validate_tuple_input(tuple_vars, converter):
            """Validate and convert tuple input values"""
            try:
                return tuple(converter(var.get()) for var in tuple_vars)
            except ValueError:
                raise ValueError(f"Invalid tuple values: {[var.get() for var in tuple_vars]}")

        # Create input fields for each parameter
        for label_text, param_info in parameters.items():
            # Unpack parameter info
            attr_name = param_info[0]
            current_value = param_info[1]
            converter = param_info[2]

            # Parameter container
            param_frame = tk.Frame(frame)
            param_frame.pack(fill='x', pady=5)

            # Label
            label = tk.Label(param_frame, text=f"{label_text}:")
            label.pack(side='left', padx=5)

            if isinstance(converter, tuple) and converter[0] == tuple:
                # Handle tuple parameters
                tuple_length = converter[1]
                param_vars[attr_name] = create_tuple_entry(param_frame, attr_name, current_value, tuple_length)
            elif converter == bool:
                # Checkbutton for boolean parameters
                param_vars[attr_name] = tk.BooleanVar(value=current_value)
                checkbox = tk.Checkbutton(param_frame, variable=param_vars[attr_name])
                checkbox.pack(side='right', padx=5)
            else:
                # Entry with StringVar for numeric parameters
                param_vars[attr_name] = tk.StringVar(value=str(current_value))
                entry = tk.Entry(param_frame, textvariable=param_vars[attr_name])
                entry.pack(side='right', padx=5, expand=True, fill='x')

        def apply_values():
            try:
                # Update all parameters
                for label_text, param_info in parameters.items():
                    attr_name, current_value, converter = param_info[:3]
                    var = param_vars[attr_name]

                    if isinstance(converter, tuple) and converter[0] == tuple:
                        # Handle tuple parameters
                        value = validate_tuple_input(var, int)  # Assuming RGB values are integers
                    elif converter == bool:
                        value = var.get()
                    else:
                        value = converter(var.get())

                    setattr(self, attr_name, value)

                line_params_window.destroy()
            except ValueError as e:
                messagebox.showerror("Error", f"Invalid input: {str(e)}")

        # Button container
        button_frame = tk.Frame(frame)
        button_frame.pack(fill='x', pady=10)

        # Apply button
        apply_button = tk.Button(button_frame, text="Apply", command=apply_values)
        apply_button.pack(expand=True)

        # Cancel button
        cancel_button = tk.Button(button_frame, text="Cancel", command=line_params_window.destroy)
        cancel_button.pack(expand=True, padx=5)

    def set_association_radius(self):
        self.association_radius = tk.simpledialog.askfloat(prompt="Enter Object association radius",
                                                           title="Enter association ( Radius",
                                                           initialvalue=self.association_radius)

    # endregion

    # region File Operations

    def save_workbook(self):
        # Save the workbook
        if self.write_mode == 'openpyxl':
            self.wb.save(self.workbook_path)
            print('wb saved')
        else:
            print('not in openpyxl write mode. cant save')

    def merge_pdfs(self):
        folder_that_has_pdfs = filedialog.askdirectory(title='Folder that has PDFs')
        merge_pdf(folder_that_has_pdfs)

    def make_pid_page_xlsx(self):

        pid_page_xlsx = openpyxl.Workbook()
        ws = pid_page_xlsx.create_sheet('pid_page')
        image_folder = filedialog.askdirectory(title='Folder that has PNGs')
        if self.pid_coords:

            for filename in os.listdir(image_folder):
                if filename.endswith(".png") or filename.endswith(".jpg"):
                    file_path = os.path.join(image_folder, filename)
                    img = cv2.imread(file_path)

                    self.cropped_x1, self.cropped_y1, self.cropped_x2, self.cropped_y2 = self.pid_coords
                    cropped_image = img[self.cropped_y1:self.cropped_y2, self.cropped_x1:self.cropped_x2]
                    self.capture_pid(cropped_image)

                    row = [filename, self.pid if self.pid else ""]
                    ws.append(row)

        counter = 1
        while True:
            try:
                save_location = os.path.join(image_folder, f"pid_page_{counter}.xlsx")
                pid_page_xlsx.save(save_location)
                print(f"File saved as: {save_location}")
                break
            except Exception as e:
                print(f"Error saving file: {e}")
                counter += 1

    def generate_instrument_count(self):
        ocr_needed = tk.messagebox.askyesno(
            title='OCR NEEDED?',
            message='Do we need to do OCR (for comments)?\nNote if available we use existing results'
        )

        overwrite = tk.messagebox.askyesno(
            title='OVERWRITE?',
            message='Do you want to overwrite Existing counts?'
        )

        sure = tk.messagebox.askyesno(
            title='SURE?',
            message='ARE U SURE'
        )

        if not sure:
            return

        # Create progress window
        progress_window = ProgressWindow(self.root, len(self.image_list))

        if ocr_needed:
            self.get_all_ocr()

        rel = self.re_line
        slm = self.simple_line_mode
        dbl = self.debug_line
        dlo = self.do_local_ocr

        self.re_line = None
        self.simple_line_mode = True
        self.do_local_ocr = False
        self.debug_line = False

        try:
            for i in range(len(self.image_list)):
                if progress_window.cancelled:
                    break

                self.go_to_page(i)
                self.one_instrument_count(overwrite=overwrite)

                # Update progress
                progress_window.update(i + 1)

            if not progress_window.cancelled:
                self.compile_excels()
        finally:
            # Restore original settings
            self.re_line = rel
            self.simple_line_mode = slm
            self.debug_line = dbl
            self.do_local_ocr = dlo
            progress_window.destroy()

    def compile_excels(self):
        # NOW we compile all the xlsxs into one
        output_file = compile_excels(self.folder_path, self.folder_path, prefix='Instrument_Count', timestamp=True,
                                     recursive=True)
        if tk.messagebox.askyesno('Open Results?', 'Open the Instrument Count Report?'):
            os.startfile(output_file)

    def one_instrument_count(self, overwrite=True):
        """
        Generate instrument count for a single page by leveraging the capture_instruments method.

        Args:
            overwrite (bool): Whether to overwrite existing count files. Defaults to True.
        """
        save_location = os.path.join(self.results_folder, "Instrument_Count.xlsx")
        if os.path.exists(save_location) and not overwrite:
            return

        inst_count_xlsx = openpyxl.Workbook()
        ws = inst_count_xlsx.active
        ws.title = 'Instrument Count'

        print('workbook, ', inst_count_xlsx)

        # Clear current state
        self.inst_data = []
        self.persistent_boxes = []
        self.active_inst_box_count = 0

        # Use capture_instruments with full image
        self.cropped_x1, self.cropped_y1 = 0, 0
        self.cropped_x2, self.cropped_y2 = self.width, self.height

        self.capture_instruments(self.cv2img)


        # Write data to Excel
        current_row = 1
        for data in self.inst_data:
            data['pid'] = self.pid
            data['file'] = self.image_path

            if not ws['A1'].value:  # Check if header doesn't exist
                print('making header')
                self.create_excel_header(ws, data)
                current_row += 1

            self.populate_excel_row(ws, data, current_row)
            current_row += 1

        try:
            inst_count_xlsx.save(save_location)
            print(f"File saved as: {save_location}")
        except Exception as e:
            print(f"Save to {save_location} failed: {e}")

    # endregion

    # region UI Helper Methods

    def shift_pressed(self, event):
        self.shift_held = True

    def shift_released(self, event):
        self.shift_held = False

    def ctrl_pressed(self, event):
        self.ctrl_held = True

    def ctrl_released(self, event):
        self.ctrl_held = False

    def show_keybindings(self):
        keybindings = """
            n/N: Next image
            b/B: Previous image
            p/P: Capture PID
            a/A: Capture Instrument Group
            f/F: Capture Line
            e/E: Capture Equipment
            z/Z: Capture Service In
            x/X: Capture Service Out
            w/W: Write Data to Excel
            c/C: Clear Instrument Group
            v/V: Vote to normalize tag numbers
            s/S: Swap Services
            g/G: Set Comment
            left shift: add to service with formatting
            left ctrl: add to service
            """
        tk.messagebox.showinfo("Keybindings", keybindings)

    def show_labels(self):

        tk.messagebox.showinfo("object recognition labels", self.detection_labels)

    def open_console(self):
        self.console_popup.create_console_popup()

    def open_image_path(self):
        os.startfile(self.image_path)

    def open_project_folder(self):
        os.startfile(self.folder_path)

    def open_page_results(self):
        os.startfile(self.results_folder)

    def open_detecto_gui(self):
        detecto_gui_window = tk.Toplevel(self.root)
        ObjectDetectionApp(detecto_gui_window)

    def open_FAIA(self):
        faia_window = tk.Toplevel(self.root)
        data_puller_window = tk.Toplevel(self.root)

        ExcelDataPullApp(data_puller_window)
        FindAnInstrumentApp(faia_window, img_path=self.image_path)

    def open_image_editor(self):
        image_editor_window = tk.Toplevel(self.root)
        # Pass current image path instead of static filename
        ImageEditor(image_editor_window, folder=self.folder_path, image=self.image_path)

    def open_ocr_results_viewer(self):
        ocr_viewer_window = tk.Toplevel(self.root)
        OCRViewer(ocr_viewer_window, self.folder_path)

    def open_workbook(self):
        os.startfile(self.workbook_path)

    def save_page_data(self):
        """Save the page data to a pickle file"""
        if not self.folder_path:
            tk.messagebox.showerror("Error", "No project folder loaded")
            return
            
        try:
            save_path = os.path.join(self.folder_path, "page_data.pkl")
            with open(save_path, 'wb') as f:
                pickle.dump(self.page_data, f)
            tk.messagebox.showinfo("Success", f"Page data saved to {save_path}")
        except Exception as e:
            tk.messagebox.showerror("Error", f"Failed to save page data: {str(e)}")

    def load_page_data(self):
        """Load the page data from a pickle file"""
        if not self.folder_path:
            tk.messagebox.showerror("Error", "No project folder loaded")
            return
            
        load_path = os.path.join(self.folder_path, "page_data.pkl")
        if not os.path.exists(load_path):
            tk.messagebox.showerror("Error", "No saved page data found")
            return
            
        try:
            with open(load_path, 'rb') as f:
                self.page_data = pickle.load(f)
            
            # Recreate boxes for current page if it exists in loaded data
            if self.image_path in self.page_data:
                self.recreate_boxes()
                
            tk.messagebox.showinfo("Success", "Page data loaded successfully")
        except Exception as e:
            tk.messagebox.showerror("Error", f"Failed to load page data: {str(e)}")

    # endregion

    # region License Managment

    def activate_license(self):
        """Show license activation dialog"""
        # Create a custom dialog with better formatting
        activation_dialog = tk.Toplevel(self.root)
        activation_dialog.title("License Activation")
        activation_dialog.geometry("500x200")
        activation_dialog.resizable(False, False)
        
        # Make it modal
        activation_dialog.transient(self.root)
        activation_dialog.grab_set()
        
        # Add instructions
        tk.Label(activation_dialog, text="Enter your license key:", font=("Arial", 12)).pack(pady=(15, 5))
        
        # Add entry field with more space
        license_entry = tk.Entry(activation_dialog, width=50)
        license_entry.pack(pady=5, padx=20)
        
        # Add activate button
        def do_activate():
            license_key = license_entry.get().strip()
            if license_key:
                success, message = self.license_manager.activate_license(license_key)
                if success:
                    tk.messagebox.showinfo("Success", message)
                    activation_dialog.destroy()
                else:
                    tk.messagebox.showerror("Error", message)
            else:
                tk.messagebox.showerror("Error", "Please enter a license key")
        
        # Add buttons for activation and cancel
        button_frame = tk.Frame(activation_dialog)
        button_frame.pack(pady=15)
        
        tk.Button(button_frame, text="Activate", command=do_activate, width=10).pack(side=tk.LEFT, padx=10)
        tk.Button(button_frame, text="Cancel", command=activation_dialog.destroy, width=10).pack(side=tk.LEFT, padx=10)
        
        # Add info about purchase
        tk.Label(activation_dialog, text="Don't have a license? Purchase one from our website.",
                 font=("Arial", 9), fg="blue", cursor="hand2").pack(pady=10)
        
        # Center the dialog on the parent window
        activation_dialog.update_idletasks()
        width = activation_dialog.winfo_width()
        height = activation_dialog.winfo_height()
        x = (self.root.winfo_width() // 2) - (width // 2) + self.root.winfo_x()
        y = (self.root.winfo_height() // 2) - (height // 2) + self.root.winfo_y()
        activation_dialog.geometry(f"+{x}+{y}")
        
        # Wait for dialog to close
        self.root.wait_window(activation_dialog)


    def check_license_status(self):
        """Show current license status"""
        is_licensed, status = self.license_manager.check_license()
        tk.messagebox.showinfo("License Status", f"Current Status: {status}")


    def open_purchase_page(self):
        """Open the license purchase page in a browser"""
        import webbrowser
        webbrowser.open("https://pidvision.com")
        tk.messagebox.showinfo(
            "Purchase License", 
            "The license purchase page has been opened in your browser.\n\n"
            "After completing your purchase, you will receive a license key by email.\n"
            "You can then activate your license using the 'Activate License' option."
        )
    # endregion

    def save_current_group(self):
        """Save the current capture group to page_data"""
        # Get current data from window
        self.get_data_from_window()
        
        # Update current capture group with window data
        self.current_capture_group.pid = self.pid
        self.current_capture_group.line = self.line
        self.current_capture_group.service_in = self.service_in
        self.current_capture_group.service_out = self.service_out
        self.current_capture_group.equipment = self.equipment
        self.current_capture_group.comment = self.comment
        self.current_capture_group.file = self.image_path

        # Create CaptureBox for each instrument group
        if self.inst_data:
            # Get all orange boxes from persistent_boxes that are rectangles with orange outline
            # and haven't been saved before
            orange_boxes = []
            for box_id in self.persistent_boxes:
                if box_id not in self.saved_boxes and self.canvas.type(box_id) == "rectangle":
                    outline_color = self.canvas.itemcget(box_id, "outline")
                    if outline_color == "orange":
                        bbox = self.canvas.bbox(box_id)
                        if bbox:
                            # Convert canvas coordinates back to image coordinates
                            x1, y1 = self.canvas_to_image(bbox[0], bbox[1])
                            x2, y2 = self.canvas_to_image(bbox[2], bbox[3])
                            orange_boxes.append((box_id, (x1, y1, x2, y2)))
            
            # Create a CaptureBox for each new orange box
            for box_id, group_coords in orange_boxes:
                # Find instrument data associated with this box
                box_inst_data = []
                for inst in self.inst_data:
                    # Check if the instrument's visual elements are not in saved boxes
                    if ('visual_elements' in inst and 
                        inst['visual_elements']['box'] not in self.saved_boxes):
                        box_inst_data.append(inst.copy())
                
                box = CaptureBox(
                    'instrument_group',
                    group_coords,
                    box_inst_data
                )
                self.current_capture_group.add_box(box)
                
                # Mark this box and its associated visual elements as saved
                self.saved_boxes.add(box_id)
                # Change only the outline to green to indicate it's saved
                self.canvas.itemconfig(box_id, outline='#90EE90')
                
                for inst in box_inst_data:
                    if 'visual_elements' in inst:
                        self.saved_boxes.add(inst['visual_elements']['box'])
                        self.saved_boxes.add(inst['visual_elements']['text'])

        # Initialize list for this page if it doesn't exist
        if self.image_path not in self.page_data:
            self.page_data[self.image_path] = []
            
        # Only append if we have new boxes
        if self.current_capture_group.boxes:
            self.page_data[self.image_path].append(self.current_capture_group)
            
            # Keep the boxes visible but mark them as saved with green outline only
            for box in self.persistent_boxes[-self.active_inst_box_count:]:
                if self.canvas.type(box) == "rectangle":
                    self.canvas.itemconfig(box, outline='#90EE90')
            
            # Create new capture group for next captures
            self.current_capture_group = CaptureGroup()
            
            # Clear current data but keep the boxes visible
            self.inst_data = []
            self.active_inst_box_count = 0
            
            # Update display
            self.update_data_display()
            
            print(f"Saved new capture group for {self.image_path}")
        else:
            print("No new boxes to save")

    def recreate_boxes(self):
        """Recreate visual elements for all capture groups on a page"""
        if self.image_path not in self.page_data:
            return
            
        capture_groups = self.page_data[self.image_path]
        
        # Clear existing boxes and tracking sets
        self.clear_boxes()
        self.inst_data = []
        self.saved_boxes.clear()
        
        # Set flag for box recreation
        self.recreating_boxes = True
        
        try:
            # Recreate boxes for each capture group
            for capture_group in capture_groups:
                # Restore page data from the last capture group (most recent)
                self.pid = capture_group.pid
                self.line = capture_group.line
                self.service_in = capture_group.service_in
                self.service_out = capture_group.service_out
                self.equipment = capture_group.equipment
                self.comment = capture_group.comment
                
                # Recreate boxes for each capture in the group
                for box in capture_group.boxes:
                    if box.box_type == 'instrument_group':
                        # Set crop coordinates for visual element creation
                        self.cropped_x1, self.cropped_y1, self.cropped_x2, self.cropped_y2 = box.coordinates
                        
                        # Create the orange group box with green outline to indicate it's saved
                        x1, y1 = self.canvas_coords_from_image(self.cropped_x1, self.cropped_y1)
                        x2, y2 = self.canvas_coords_from_image(self.cropped_x2, self.cropped_y2)
                        orange_box = self.canvas.create_rectangle(
                            x1, y1, x2, y2,
                            outline='#90EE90',  # Light green outline
                            fill='#90EE90',     # Light green fill
                            stipple='gray12'    # Keep stipple pattern
                        )
                        self.persistent_boxes.append(orange_box)
                        self.saved_boxes.add(orange_box)  # Track this box as saved
                        
                        # Extend instrument data
                        self.inst_data.extend(box.data)
                        
                        # Get labels, boxes, scores from instrument data
                        labels = []
                        boxes = []
                        scores = []
                        for inst in box.data:
                            if 'label' in inst and 'box' in inst and 'score' in inst:
                                labels.append(inst['label'])
                                boxes.append(inst['box'])
                                scores.append(inst['score'])
                        
                        if labels and boxes and scores:
                            # Process visual elements
                            visual_elements_list, new_active_boxes = self.process_visual_elements(
                                labels, boxes, scores, (self.cropped_x1, self.cropped_y1)
                            )
                            
                            # Update persistent boxes and box count
                            self.persistent_boxes.extend(new_active_boxes)
                            self.active_inst_box_count += len(new_active_boxes)
                            
                            # Add new boxes to saved set
                            self.saved_boxes.update(new_active_boxes)
                            
                            # Update visual elements in instrument data
                            for inst_data, visual_elements in zip(box.data[-len(visual_elements_list):], visual_elements_list):
                                inst_data['visual_elements'] = visual_elements

                        # Add group annotation
                        annotation = self.create_group_annotation(capture_group, x1, y1)
                        self.persistent_boxes.append(annotation)
                        self.saved_boxes.add(annotation)
        finally:
            # Always clear the flag when done
            self.recreating_boxes = False
        
        # Update display
        self.update_data_display()

    def create_group_annotation(self, group, x1, y1):
        """Create a text annotation showing group data"""
        # Format the annotation text
        annotation_lines = []
        if group.pid:
            annotation_lines.append(f"PID: {group.pid}")
        if group.line:
            annotation_lines.append(f"Line: {group.line}")
        if group.service_in:
            annotation_lines.append(f"Service In: {group.service_in}")
        if group.service_out:
            annotation_lines.append(f"Service Out: {group.service_out}")
        if group.comment:
            annotation_lines.append(f"Comment: {group.comment}")
        
        annotation_text = '\n'.join(annotation_lines)
        
        # Create text annotation on canvas with bold font
        text_id = self.canvas.create_text(
            x1, y1 - 10,  # Position above the box
            text=annotation_text,
            anchor='sw',
            fill='#90CC90',
            font=('Arial', 9, 'bold'),  # Set font to bold
            justify='left',
            state='hidden'  # Add this line to make annotations start hidden
        )
        
        # Add to annotations list
        self.annotations.append(text_id)
        
        return text_id

    def handle_right_click(self, event):
        """Handle initial right click - start both potential pan and edit"""
        self.right_click_start = (event.x, event.y)
        self.right_click_moved = False
        self.canvas.scan_mark(event.x, event.y)  # Start pan

    def handle_right_motion(self, event):
        """Handle right button motion - pan the canvas"""
        if self.right_click_start:
            self.right_click_moved = True
            self.canvas.scan_dragto(event.x, event.y, gain=1)
            self.show_image()

    def handle_right_release(self, event):
        """Handle right click release - edit if no movement occurred"""
        if not self.right_click_moved and self.right_click_start:
            # Convert canvas coordinates to image coordinates
            x = self.canvas.canvasx(event.x)
            y = self.canvas.canvasy(event.y)
            img_x, img_y = self.canvas_to_image(x, y)
            
            # Check if click is inside any saved group
            if self.image_path in self.page_data:
                for group in self.page_data[self.image_path]:
                    for box in group.boxes:
                        if box.box_type == 'instrument_group':
                            x1, y1, x2, y2 = box.coordinates
                            if x1 <= img_x <= x2 and y1 <= img_y <= y2:
                                self.edit_group(group)
                                break
    
        self.right_click_start = None
        self.right_click_moved = False

    def edit_group(self, group):
        """Create popup window for editing group data"""
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Edit Group")
        edit_window.bind('<Return>', lambda e: save_changes())
        
        # Create and pack a frame for better layout
        frame = ttk.Frame(edit_window, padding="10")
        frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Create entry fields with labels
        fields = [
            ("PID:", group.pid),
            ("Line:", group.line),
            ("Service In:", group.service_in),
            ("Service Out:", group.service_out),
            ("Equipment:", group.equipment),
            ("Comment:", group.comment)
        ]
        
        entries = {}
        for i, (label_text, value) in enumerate(fields):
            ttk.Label(frame, text=label_text).grid(row=i, column=0, padx=5, pady=5, sticky=tk.W)
            entry = ttk.Entry(frame, width=40)
            if value:
                entry.insert(0, value)
            entry.grid(row=i, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))
            entries[label_text] = entry
        
        def save_changes():
            # Update group data
            group.pid = entries["PID:"].get()
            group.line = entries["Line:"].get()
            group.service_in = entries["Service In:"].get()
            group.service_out = entries["Service Out:"].get()
            group.equipment = entries["Equipment:"].get()
            group.comment = entries["Comment:"].get()
            
            # Update display if this is the most recent group
            if self.page_data[self.image_path][-1] == group:
                self.pid = group.pid
                self.line = group.line
                self.service_in = group.service_in
                self.service_out = group.service_out
                self.equipment = group.equipment
                self.comment = group.comment
                self.update_data_display()
            
            edit_window.destroy()
        
        def delete_group():
            """Delete the group and its associated visual elements"""
            if tk.messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this group?"):
                # Remove visual elements from canvas
                for box in group.boxes:
                    if box.box_type == 'instrument_group':
                        # Remove the orange/green group box
                        x1, y1 = self.canvas_coords_from_image(box.coordinates[0], box.coordinates[1])
                        x2, y2 = self.canvas_coords_from_image(box.coordinates[2], box.coordinates[3])
                        overlapping = self.canvas.find_overlapping(x1, y1, x2, y2)
                        for item_id in overlapping:
                            if item_id in self.persistent_boxes:
                                self.canvas.delete(item_id)
                                self.persistent_boxes.remove(item_id)
                                if item_id in self.saved_boxes:
                                    self.saved_boxes.remove(item_id)
                    
                    # Remove associated instrument boxes and labels
                    for inst_data in box.data:
                        if 'visual_elements' in inst_data:
                            box_id = inst_data['visual_elements']['box']
                            text_id = inst_data['visual_elements']['text']
                            self.canvas.delete(box_id)
                            self.canvas.delete(text_id)
                            if box_id in self.persistent_boxes:
                                self.persistent_boxes.remove(box_id)
                            if text_id in self.persistent_boxes:
                                self.persistent_boxes.remove(text_id)
                            if box_id in self.saved_boxes:
                                self.saved_boxes.remove(box_id)
                            if text_id in self.saved_boxes:
                                self.saved_boxes.remove(text_id)

                # Remove group from page_data
                if self.image_path in self.page_data:
                    self.page_data[self.image_path].remove(group)
                    
                    # Update display with last group's data if available
                    if self.page_data[self.image_path]:
                        last_group = self.page_data[self.image_path][-1]
                        self.pid = last_group.pid
                        self.line = last_group.line
                        self.service_in = last_group.service_in
                        self.service_out = last_group.service_out
                        self.equipment = last_group.equipment
                        self.comment = last_group.comment
                    else:
                        # Clear current data if no groups remain
                        self.pid = None
                        self.line = None
                        self.service_in = None
                        self.service_out = None
                        self.equipment = None
                        self.comment = None
                    
                    self.update_data_display()
                
                edit_window.destroy()
        
        # Create button frame
        button_frame = ttk.Frame(frame)
        button_frame.grid(row=len(fields), column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="Save", command=save_changes).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Delete", command=delete_group, style='Delete.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=edit_window.destroy).pack(side=tk.LEFT, padx=5)
        
        # Create a style for the delete button
        style = ttk.Style()
        style.configure('Delete.TButton', foreground='red')
        
        # Configure grid weights
        frame.columnconfigure(1, weight=1)
        
        # Make dialog modal
        edit_window.transient(self.root)
        edit_window.grab_set()
        
        # Center the window
        edit_window.update_idletasks()
        width = edit_window.winfo_width()
        height = edit_window.winfo_height()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (width // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (height // 2)
        edit_window.geometry(f'+{x}+{y}')

    def generate_page_data_report(self):
        """Generate an Excel report from the saved page data"""
        if not self.page_data:
            tk.messagebox.showerror("Error", "No page data available")
            return

        try:
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Page Data Report'

            # Define headers
            headers = ['File', 'PID', 'Line', 'Service In', 'Service Out', 'Equipment', 'Comment', 
                      'Instrument Tag', 'Instrument Tag No', 'Instrument Type', 'Instrument Line', 'Instrument Comment']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Current row for writing data
            current_row = 2

            # Process each page's data
            for file_path, capture_groups in self.page_data.items():
                for group in capture_groups:
                    # Get base data that will be repeated for each instrument
                    base_data = {
                        'File': file_path,
                        'PID': group.pid,
                        'Line': group.line,
                        'Service In': group.service_in,
                        'Service Out': group.service_out,
                        'Equipment': group.equipment,
                        'Comment': group.comment
                    }

                    # Find instrument data in the group's boxes
                    instruments_found = False
                    for box in group.boxes:
                        if box.box_type == 'instrument_group' and box.data:
                            for inst in box.data:
                                # Combine base data with instrument data
                                row_data = base_data.copy()
                                row_data.update({
                                    'Instrument Tag': inst.get('tag', ''),
                                    'Instrument Tag No': inst.get('tag_no', ''),
                                    'Instrument Type': inst.get('type', ''),
                                    'Instrument Line': inst.get('line', ''),
                                    'Instrument Comment': inst.get('comment', '')
                                })

                                # Write row to worksheet
                                for col, header in enumerate(headers, 1):
                                    ws.cell(row=current_row, column=col, value=row_data.get(header, ''))
                                current_row += 1
                                instruments_found = True

                    # If no instruments in group, write a row with just the base data
                    if not instruments_found:
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=current_row, column=col, value=base_data.get(header, ''))
                        current_row += 1

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # Save the workbook
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            save_path = os.path.join(self.folder_path, f"page_data_report_{timestamp}.xlsx")
            wb.save(save_path)

            # Ask if user wants to open the file
            if tk.messagebox.askyesno("Success", f"Report saved to {save_path}\nWould you like to open it?"):
                os.startfile(save_path)

        except Exception as e:
            tk.messagebox.showerror("Error", f"Failed to generate report: {str(e)}")

    def toggle_annotations(self):
        """Toggle visibility of group annotations"""
        self.annotations_visible = not self.annotations_visible
        self.config.annotations_visible = self.annotations_visible
        for annotation in self.annotations:
            if self.annotations_visible:
                self.canvas.itemconfig(annotation, state='normal')
            else:
                self.canvas.itemconfig(annotation, state='hidden')

    def open_text_corrections(self):
        TextCorrectionsEditor(self.root, self.text_corrections)


def set_window_logo(window, png_path, size=(64, 64)):
    """
    Set a PNG image as the window logo.

    Args:
        window: Tkinter window object
        png_path (str): Path to the PNG file
        size (tuple): Desired icon size (width, height)
    """
    try:
        # Open and resize the PNG
        img = Image.open(png_path)
        img = img.resize(size, Image.Resampling.LANCZOS)

        # Save as temporary ICO
        img.save('temp_icon.ico', format='ICO')

        # Set the window icon
        window.iconbitmap('temp_icon.ico')

        # Clean up
        import os
        os.remove('temp_icon.ico')

    except Exception as e:
        print(f"Error setting icon: {e}")


if __name__ == "__main__":
    root = tk.Tk()

    png_path = r"LOGO.png"
    set_window_logo(root, png_path)

    app = PIDVisionApp(root)
    root.mainloop()
