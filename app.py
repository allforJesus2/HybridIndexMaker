from PIL import Image, ImageTk
from console_redirect import *
from convolutioner import ConvolutionReplacer
from detecto.core import Model
from detecto_gui import ObjectDetectionApp
from easyocr_mosaic import *
from find_an_instrument import FindAnInstrumentApp
from functions import *
from group_manager import GroupManager
from image_editor import ImageEditor
from minscore_edit import SliderApp
from model_predict_mosaic import *
from multi_window_dictionary_maker import DictionaryBuilder
from ocr_results_viewer import OCRViewer
from reader_settings import SetReaderSettings
from tkinter import filedialog, ttk
from auto_scroll import AutoScrollbar
import easyocr
import importlib.util
import json
import openpyxl
import os
import re
import threading
import tkinter as tk
import tkinter.messagebox
import tkinter.simpledialog
import xlwings as xw


class PIDVisionApp:

    # region Initialization and Setup

    def __init__(self, root):
        # Initialize root window
        self.root = root

        # Hide the main window initially
        self.root.withdraw()
        # Create and show splash screen

        self.splash = SplashScreen(root, r"C:\Users\dcaoili\OneDrive - Samuel Engineering\Pictures\logo\logo-big.png")

        # Start initialization in a separate thread
        init_thread = threading.Thread(target=self.initialize_app)
        init_thread.start()

        # Check if initialization is complete
        self.check_initialization(init_thread)

    def initialize_app(self):

        # Create menu bar
        self.create_menu_bar()

        # Initialize console redirects
        self.console_popup = ConsolePopup(self.root)

        # Initialize canvas and scrollbars
        self.create_canvas_and_scrollbars()

        # Initialize image attributes
        self.initialize_image_attributes()

        # Initialize instrument and equipment models
        self.initialize_models()

        # Initialize reader settings
        self.initialize_reader_settings()

        # Initialize other attributes
        self.initialize_other_attributes()

        # Bind key shortcuts to the respective commands
        self.bind_key_shortcuts()
        # Create data window
        self.create_data_window()
        # Initialize data display
        self.update_data_display()

        # Initialize captured data
        self.capture_mode = 'pid'

    def check_initialization(self, init_thread):
        """Check if initialization is complete and show main window"""
        if init_thread.is_alive():
            # Check again in 100ms
            self.root.after(100, lambda: self.check_initialization(init_thread))
        else:
            # Initialization complete, show main window
            self.show_main_window()

    def show_main_window(self):
        """Show the main window and destroy splash screen"""
        self.splash.destroy()
        self.root.deiconify()
        self.root.title("PIDVision.AI")

    def create_menu_bar(self):
        # Create a menu bar
        self.menu_bar = tk.Menu(self.root)
        self.file_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.file_menu.add_command(label="Open Project Folder", command=self.open_folder)
        self.file_menu.add_command(label="Open Current Image", command=self.open_image_path)
        self.file_menu.add_command(label="Open Index", command=self.open_workbook)

        self.file_menu.add_command(label="Load Object detection model", command=self.load_pretrained_model)

        self.menu_bar.add_cascade(label="File", menu=self.file_menu)

        # Create App menu
        self.app_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.app_menu.add_command(label="Image Editor", command=self.open_image_editor)
        self.app_menu.add_command(label="Find an instrument App", command=self.open_FAIA)
        self.app_menu.add_command(label="Search OCR Results", command=self.open_ocr_results_viewer)
        self.app_menu.add_command(label="Train Object Detection Model", command=self.open_detecto_gui)
        self.app_menu.add_command(label="Open Console Popup", command=self.open_console)

        self.menu_bar.add_cascade(label="Apps", menu=self.app_menu)

        # Create a commands menu
        self.command_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.command_menu.add_command(label="Create images from PDF", command=self.open_pdf2png)
        self.command_menu.add_command(label="Merge pdfs", command=self.merge_pdfs)

        # self.command_menu.add_command(label="Load a Tag Correction Function", command=self.load_correct_fn)
        self.command_menu.add_command(label="Append Data to Index", command=self.append_data_to_excel)
        self.command_menu.add_command(label="Live Write Mode", command=lambda: self.set_write_mode('xlwings'))
        self.command_menu.add_command(label="Silent/quick Write Mode", command=lambda: self.set_write_mode('openpyxl'))
        self.command_menu.add_command(label="Save workbook", command=self.save_workbook)
        # self.command_menu.add_command(label="Auto Generate Index", command=self.auto_generate_index)
        self.command_menu.add_command(label="Generate type xlsx via convolution", command=self.create_tag_type_xlsx)
        self.command_menu.add_command(label="Generate All pages Instrument Count",
                                      command=self.generate_instrument_count)
        self.command_menu.add_command(label="Generate Single Page Instrument Count", command=self.one_instrument_count)
        # self.command_menu.add_command(label="Compile Instrument counts", command=self.compile_excels)
        self.command_menu.add_command(label="Generate Filename PID xlsx", command=self.make_pid_page_xlsx)
        self.command_menu.add_command(label="Test instrument Model on Image", command=self.test_model_mosaic)

        self.command_menu.add_command(label="Get OCR", command=self.get_ocr)
        self.command_menu.add_command(label="Get all pages OCR", command=self.get_all_ocr)
        self.command_menu.add_command(label="Open Page Results folder", command=self.open_page_results)
        # self.command_menu.add_command(label="Print class vairables", command=self.print_class_vars)

        self.menu_bar.add_cascade(label="Commands", menu=self.command_menu)

        # Create a capture menu
        self.capture_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.capture_menu.add_command(label="Capture PID", command=lambda: self.set_capture('pid'))
        self.capture_menu.add_command(label="Capture Instrument Group", command=lambda: self.set_capture('instruments'))
        self.capture_menu.add_command(label="Capture Line", command=lambda: self.set_capture('line'))
        self.capture_menu.add_command(label="Capture Equipment", command=lambda: self.set_capture('equipment'))
        self.capture_menu.add_command(label="Capture Service In", command=lambda: self.set_capture('service_in'))
        self.capture_menu.add_command(label="Capture Service Out", command=lambda: self.set_capture('service_out'))
        self.capture_menu.add_command(label="Capture comment", command=lambda: self.set_capture('comment'))

        self.capture_menu.add_command(label="Swap services", command=self.swap_services)
        self.capture_menu.add_command(label="Clear instrument group", command=self.clear_instrument_group)

        self.menu_bar.add_cascade(label="Capture", menu=self.capture_menu)

        # Create page menu
        self.page_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.page_menu.add_command(label="Next", command=self.next_image)
        self.page_menu.add_command(label="Previous", command=self.previous_image)
        self.page_menu.add_command(label="Go to Page", command=self.open_go_to_page)
        self.menu_bar.add_cascade(label="Page", menu=self.page_menu)

        # settings menu
        self.settings_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.settings_menu.add_command(label="Instrument Reader Settings",
                                       command=self.open_instrument_reader_settings)
        self.settings_menu.add_command(label="General Reader Settings", command=self.open_general_reader_settings)
        self.settings_menu.add_command(label="instrument comment box expand", command=self.set_comment_box_expand)
        self.settings_menu.add_command(label="Tag Prefix Groups", command=self.set_tag_label_groups)
        self.settings_menu.add_command(label="Instrument Groups", command=self.categorize_labels)
        self.settings_menu.add_command(label="Group Association radius", command=self.set_association_radius)
        self.settings_menu.add_command(label="Object Min Scores", command=self.set_object_scores)
        self.settings_menu.add_command(label="object box expand %", command=self.set_object_box_expand)
        self.settings_menu.add_command(label="NMS threshold", command=self.set_nms_threshold)
        self.settings_menu.add_command(label="Save Settings", command=self.save_attributes)

        self.menu_bar.add_cascade(label="Settings", menu=self.settings_menu)

        # Create a Help menu
        self.help_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.help_menu.add_command(label="Keybindings", command=self.show_keybindings)
        self.help_menu.add_command(label="Object detection lables", command=self.show_labels)
        self.menu_bar.add_cascade(label="Help", menu=self.help_menu)

        self.root.config(menu=self.menu_bar)

    def create_canvas_and_scrollbars(self):
        # Vertical and horizontal scrollbars for canvas
        vbar = AutoScrollbar(self.root, orient='vertical')
        hbar = AutoScrollbar(self.root, orient='horizontal')
        vbar.grid(row=0, column=1, sticky='ns')
        hbar.grid(row=1, column=0, sticky='we')

        # Create canvas and put image on it
        self.canvas = tk.Canvas(self.root, highlightthickness=0,
                                xscrollcommand=hbar.set, yscrollcommand=vbar.set)
        self.canvas.grid(row=0, column=0, sticky='nswe')
        self.canvas.update()  # wait till canvas is created
        vbar.configure(command=self.scroll_y)  # bind scrollbars to the canvas
        hbar.configure(command=self.scroll_x)

        # Make the canvas expandable
        self.root.rowconfigure(0, weight=1)
        self.root.columnconfigure(0, weight=1)

        self.crop_rectangle = None
        self.image_container = None

    def initialize_image_attributes(self):
        # Initialize image attributes
        self.image_list = []
        self.image_path = None
        self.current_image_index = 0
        self.original_image = None
        self.imscale = 1.0  # scale for the canvas image
        self.delta = 1.3  # zoom magnitude
        self.cv2img = None

    def initialize_models(self):
        # Initialize instrument and equipment models
        self.model_inst_path = "models/vortex_large.pth"
        # self.model_equipment_path = "models/equipment_services_v2.pth"

        print('loading models from ', self.model_inst_path)
        try:
            # load instrument recognition model
            self.load_pretrained_model(self.model_inst_path)
        except Exception as e:
            print('Error', e)
            print('load model failed. you will likely have to load the model from command')
        '''
        labels = 'tank, pump, service_in, service_out'
        self.labels_equipment = labels.split(', ')
        try:
            # load instrument recognition model
            self.model_equip = Model.load(self.model_equipment_path, self.labels_equipment)
        except Exception as e:
            print('Error', e)
            print('load model failed. you will likely have to load the model from command')'''

    def initialize_reader_settings(self):
        # Initialize reader settings
        self.instrument_reader_settings = {
            "low_text": 0.3,
            "min_size": 10,
            "ycenter_ths": 0.5,
            "height_ths": 0.5,
            "width_ths": 6.0,
            "add_margin": -0.1,
            "link_threshold": 0.2,
            "text_threshold": 0.3,
            "mag_ratio": 3.0,
            "allowlist": '0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ()',
            "decoder": 'beamsearch',
            "batch_size": 1
        }

        self.reader_settings = {
            "low_text": 0.4,
            "min_size": 10,
            "ycenter_ths": 0.5,
            "height_ths": 0.5,
            "width_ths": 2.4,
            "add_margin": 0.1,
            "link_threshold": 0.13,
            "text_threshold": 0.3,
            "mag_ratio": 1.0,
            "allowlist": '',
            "decoder": 'beamsearch',
            "batch_size": 1

        }

    def initialize_other_attributes(self):
        # Initialize other attributes
        self.ocr_results = None
        self.comment_box_expand = 20
        self.line = None
        self.service_in = None
        self.service_out = None
        self.equipment = None
        self.inst_data = None
        self.pid = None
        self.pid_coords = None
        self.comment = None
        self.persistent_boxes = []
        self.persistent_texts = []
        self.correct_fn = None
        self.correct_fn_path = None
        self.mouse_pressed = False
        self.start_x = 0
        self.start_y = 0
        self.current_selection_box = None
        self.current_text = None
        self.wb = None
        self.sheet = None
        self.capture_mode = 'pid'
        self.capture_actions = {
            'pid': self.capture_pid,
            'instruments': self.capture_instruments,
            'line': self.capture_line,
            'equipment': self.capture_equipment,
            'service_in': self.capture_service_in,
            'service_out': self.capture_service_out,
            'comment': self.capture_comment
        }
        self.whole_page_ocr_results = None
        self.reader = easyocr.Reader(['en'])
        self.detection_labels = []
        self.tag_label_groups = {"FE FIT": ["CORIOLIS", "MAGNETIC", "PITOT", "TURBINE", "ULTRASONIC", "VORTEX"],
                                 "PCV TCV LCV SDV AV XV HCV FCV FV PV TV LV": ["BALL", "BUTTERFLY", "DIAPHRAM", "GATE",
                                                                               "GLOBE", "KNIFE", "PLUG", "VBALL"],
                                 "LE LIT LT": ["GWR", "PR"], "PT PIT PI DPIT": ["SEAL"]}
        self.group_inst = []
        self.object_box_expand = 0.0
        self.group_other = []
        self.min_scores = {}
        self.association_radius = 33
        self.minscore_inst = 0.74
        self.nms_threshold = 0.5
        self.inst_data = []
        self.active_inst_box_count = 0
        self.write_mode = 'xlwings'
        self.equipment_defined = None
        self.crop_start = None
        self.crop_end = None
        self.workbook_path = None

    def bind_events_to_canvas(self):
        # Bind events to the canvas
        self.canvas.bind('<Configure>', self.show_image)  # canvas is resized
        self.canvas.bind('<ButtonPress-3>', self.move_from)
        self.canvas.bind('<B3-Motion>', self.move_to)
        self.canvas.bind('<MouseWheel>', self.wheel)  # with Windows and MacOS, but not Linux
        self.canvas.bind('<Button-5>', self.wheel)  # only with Linux, wheel scroll down
        self.canvas.bind('<Button-4>', self.wheel)  # only with Linux, wheel scroll up
        self.canvas.bind('<ButtonPress-1>', self.start_crop)
        self.canvas.bind('<B1-Motion>', self.update_crop)
        self.canvas.bind('<ButtonRelease-1>', self.end_crop)

    def bind_key_shortcuts(self):
        # Bind key shortcuts to the respective commands
        self.root.bind('n', lambda event: self.next_image())
        self.root.bind('b', lambda event: self.previous_image())
        self.root.bind('p', lambda event: self.set_capture('pid'))
        self.root.bind('a', lambda event: self.set_capture('instruments'))
        self.root.bind('f', lambda event: self.set_capture('line'))
        self.root.bind('e', lambda event: self.set_capture('equipment'))
        self.root.bind('z', lambda event: self.set_capture('service_in'))
        self.root.bind('x', lambda event: self.set_capture('service_out'))
        self.root.bind('w', lambda event: self.append_data_to_excel())
        self.root.bind('c', lambda event: self.clear_instrument_group())
        self.root.bind('v', lambda event: self.vote())
        self.root.bind('s', lambda event: self.swap_services())
        self.root.bind('g', lambda event: self.set_capture('comment'))
        self.root.bind('<Return>', lambda event: self.set_comment())

        # Bind key shortcuts to the respective commands
        self.root.bind('N', lambda event: self.next_image())
        self.root.bind('B', lambda event: self.previous_image())
        self.root.bind('P', lambda event: self.set_capture('pid'))
        self.root.bind('A', lambda event: self.set_capture('instruments'))
        self.root.bind('F', lambda event: self.set_capture('line'))
        self.root.bind('E', lambda event: self.set_capture('equipment'))
        self.root.bind('Z', lambda event: self.set_capture('service_in'))
        self.root.bind('X', lambda event: self.set_capture('service_out'))
        self.root.bind('W', lambda event: self.append_data_to_excel())
        self.root.bind('C', lambda event: self.clear_instrument_group())
        self.root.bind('V', lambda event: self.vote())
        self.root.bind('S', lambda event: self.swap_services())
        self.root.bind('G', lambda event: self.set_capture('comment'))

        # Initialize shift key binding
        self.shift_held = False
        self.root.bind('<KeyPress-Shift_L>', self.shift_pressed)
        self.root.bind('<KeyRelease-Shift_L>', self.shift_released)

        self.ctrl_held = False
        self.root.bind('<KeyPress-Control_L>', self.ctrl_pressed)
        self.root.bind('<KeyRelease-Control_L>', self.ctrl_released)

    def save_attributes(self):
        """Save class attributes to a JSON file"""
        attributes_to_save = [
            'pid_coords',
            'current_image_index',
            'instrument_reader_settings',
            'reader_settings',
            'model_inst_path',
            'labels',
            'group_inst',
            'group_other',
            'comment_box_expand',
            'association_radius',
            'min_scores',
            'tag_label_groups',
            # Add any other attribute names you want to save here
        ]

        attributes = {attr_name: getattr(self, attr_name) for attr_name in attributes_to_save}

        attributes_file = os.path.join(self.folder_path, 'attributes.json')
        with open(attributes_file, 'w') as file:
            json.dump(attributes, file)

    def load_attributes(self):
        """Load class attributes from a JSON file"""
        attributes_file = os.path.join(self.folder_path, 'attributes.json')
        if os.path.exists(attributes_file):
            try:
                with open(attributes_file, 'r') as file:
                    attributes = json.load(file)
                    # Automatically load attributes if they exist in the JSON file
                    for key, value in attributes.items():
                        if hasattr(self, key):
                            setattr(self, key, value)
            except json.JSONDecodeError:
                print("Error: Invalid JSON file format.")
        else:
            print("Attribute file not found. Using default values.")

    def create_capture_text(self):
        # Remove the previous capture text if it exists
        if hasattr(self, 'capture_text') and self.capture_text:
            self.canvas.delete(self.capture_text)

        # Create the capture text
        self.capture_text = self.canvas.create_text(0, 0, text=self.capture_mode, font=("Arial", 8), fill="orange")

        # Bind the mouse motion event to update the capture text position
        self.canvas.bind("<Motion>", self.update_capture_text)

    # endregion

    # region Image Loading and Navigation

    def load_image(self):
        if self.image_list:

            self.image_path = self.image_list[self.current_image_index]
            print(self.image_path)
            name, ext = os.path.splitext(self.image_path)
            self.results_folder = name + "_results"
            if not os.path.exists(self.results_folder):
                os.makedirs(self.results_folder)

            self.original_image = Image.open(self.image_path)
            self.cv2img = pil_to_cv2(self.original_image)

            self.load_ocr()

            self.width, self.height = self.original_image.size

            # Put image into container rectangle and use it to set proper coordinates to the image
            self.image_container = self.canvas.create_rectangle(0, 0, self.width, self.height, width=0)

            # Bind events to the canvas
            self.bind_events_to_canvas()

            # Get the current window size
            window_width = self.canvas.winfo_width()
            window_height = self.canvas.winfo_height()

            # Calculate the scaling factor
            width_ratio = window_width / self.width
            height_ratio = window_height / self.height
            scale_factor = min(width_ratio, height_ratio)

            # Calculate new dimensions
            new_width = int(self.width * scale_factor)
            new_height = int(self.height * scale_factor)

            # Clear the previous image from the canvas
            self.canvas.delete("all")

            # Set the initial scale
            self.imscale = scale_factor

            # Update the scroll region
            self.canvas.config(scrollregion=(0, 0, new_width, new_height))

            # Create a new image container
            self.image_container = self.canvas.create_rectangle(0, 0, new_width, new_height, width=0)

            # Show the image
            self.show_image()

            # Center the image
            self.center_image()

            self.create_capture_text()

    def center_image(self):
        # Get the current window size
        window_width = self.canvas.winfo_width()
        window_height = self.canvas.winfo_height()

        # Get the current image size
        bbox = self.canvas.bbox(self.image_container)
        image_width = bbox[2] - bbox[0]
        image_height = bbox[3] - bbox[1]

        # Calculate offsets
        offset_x = 0  # (window_width - image_width) / 2
        offset_y = 0  # (window_height - image_height) / 2

        # Move the image
        self.canvas.move(self.image_container, offset_x, offset_y)
        self.show_image()

    def go_to_page(self, page_index):
        if self.image_list:
            if page_index < 0:
                self.current_image_index = 0  # Set to the first image
            elif page_index >= len(self.image_list):
                self.current_image_index = len(self.image_list) - 1  # Set to the last image
            else:
                self.current_image_index = page_index

            self.load_image()

            if self.pid_coords:
                # Crop the image using the scaled coordinates
                print(self.pid_coords)

                cropped_image = self.original_image.crop(self.pid_coords)

                cropped_image = pil_to_cv2(cropped_image)
                # this is weird but we need to set the self.cropped coords so that we dont let the last used coord overwrite pid
                self.cropped_x1, self.cropped_y1, self.cropped_x2, self.cropped_y2 = self.pid_coords
                self.capture_pid(cropped_image)

                self.update_data_display()

            self.load_ocr()

    def next_image(self):
        self.go_to_page(self.current_image_index + 1)

    def previous_image(self):
        self.go_to_page(self.current_image_index - 1)

    def open_go_to_page(self):

        page_index = tk.simpledialog.askinteger("Go to Page", "Enter the page index:")

        if page_index is not None:
            self.go_to_page(page_index - 1)  # Adjust the index since it's 0-based
            self.clear_boxes()

    def open_folder(self, given_folder=None):

        self.initialize_models()

        # Define a custom sorting key function
        def natural_sort_key(s):
            return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', s)]

        if not given_folder:
            self.folder_path = filedialog.askdirectory()
        else:
            self.folder_path = given_folder

        if self.folder_path:
            self.workbook_path = os.path.join(self.folder_path, 'index.xlsx')

            self.image_list = [os.path.join(self.folder_path, f) for f in os.listdir(self.folder_path) if
                               f.endswith(('.png', '.jpg', '.jpeg', '.gif'))]
            # Sort the collected files using the natural sort key function
            self.image_list.sort(key=natural_sort_key)
            self.current_image_index = 0
            self.load_attributes()
            if self.current_image_index:
                self.go_to_page(self.current_image_index)
            else:
                self.go_to_page(1)

    def open_pdf2png(self):
        # Ask for DPI value
        width = tk.simpledialog.askinteger("DPI", "Enter the image width:", initialvalue=5000)

        # Open file dialog to select PDF file
        root = tk.Tk()
        root.withdraw()  # Hide the root window
        pdf_file = filedialog.askopenfilename(title="Select PDF File", filetypes=[("PDF Files", "*.pdf")])

        if pdf_file and width:
            # Call the pdf2png function with the selected PDF file and DPI value
            images_folder = pdf2png(pdf_file, width)
        else:
            return

        if tk.messagebox.askyesno("Open the project?"):
            self.open_folder(images_folder)

    # endregion

    # region Canvas and Image Manipulation

    def scroll_y(self, *args, **kwargs):
        ''' Scroll canvas vertically and redraw the image '''
        self.canvas.yview(*args, **kwargs)  # scroll vertically
        self.show_image()  # redraw the image

    def scroll_x(self, *args, **kwargs):
        ''' Scroll canvas horizontally and redraw the image '''
        self.canvas.xview(*args, **kwargs)  # scroll horizontally
        self.show_image()  # redraw the image

    def move_from(self, event):
        ''' Remember previous coordinates for scrolling with the mouse '''
        self.canvas.scan_mark(event.x, event.y)

    def move_to(self, event):
        ''' Drag (move) canvas to the new position '''
        self.canvas.scan_dragto(event.x, event.y, gain=1)
        self.show_image()  # redraw the image

    def wheel(self, event):
        ''' Zoom with mouse wheel '''
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        bbox = self.canvas.bbox(self.image_container)  # get image area
        if bbox[0] < x < bbox[2] and bbox[1] < y < bbox[3]:
            pass  # Ok! Inside the image
        else:
            return  # zoom only inside image area
        scale = 1.0
        # Respond to Linux (event.num) or Windows (event.delta) wheel event
        if event.num == 5 or event.delta == -120:  # scroll down
            i = min(self.width, self.height)
            if int(i * self.imscale) < 30: return  # image is less than 30 pixels
            self.imscale /= self.delta
            scale /= self.delta
        if event.num == 4 or event.delta == 120:  # scroll up
            i = min(self.canvas.winfo_width(), self.canvas.winfo_height())
            if i < self.imscale: return  # 1 pixel is bigger than the visible area
            self.imscale *= self.delta
            scale *= self.delta
        self.canvas.scale('all', x, y, scale, scale)  # rescale all canvas objects
        # print('scale',scale)
        self.show_image()

    def show_image(self, event=None):
        # print('self.imscale', self.imscale)
        # Get the current horizontal and vertical offsets
        x_offset = self.canvas.xview()[0]
        y_offset = self.canvas.yview()[0]

        # Print the offsets (you can adjust the formatting as needed)
        # print(f"Horizontal Offset: {x_offset}, Vertical Offset: {y_offset}")
        ''' Show image on the Canvas '''
        bbox1 = self.canvas.bbox(self.image_container)  # get image area
        # Remove 1 pixel shift at the sides of the bbox1
        bbox1 = (bbox1[0] + 1, bbox1[1] + 1, bbox1[2] - 1, bbox1[3] - 1)
        bbox2 = (self.canvas.canvasx(0),  # get visible area of the canvas
                 self.canvas.canvasy(0),
                 self.canvas.canvasx(self.canvas.winfo_width()),
                 self.canvas.canvasy(self.canvas.winfo_height()))
        bbox = [min(bbox1[0], bbox2[0]), min(bbox1[1], bbox2[1]),  # get scroll region box
                max(bbox1[2], bbox2[2]), max(bbox1[3], bbox2[3])]
        if bbox[0] == bbox2[0] and bbox[2] == bbox2[2]:  # whole image in the visible area
            bbox[0] = bbox1[0]
            bbox[2] = bbox1[2]
        if bbox[1] == bbox2[1] and bbox[3] == bbox2[3]:  # whole image in the visible area
            bbox[1] = bbox1[1]
            bbox[3] = bbox1[3]
        self.canvas.configure(scrollregion=bbox)  # set scroll region
        x1 = max(bbox2[0] - bbox1[0], 0)  # get coordinates (x1,y1,x2,y2) of the image tile
        y1 = max(bbox2[1] - bbox1[1], 0)
        x2 = min(bbox2[2], bbox1[2]) - bbox1[0]
        y2 = min(bbox2[3], bbox1[3]) - bbox1[1]
        if int(x2 - x1) > 0 and int(y2 - y1) > 0:  # show image if it in the visible area
            x = min(int(x2 / self.imscale), self.width)  # sometimes it is larger on 1 pixel...
            y = min(int(y2 / self.imscale), self.height)  # ...and sometimes not
            image = self.original_image.crop((int(x1 / self.imscale), int(y1 / self.imscale), x, y))
            imagetk = ImageTk.PhotoImage(image.resize((int(x2 - x1), int(y2 - y1))))
            imageid = self.canvas.create_image(max(bbox2[0], bbox1[0]), max(bbox2[1], bbox1[1]),
                                               anchor='nw', image=imagetk)
            self.canvas.lower(imageid)  # set image into background
            self.canvas.imagetk = imagetk  # keep an extra reference to prevent garbage-collection

        # After creating the image on the canvas, bring the crop rectangle to the front if it exists
        if self.crop_rectangle:
            self.canvas.tag_raise(self.crop_rectangle)

    def update_capture_text(self, event):
        # Update the text's position to follow the cursor
        # x, y = event.x, event.y
        x, y = self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)
        self.canvas.coords(self.capture_text, x - 8, y - 8)

    # endregion

    # region Cropping and Selection

    def start_crop(self, event):
        self.crop_start = (self.canvas.canvasx(event.x), self.canvas.canvasy(event.y))
        if self.crop_rectangle and self.crop_rectangle not in self.persistent_boxes:
            self.canvas.delete(self.crop_rectangle)

    def update_crop(self, event):
        if self.crop_start:
            x, y = self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)
            if self.crop_rectangle and self.crop_rectangle not in self.persistent_boxes:
                self.canvas.delete(self.crop_rectangle)
            # self.crop_rectangle = self.canvas.create_rectangle(self.crop_start[0], self.crop_start[1], x, y,
            #                                                   outline='orange')

        self.crop_rectangle = self.canvas.create_rectangle(
            self.crop_start[0],
            self.crop_start[1],
            x, y,
            outline='orange',
            fill='orange',
            stipple='gray25'  # Creates a checkerboard pattern for transparency effect
        )

    def end_crop(self, event):
        if self.crop_start:
            self.crop_end = (self.canvas.canvasx(event.x), self.canvas.canvasy(event.y))
            self.perform_crop()

    def perform_crop(self):
        if self.crop_start and self.crop_end:
            # Convert canvas coordinates to image coordinates
            x1c, y1c = self.canvas_to_image(self.crop_start[0], self.crop_start[1])
            x2c, y2c = self.canvas_to_image(self.crop_end[0], self.crop_end[1])
            # Ensure x1 < x2 and y1 < y2
            x1 = min(x1c, x2c)
            x2 = max(x1c, x2c)

            y1 = min(y1c, y2c)
            y2 = max(y1c, y2c)

            self.cropped_x1 = x1
            self.cropped_y1 = y1
            self.cropped_x2 = x2
            self.cropped_y2 = y2

            # Crop the image
            cropped_image = self.original_image.crop((x1, y1, x2, y2))
            self.cropped_image = pil_to_cv2(cropped_image)
            # Save or display the cropped image
            cropped_image.save("ocr_capture.png")
            print("Image cropped and saved as 'ocr_capture.png'")

            # Perform the action based on self.capture
            if self.capture_mode in self.capture_actions:
                self.capture_actions[self.capture_mode](self.cropped_image)
                self.update_data_display()
            else:
                print(f"Invalid capture action: {self.capture_mode}")

            # Clear the crop rectangle
            if self.crop_rectangle not in self.persistent_boxes:
                self.canvas.delete(self.crop_rectangle)
            self.crop_start = None
            self.crop_end = None

    def canvas_to_image(self, canvas_x, canvas_y):
        # Convert canvas coordinates to image coordinates
        bbox = self.canvas.bbox(self.image_container)
        image_x = int((canvas_x - bbox[0]) / self.imscale)
        image_y = int((canvas_y - bbox[1]) / self.imscale)
        return image_x, image_y

    def clear_boxes(self):

        for box in self.persistent_boxes:
            self.canvas.delete(box)  # Remove the previous box
            # self.canvas.delete(text)

        # not sure if this is necessary as zip clears stuff
        self.persistent_boxes = []

    # endregion

    # region Data Capture

    def capture_pid(self, cropped_image):
        print('Perform actions for capturing PID')
        try:
            # Perform actions for capturing line
            result = self.reader.readtext(cropped_image, **self.reader_settings)
            if result:
                self.pid = ' '.join([box[1] for box in result])
                self.pid_coords = (self.cropped_x1, self.cropped_y1, self.cropped_x2, self.cropped_y2)
            else:
                self.pid = ''
                print('no result')
        except Exception as e:
            print('error capturing pid \n', e)

    def capture_instruments(self, cropped_image):
        offset = (self.cropped_x1, self.cropped_y1)
        # Perform actions for capturing instruments
        # cropped_image = pil_to_cv2(cropped_image)
        labels, boxes, scores = model_predict_on_mosaic(cropped_image, self.model_inst, threshold=self.nms_threshold)
        if labels:
            self.persistent_boxes.append(self.crop_rectangle)
            print('added ', self.crop_rectangle, ' to persistent boxes')
            self.active_inst_box_count += 1
            for i, label in enumerate(labels):
                print(f"{label}: {scores[i]:.2f} | ", end='')
            # self.persistent_texts.append(self.current_text)
            print()
            inst_prediction_data = zip(labels, boxes, scores)

            inst_data = return_inst_data(inst_prediction_data, cropped_image, self.reader,
                                         self.instrument_reader_settings, radius=self.association_radius,
                                         min_scores=self.min_scores, expand=self.object_box_expand,
                                         offset=offset, comment_box_expand=self.comment_box_expand,
                                         ocr_results=self.ocr_results,
                                         inst_labels=self.group_inst, other_labels=self.group_other,
                                         tag_label_groups=self.tag_label_groups)

            for data in inst_data:
                print(data)

            self.inst_data.extend(inst_data)

    def capture_line(self, cropped_image):
        self.process_captured_text(cropped_image, 'line')

    def capture_equipment(self, cropped_image):
        self.process_captured_text(cropped_image, 'equipment')

    def capture_service_in(self, cropped_image):
        self.process_captured_text(cropped_image, 'service_in')

    def capture_service_out(self, cropped_image):
        self.process_captured_text(cropped_image, 'service_out')

    def capture_comment(self, cropped_image):
        print('Perform actions for capturing a comment')
        # Perform actions for capturing line
        result = self.reader.readtext(cropped_image, **self.reader_settings)
        if result:
            self.comment = ' '.join([box[1] for box in result])
        else:
            self.comment = ''

    def process_captured_text(self, cropped_image, target_attribute):
        """
        Unified method for processing captured text from images.

        Args:
            cropped_image: The image to process
            target_attribute: String name of attribute to update ('line', 'equipment', 'service_in', 'service_out')
        """
        # Handle rotation for line captures only
        if target_attribute == 'line':
            height, width = cropped_image.shape[:2]
            if height > width:
                cropped_image = cv2.rotate(cropped_image, cv2.ROTATE_90_CLOCKWISE)

        # Perform OCR
        result = self.reader.readtext(cropped_image, **self.reader_settings)

        if not result:
            setattr(self, target_attribute, '')
            return

        just_text = ' '.join([box[1] for box in result])
        current_value = getattr(self, target_attribute, '')  # Default to empty string if attribute doesn't exist

        # Process text based on modifier keys
        if self.ctrl_held:
            new_text = merge_common_substrings(current_value, just_text)
        elif self.shift_held:
            new_text = f"{current_value} {just_text}".strip()
        else:
            new_text = just_text

        # Update the target attribute
        setattr(self, target_attribute, new_text)

        # Clear other attributes based on the capture type
        if target_attribute in ('line', 'equipment'):
            other_attr = 'equipment' if target_attribute == 'line' else 'line'
            setattr(self, other_attr, None)
        elif target_attribute in ('service_in', 'service_out'):
            self.equipment = None

    def set_capture(self, capture_type):
        self.capture_mode = capture_type
        self.canvas.itemconfig(self.capture_text, text=self.capture_mode)

    # endregion

    # region Data Management

    def create_data_window(self):
        self.data_window = tk.Toplevel(self.root)
        self.data_window.title("Captured Data")

        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        data_window_width = 250
        data_window_x = screen_width - data_window_width
        self.data_window.geometry(f'{data_window_width}x{screen_height}+{data_window_x}+0')

        # Create a frame to hold all widgets
        self.data_frame = ttk.Frame(self.data_window)
        self.data_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # PID Entry
        ttk.Label(self.data_frame, text="PID:").pack(anchor='w')
        self.pid_entry = ttk.Entry(self.data_frame)
        self.pid_entry.pack(fill=tk.X, padx=5, pady=(0, 10))

        # Page display (non-editable)
        self.page_label = ttk.Label(self.data_frame, text="")
        self.page_label.pack(anchor='w', pady=(0, 10))

        # Line Entry
        ttk.Label(self.data_frame, text="Line:").pack(anchor='w')
        self.line_entry = ttk.Entry(self.data_frame)
        self.line_entry.pack(fill=tk.X, padx=5, pady=(0, 10))

        # Service In Entry
        ttk.Label(self.data_frame, text="Service In:").pack(anchor='w')
        self.service_in_entry = ttk.Entry(self.data_frame)
        self.service_in_entry.pack(fill=tk.X, padx=5, pady=(0, 10))

        # Service Out Entry
        ttk.Label(self.data_frame, text="Service Out:").pack(anchor='w')
        self.service_out_entry = ttk.Entry(self.data_frame)
        self.service_out_entry.pack(fill=tk.X, padx=5, pady=(0, 10))

        # Equipment Entry
        ttk.Label(self.data_frame, text="Equipment:").pack(anchor='w')
        self.equipment_entry = ttk.Entry(self.data_frame)
        self.equipment_entry.pack(fill=tk.X, padx=5, pady=(0, 10))

        # Comment Entry
        ttk.Label(self.data_frame, text="Comment:").pack(anchor='w')
        self.comment_entry = ttk.Entry(self.data_frame)
        self.comment_entry.pack(fill=tk.X, padx=5, pady=(0, 10))

        # Modify entry creation sections to add FocusOut binding
        self.pid_entry.bind('<FocusOut>', lambda e: self.get_data_from_window())
        self.line_entry.bind('<FocusOut>', lambda e: self.get_data_from_window())
        self.service_in_entry.bind('<FocusOut>', lambda e: self.get_data_from_window())
        self.service_out_entry.bind('<FocusOut>', lambda e: self.get_data_from_window())
        self.equipment_entry.bind('<FocusOut>', lambda e: self.get_data_from_window())
        self.comment_entry.bind('<FocusOut>', lambda e: self.get_data_from_window())

        # Instrument Data Tree
        ttk.Label(self.data_frame, text="Instrument Data:").pack(anchor='w')
        self.inst_tree = ttk.Treeview(self.data_frame, columns=('Tag', 'Tag No', 'Type'), show='headings', height=10)
        self.inst_tree.heading('Tag', text='Tag')
        self.inst_tree.heading('Tag No', text='Tag No')
        self.inst_tree.heading('Type', text='Type')
        self.inst_tree.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # Add scrollbar for tree
        scrollbar = ttk.Scrollbar(self.data_frame, orient=tk.VERTICAL, command=self.inst_tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.inst_tree.configure(yscrollcommand=scrollbar.set)

        # Bind double-click event for editing instrument data
        self.inst_tree.bind('<Double-1>', self.edit_instrument)

        self.data_window.protocol("WM_DELETE_WINDOW", self.update_data_display)
        self.root.bind("<FocusIn>", lambda event: self.data_window.lift())

        # Position main window
        root_window_width = screen_width - data_window_width
        self.root.geometry(f'{root_window_width}x{screen_height}+0+0')

    def update_data_display(self):
        # Update entries with current values
        self.pid_entry.delete(0, tk.END)
        self.pid_entry.insert(0, str(self.pid) if self.pid else "")

        self.page_label.config(text=f"Page: {self.current_image_index + 1} of {len(self.image_list)}")

        self.line_entry.delete(0, tk.END)
        self.line_entry.insert(0, str(self.line) if self.line else "")

        self.service_in_entry.delete(0, tk.END)
        self.service_in_entry.insert(0, condense_hyphen_string(self.service_in) if self.service_in else "")

        self.service_out_entry.delete(0, tk.END)
        self.service_out_entry.insert(0, condense_hyphen_string(self.service_out) if self.service_out else "")

        self.equipment_entry.delete(0, tk.END)
        self.equipment_entry.insert(0, str(self.equipment) if self.equipment else "")

        self.comment_entry.delete(0, tk.END)
        self.comment_entry.insert(0, str(self.comment) if self.comment else "")

        # Clear and update instrument tree
        for item in self.inst_tree.get_children():
            self.inst_tree.delete(item)

        for data in self.inst_data:
            self.inst_tree.insert('', tk.END, values=(data['tag'], data['tag_no'], data['type']))

    def get_data_from_window(self):
        """Retrieve data from the window widgets and update instrument data"""
        self.pid = self.pid_entry.get()
        self.line = self.line_entry.get()
        self.service_in = self.service_in_entry.get()
        self.service_out = self.service_out_entry.get()
        self.equipment = self.equipment_entry.get()
        self.comment = self.comment_entry.get()

        # Update existing instrument data while preserving structure
        updated_data = []
        for item in self.inst_tree.get_children():
            values = self.inst_tree.item(item)['values']
            inst_dict = {
                'tag': values[0] if values[0] else '',  # Ensure empty string instead of None
                'tag_no': values[1] if values[1] else '',  # Ensure empty string instead of None
                'type': values[2] if values[2] else ''  # Ensure empty string instead of None
            }
            # Preserve any existing additional data
            if self.inst_data:
                existing = next((d for d in self.inst_data if d['tag'] == values[0]), {})
                inst_dict.update({k: v for k, v in existing.items() if k not in ['tag', 'tag_no', 'type']})
            updated_data.append(inst_dict)

        self.inst_data = updated_data

    def edit_instrument(self, event):
        """Handle double-click editing of instrument data"""
        item = self.inst_tree.selection()[0]
        values = self.inst_tree.item(item)['values']

        # Create popup window for editing
        edit_window = tk.Toplevel(self.data_window)
        edit_window.title("Edit Instrument")

        ttk.Label(edit_window, text="Tag:").grid(row=0, column=0, padx=5, pady=5)
        tag_entry = ttk.Entry(edit_window)
        tag_entry.insert(0, values[0])
        tag_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(edit_window, text="Tag No:").grid(row=1, column=0, padx=5, pady=5)
        tag_no_entry = ttk.Entry(edit_window)
        tag_no_entry.insert(0, values[1])
        tag_no_entry.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(edit_window, text="Type:").grid(row=2, column=0, padx=5, pady=5)
        type_entry = ttk.Entry(edit_window)
        type_entry.insert(0, values[2])
        type_entry.grid(row=2, column=1, padx=5, pady=5)

        def save_changes():
            self.inst_tree.item(item, values=(tag_entry.get(), tag_no_entry.get(), type_entry.get()))
            edit_window.destroy()

        ttk.Button(edit_window, text="Save", command=save_changes).grid(row=3, column=0, columnspan=2, pady=10)

    def append_data(self, excel_type='xlwings'):
        """
        Append data to Excel file using either xlwings or openpyxl.

        Args:
            excel_type (str): Either 'xlwings' or 'openpyxl'
        """
        self.get_data_from_window()

        # Initialize workbook and worksheet
        if excel_type == 'xlwings':
            if not os.path.exists(self.workbook_path):
                self.wb = xw.Book()
                self.wb.save(self.workbook_path)
            wb = xw.Book(self.workbook_path)
            if 'Instrument Index' not in wb.sheet_names:
                wb.sheets.add(name='Instrument Index')
            ws = wb.sheets['Instrument Index']
        else:  # openpyxl
            if not os.path.exists(self.workbook_path):
                self.wb = openpyxl.Workbook()
            elif not self.wb:
                self.wb = openpyxl.load_workbook(self.workbook_path)
            if 'Instrument Index' not in self.wb.sheetnames:
                self.ws = self.wb.create_sheet('Instrument Index')
            else:
                self.ws = self.wb['Instrument Index']
            ws = self.ws

        for data in self.inst_data:
            # Initialize default data dictionary
            data.update({
                'PID': self.pid,
                'SERVICE': '',
                'LINE/EQUIP': '',
                'INPUT COMMENT': self.comment if self.comment else '',
                'FILE': self.image_path,
            })

            # Process service information
            si = condense_hyphen_string(self.service_in)
            so = condense_hyphen_string(self.service_out)

            if self.service_in and self.service_out:
                data['SERVICE'] = f"{si} TO {so}" if self.service_in != self.service_out else f"{si} RECIRCULATION"
            elif self.service_in:
                data['SERVICE'] = f"{si} OUTLET"
            elif self.service_out:
                data['SERVICE'] = f"TO {so}"

            # Process line/equipment information
            if self.line:
                data['LINE/EQUIP'] = self.line
            elif self.equipment:
                words = self.equipment.split(' ')
                data['LINE/EQUIP'] = words[0]
                data['SERVICE'] = ' '.join(words[1:])

            # Get last row and handle header
            if excel_type == 'xlwings':
                last_row = ws.range('A1').expand('down').last_cell.row if ws.range('A1').value is not None else 0
            else:
                last_row = len(list(ws.rows))

            if (excel_type == 'xlwings' and ws.range('A1').value is None) or \
                    (excel_type == 'openpyxl' and ws['A1'].value is None):
                self.create_excel_header(ws, data)
                last_row += 1

            self.populate_excel_row(ws, data, last_row + 1)

        if excel_type == 'openpyxl':
            self.wb.save(self.workbook_path)

        self.turn_boxes_blue()

    def append_data_to_excel(self):
        try:
            self.append_data(self.write_mode)
        except Exception as e:
            tk.messagebox.showerror(e)
            print(f"error {e}")

    def create_excel_header(self, worksheet, data):
        """
        Create a header row in the given worksheet based on the keys in the data dictionary
        and format entire columns as text.
        :param worksheet: The worksheet to add the header to (either xlwings or openpyxl worksheet)
        :param data: A dictionary containing the data structure
        :return: None
        """
        header = list(data.keys())
        print(header)

        if isinstance(worksheet, xw.main.Sheet):
            # xlwings worksheet
            for i, column_header in enumerate(header, start=1):
                cell = worksheet.range((1, i))
                cell.value = column_header
                # Format entire column as text
                # Get the last row in the worksheet
                last_row = worksheet.cells.last_cell.row
                # Format the entire column from row 1 to last row
            column_range = worksheet.range((1, 2), (last_row, 2))
            column_range.number_format = '@'

        elif isinstance(worksheet, openpyxl.worksheet.worksheet.Worksheet):
            # openpyxl worksheet
            for col, column_header in enumerate(header, start=1):
                worksheet.cell(row=1, column=col).value = column_header
            # Format entire column as text
            column_letter = openpyxl.utils.get_column_letter(2)
            for cell in worksheet[column_letter]:
                cell.number_format = '@'
        else:
            raise ValueError("Unsupported worksheet type")

    def populate_excel_row(self, worksheet, data, row):
        """
        Populate a row in the given worksheet with the provided data.

        :param worksheet: The worksheet to populate (either xlwings or openpyxl worksheet)
        :param data: A dictionary containing the data to be written
        :param row: The row number to populate
        :return: None
        """
        # Ensure all dictionary values are strings and empty strings for None
        processed_data = {k: str(v) if v is not None else '' for k, v in data.items()}

        for col, (key, value) in enumerate(processed_data.items(), start=1):
            if isinstance(worksheet, xw.main.Sheet):
                # xlwings worksheet
                cell = worksheet.range((row, col))
                cell.value = value
            elif isinstance(worksheet, openpyxl.worksheet.worksheet.Worksheet):
                # openpyxl worksheet
                cell = worksheet.cell(row=row, column=col)
                cell.value = value
            else:
                raise ValueError("Unsupported worksheet type")

        if isinstance(worksheet, openpyxl.worksheet.worksheet.Worksheet):
            worksheet.parent.save(self.workbook_path)

    def turn_boxes_blue(self):
        # Slice the list to get the last 4 items
        active_boxes = self.persistent_boxes[-self.active_inst_box_count:]

        # Iterate over the sliced list and change the outline color of each item
        for box_id in active_boxes:
            self.canvas.itemconfig(box_id, outline='#87CEEB', fill='#87CEEB')

        self.active_inst_box_count = 0
        self.inst_data = []
        # self.comment = ''
        self.update_data_display()

    def vote(self):
        # Create a dictionary to store tag_no counts
        tag_counts = {}

        # Count the occurrences of each tag_no
        for data in self.inst_data:
            tag_no = data['tag_no']
            tag_counts[tag_no] = tag_counts.get(tag_no, 0) + 1

        # Find the most frequent tag_no
        most_frequent_tag_no = max(tag_counts, key=tag_counts.get)

        # Assign the most frequent tag_no to each entry
        for data in self.inst_data:
            data['tag_no'] = most_frequent_tag_no

        self.update_data_display()

    def swap_services(self):
        self.service_in, self.service_out = self.service_out, self.service_in
        self.update_data_display()

    def set_comment(self):
        self.comment = tk.simpledialog.askstring("Input", "Please enter a Comment:")
        self.update_data_display()

    def clear_instrument_group(self):

        if self.inst_data:
            for box in self.persistent_boxes[-self.active_inst_box_count:]:
                self.canvas.delete(box)  # Remove the previous box
        self.inst_data = []
        self.update_data_display()

    # endregion

    # region OCR and Model Operations

    def get_ocr(self):
        self.ocr_results = HardOCR(self.cv2img, self.reader, self.reader_settings)
        img_ocr_results = plot_ocr_results(self.cv2img, self.ocr_results)
        print(self.ocr_results)
        # save a pic
        cv2.imwrite(os.path.join(self.results_folder, 'ocr_img.png'), img_ocr_results)

        # save actual results
        save_path = os.path.join(self.results_folder, 'ocr.pkl')
        save_easyocr_results_pickle(self.ocr_results, filename=save_path)

    def get_all_ocr(self):
        self.current_image_index = 0

        for i in range(len(self.image_list)):
            self.load_image()
            self.get_ocr()
            self.current_image_index += 1

    def load_ocr(self):
        file_name = os.path.join(self.results_folder, 'ocr.pkl')
        try:
            self.ocr_results = load_easyocr_results_pickle(filename=file_name)
            print('ocr results loaded')
        except:
            print('error loading ocr results. setting results to none')
            self.ocr_results = None

    def load_model(self):
        # Create a module specification
        self.model_inst_path = filedialog.askopenfilename(filetypes=[('PTH Files', '*.pth')])
        # labels = ['inst', 'dcs', 'ball', 'globe', 'diaphragm', 'knife', 'vball', 'plug', 'butterfly', 'gate']
        # self.model_inst_path = r"models\saved_model_vid-v3.18_GEVO.pth"
        # load instrument recognition model
        print('loading model')
        self.model_inst = Model.load(self.model_inst_path, self.detection_labels)

    def load_pretrained_model(self, model_path=None):
        if model_path == None:
            model_path = filedialog.askopenfilename(title="Select pretrained model file",
                                                    filetypes=[("Model files", "*.pth")])
        if model_path:
            label_path = model_path.replace(".pth", ".txt")
            if os.path.exists(label_path):
                with open(label_path, "r") as f:
                    self.detection_labels = [x.strip() for x in f.read().split(",")]
                    print(self.detection_labels)
            else:
                tk.messagebox.showinfo("Load error", f"{label_path} with comma separated labels not found")
            print('make sure labels are loaded correctly')
            self.model_inst = Model.load(model_path, self.detection_labels)
            print("Pretrained model loaded successfully.")

    def set_write_mode(self, mode):
        self.write_mode = mode

    def test_model_mosaic(self):

        test_on_capture = tk.messagebox.askyesno(title='test on capture?',
                                                 message='Do you want to test the model on the Capture?')

        if self.model_inst is None:
            tk.messagebox.showerror("Error", "Model not loaded. Load a pretrained model first.")
            return

        minscore = tk.simpledialog.askinteger("Scale percent", "Enter a minscore 1-100:", initialvalue=50) / 100

        if test_on_capture == False:
            # Ask for an image to test
            image_path = filedialog.askopenfilename(title="Select an image for testing", initialfile=self.image_path,
                                                    filetypes=(("PNG files", "*.png"), ("JPEG files", "*.jpg")))
        else:
            image_path = 'ocr_capture.png'

        if image_path:
            image = cv2.imread(image_path)

            labels, boxes, scores = model_predict_on_mosaic(image, self.model_inst, threshold=self.nms_threshold)

            # Overlay boxes and labels on the image
            img_with_boxes = draw_detection_boxes(image, labels, boxes, scores, minscore=minscore)

            # Save the image with overlaid boxes and labels
            # temp_dir = tempfile.gettempdir()
            output_image_path = "result_image.png"
            cv2.imwrite(output_image_path, img_with_boxes)

            # Open the saved image using the default image viewer
            os.startfile(output_image_path)

    # endregion

    # region Settings and Configuration

    def set_instrument_reader_settings(self, rs):
        self.instrument_reader_settings = rs
        self.save_attributes()
        print('instrument reader settings: ', rs)

    def set_reader_settings(self, rs):
        self.reader_settings = rs
        self.save_attributes()
        print('reader settings: ', rs)

    def open_instrument_reader_settings(self):
        # Create a new window for ObjectDetectionApp
        img_path = 'instrument_capture.png'
        if os.path.exists(img_path):
            reader_settings_root = tk.Toplevel()
            SetReaderSettings(reader_settings_root, img_path, self.reader,
                              reader_settings=self.instrument_reader_settings,
                              callback=self.set_instrument_reader_settings)
        else:
            print('first capture an instrument')

    def open_general_reader_settings(self):
        # Create a new window for ObjectDetectionApp
        img_path = 'ocr_capture.png'
        if os.path.exists(img_path):
            reader_settings_root = tk.Toplevel()
            SetReaderSettings(reader_settings_root, img_path, self.reader,
                              reader_settings=self.reader_settings,
                              callback=self.set_reader_settings)
        else:
            print('first capture an instrument')

    def set_tag_label_groups(self):
        group_window = tk.Toplevel(self.root)
        print(self.tag_label_groups)
        app = DictionaryBuilder(group_window, self.tag_label_groups, self.detection_labels)
        self.tag_label_groups = app.run()
        print(self.tag_label_groups)

    def set_nms_threshold(self):
        response = tkinter.simpledialog.askfloat(title='Non-Maximum-Supression',
                                                 prompt='Set NMS overlap threshold (smaller = fewer boxes): ',
                                                 initialvalue=self.nms_threshold)
        try:
            self.nms_threshold = float(response)
        except Exception as e:
            tk.messagebox.showinfo('set NMS fail', e)

    def set_object_box_expand(self):
        response = tkinter.simpledialog.askfloat(title='Set percent box expand',
                                                 prompt='Enter the % box expand for group inst',
                                                 initialvalue=self.object_box_expand)
        try:
            self.object_box_expand = float(response)
        except Exception as e:
            tk.messagebox.showinfo('set box expand fail', e)

    def set_object_scores(self):
        def set_minscores(score_dict):
            self.min_scores = score_dict

        slider_window = tk.Toplevel(self.root)

        for label in self.detection_labels:
            if self.min_scores.get(label):
                continue
            else:
                self.min_scores[label] = self.minscore_inst

        SliderApp(slider_window, self.min_scores, callback=set_minscores)

    def categorize_labels(self):
        manager = GroupManager(self.detection_labels, self.group_inst, self.group_other)
        manager.run()
        self.group_inst = manager.group_capture
        self.group_other = manager.group_association
        print(self.group_other)

    def set_comment_box_expand(self):
        self.minscore_inst = tk.simpledialog.askinteger(prompt="Enter Comment Box Expand",
                                                        title="Enter Comment Box Expand",
                                                        initialvalue=self.comment_box_expand)

    def set_association_radius(self):
        self.association_radius = tk.simpledialog.askfloat(prompt="Enter Object association radius",
                                                           title="Enter association ( Radius",
                                                           initialvalue=self.association_radius)

    # endregion

    # region File Operations

    def save_workbook(self):
        # Save the workbook
        if self.write_mode == 'openpyxl':
            self.wb.save(self.workbook_path)
            print('wb saved')
        else:
            print('not in openpyxl write mode. cant save')

    def merge_pdfs(self):
        folder_that_has_pdfs = filedialog.askdirectory(title='Folder that has PDFs')
        merge_pdf(folder_that_has_pdfs)
        self.open_folder(folder_that_has_pdfs)

    def make_pid_page_xlsx(self):

        pid_page_xlsx = openpyxl.Workbook()
        ws = pid_page_xlsx.create_sheet('pid_page')
        image_folder = filedialog.askdirectory(title='Folder that has PNGs')
        if self.pid_coords:

            for filename in os.listdir(image_folder):
                if filename.endswith(".png") or filename.endswith(".jpg"):
                    file_path = os.path.join(image_folder, filename)
                    img = cv2.imread(file_path)

                    self.cropped_x1, self.cropped_y1, self.cropped_x2, self.cropped_y2 = self.pid_coords
                    cropped_image = img[self.cropped_y1:self.cropped_y2, self.cropped_x1:self.cropped_x2]
                    self.capture_pid(cropped_image)

                    row = [filename, self.pid if self.pid else ""]
                    ws.append(row)

        counter = 1
        while True:
            try:
                save_location = os.path.join(image_folder, f"pid_page_{counter}.xlsx")
                pid_page_xlsx.save(save_location)
                print(f"File saved as: {save_location}")
                break
            except Exception as e:
                print(f"Error saving file: {e}")
                counter += 1

    def create_tag_type_xlsx(self):
        kernel_folder = filedialog.askdirectory(title='Folder that has Kernels')
        image_folder = filedialog.askdirectory(title='Folder that has PNGs')
        scale = tk.simpledialog.askinteger("Scale percent", "Enter a scale percent 1-100:", initialvalue=100) / 100
        expansion_pixels = tk.simpledialog.askinteger("Expand box", "Enter box expansion pixels:", initialvalue=180)
        confidence = tk.simpledialog.askinteger("Confidence", "Enter convolution confidence threshold:",
                                                initialvalue=80) / 100
        rotation = tk.simpledialog.askinteger("Rotation", "Enter rotation directions 1-4:", initialvalue=1)

        CR = ConvolutionReplacer(kernel_folder, scale, rotation)

        tag_type_xlsx = openpyxl.Workbook()
        ws = tag_type_xlsx.create_sheet('tagtype')
        # Define the header row
        header = ['TAG', 'TAG_NO', 'TYPE', 'PAGE', 'PID']
        # Write the header row to the first sheet of the workbook
        ws.append(header)

        # for img in image_folder
        for filename in os.listdir(image_folder):
            if filename.endswith(".png") or filename.endswith(".jpg"):
                file_path = os.path.join(image_folder, filename)
                img = cv2.imread(file_path)

                if self.pid_coords:
                    self.cropped_x1, self.cropped_y1, self.cropped_x2, self.cropped_y2 = self.pid_coords
                    cropped_image = img[self.cropped_y1:self.cropped_y2, self.cropped_x1:self.cropped_x2]
                    self.capture_pid(cropped_image)

                result_boxes_image, final_detections_rescaled = CR.detect(img, threshold=confidence)
                # final_detections_rescaled = [(top_left, bottom_right, label, rotation, score)...]
                cv2.imwrite('temp.png', result_boxes_image)
                # os.startfile('temp.png')

                for detection in final_detections_rescaled:
                    label = detection[2]

                    top_left, bottom_right = detection[0], detection[1]
                    # Expand the box by 100px (adjustable parameter)

                    x1, y1 = int(max(0, top_left[0] - expansion_pixels)), int(max(0, top_left[1] - expansion_pixels))
                    x2, y2 = int(min(img.shape[1], bottom_right[0] + expansion_pixels)), int(
                        min(img.shape[0], bottom_right[1] + expansion_pixels))

                    # Crop the image
                    cropped_image = img[y1:y2, x1:x2]

                    self.capture_instruments(cropped_image)

                    for data in self.inst_data:
                        data['label'] = label
                        row = [data['tag'], data['tag_no'], data['label'], filename, self.pid if self.pid else ""]
                        ws.append(row)

                    self.inst_data = []

        counter = 1
        while True:
            try:
                save_location = os.path.join(image_folder, f"tag_type{counter}.xlsx")
                tag_type_xlsx.save(save_location)
                print(f"File saved as: {save_location}")
                break
            except Exception as e:
                print(f"Error saving file: {e}")
                counter += 1

    def generate_instrument_count(self):

        ocr_needed = tk.messagebox.askyesno(title='OCR NEEDED?',
                                            message='Do we need to do OCR (for comments)?\nNote if avaliable we use existing results')
        type_needed = tk.messagebox.askyesno(title='TYPE NEEDED?',
                                             message='Do we need to do instrument types?')
        overwrite = tk.messagebox.askyesno(title='OVERWRITE?',
                                           message='Do you want to overwrite Existing counts?')

        sure = tk.messagebox.askyesno(title='SURE?',
                                      message='ARE U SURE')
        if not sure:
            return

        for i in range(len(self.image_list)):
            self.go_to_page(i)
            self.one_instrument_count(overwrite=overwrite)

        self.compile_excels()

    def compile_excels(self):
        # NOW we compile all the xlsxs into one
        output_file = compile_excels(self.folder_path, self.folder_path, prefix='Instrument_Count', timestamp=True,
                                     recursive=True)
        if tk.messagebox.askyesno('Open Results?'):
            os.startfile(output_file)

    def one_instrument_count(self, overwrite=True):

        save_location = os.path.join(self.results_folder, "Instrument_Count.xlsx")
        if os.path.exists(save_location) and not overwrite:
            return

        inst_count_xlsx = openpyxl.Workbook()
        ws = inst_count_xlsx.active
        ws.title = 'Instrument Count'

        img = self.cv2img

        labels, boxes, scores = model_predict_on_mosaic(img, self.model_inst, threshold=self.nms_threshold)
        results = zip(labels, boxes, scores)

        inst_img = draw_detection_boxes(img, labels, boxes, scores)
        cv2.imwrite(os.path.join(self.results_folder, 'inst_img.png'), inst_img)

        inst_data = return_inst_data(results, img, self.reader, self.instrument_reader_settings,
                                     inst_labels=self.group_inst, other_labels=self.group_other,
                                     min_scores=self.min_scores, expand=self.object_box_expand,
                                     comment_box_expand=self.comment_box_expand,
                                     ocr_results=self.ocr_results, offset=(0, 0),
                                     radius=self.association_radius,
                                     tag_label_groups=self.tag_label_groups)

        print('there are ', len(inst_data), ' inst_data items')

        just_lines_img = remove_objects_from_image(img, inst_data, self.ocr_results)
        cv2.imwrite(os.path.join(self.results_folder, 'blank.png'), just_lines_img)

        current_row = 1
        for data in inst_data:
            data['pid'] = self.pid
            data['file'] = self.image_path
            if not ws['A1'].value:  # Check if header doesn't exist
                print('making header')
                self.create_excel_header(ws, data)
                current_row += 1

            self.populate_excel_row(ws, data, current_row)
            current_row += 1
        try:
            inst_count_xlsx.save(save_location)
            print(f"File saved as: {save_location}")
        except:
            print(f"Save to {save_location} failed")

    # endregion

    # region UI Helper Methods

    def shift_pressed(self, event):
        self.shift_held = True

    def shift_released(self, event):
        self.shift_held = False

    def ctrl_pressed(self, event):
        self.ctrl_held = True

    def ctrl_released(self, event):
        self.ctrl_held = False

    def show_keybindings(self):
        keybindings = """
            n/N: Next image
            b/B: Previous image
            p/P: Capture PID
            a/A: Capture Instrument Group
            f/F: Capture Line
            e/E: Capture Equipment
            z/Z: Capture Service In
            x/X: Capture Service Out
            w/W: Write Data to Excel
            c/C: Clear Instrument Group
            v/V: Vote to normalize tag numbers
            s/S: Swap Services
            g/G: Set Comment
            left shift: add to service with formatting
            left ctrl: add to service
            """
        tk.messagebox.showinfo("Keybindings", keybindings)

    def show_labels(self):

        tk.messagebox.showinfo("object recognition labels", self.detection_labels)

    def open_console(self):
        self.console_popup.create_console_popup()

    def open_image_path(self):
        os.startfile(self.image_path)

    def open_page_results(self):
        os.startfile(self.results_folder)

    def open_detecto_gui(self):
        detecto_gui_window = tk.Toplevel(self.root)
        ObjectDetectionApp(detecto_gui_window)

    def open_FAIA(self):
        faia_window = tk.Toplevel(self.root)
        FindAnInstrumentApp(faia_window, img_path=self.image_path)

    def open_image_editor(self):
        image_editor_window = tk.Toplevel(self.root)
        ImageEditor(image_editor_window, folder=self.folder_path, image="ocr_capture.png")

    def open_ocr_results_viewer(self):
        ocr_viewer_window = tk.Toplevel(self.root)
        OCRViewer(ocr_viewer_window, self.folder_path)

    def open_workbook(self):
        os.startfile(self.workbook_path)
    # endregion

def set_window_logo(window, png_path, size=(64, 64)):
    """
    Set a PNG image as the window logo.

    Args:
        window: Tkinter window object
        png_path (str): Path to the PNG file
        size (tuple): Desired icon size (width, height)
    """
    try:
        # Open and resize the PNG
        img = Image.open(png_path)
        img = img.resize(size, Image.Resampling.LANCZOS)

        # Save as temporary ICO
        img.save('temp_icon.ico', format='ICO')

        # Set the window icon
        window.iconbitmap('temp_icon.ico')

        # Clean up
        import os
        os.remove('temp_icon.ico')

    except Exception as e:
        print(f"Error setting icon: {e}")

class SplashScreen:
    def __init__(self, parent, image_path):
        self.parent = parent

        # Create a toplevel window
        self.splash = tk.Toplevel(parent)
        self.splash.overrideredirect(True)  # Remove window decorations

        # Configure the window to handle transparency
        self.splash.attributes('-alpha', 1.0)  # Make the window fully opaque
        self.splash.wm_attributes('-transparentcolor', 'black')  # Set transparent color

        # Load the image preserving transparency
        image = Image.open(image_path)
        # Convert to RGBA if it isn't already
        if image.mode != 'RGBA':
            image = image.convert('RGBA')

        splash_width, splash_height = image.size

        # Create a fully transparent background
        background = Image.new('RGBA', image.size, (0, 0, 0, 0))
        # Composite the image onto the transparent background
        background.paste(image, (0, 0), image)

        # Resize if needed
        background = background.resize((splash_width, splash_height), Image.Resampling.LANCZOS)
        self.photo = ImageTk.PhotoImage(background)

        # Create and pack the image label with transparent background
        self.label = tk.Label(self.splash, image=self.photo, border=0, bg='black')
        self.label.pack()

        # Add a loading label with transparent background
        self.loading_label = tk.Label(
            self.splash,
            text="Loading...",
            font=("Arial", 12),
            bg='black',  # Match the transparent color
            fg='white'  # Make text visible
        )
        self.loading_label.pack(pady=10)

        # Configure the splash window background
        self.splash.configure(bg='black')  # Match the transparent color

        # Center the splash screen
        screen_width = parent.winfo_screenwidth()
        screen_height = parent.winfo_screenheight()
        x = (screen_width - splash_width) // 2
        y = (screen_height - splash_height) // 2
        self.splash.geometry(f'{splash_width}x{splash_height}+{x}+{y}')

        # Ensure splash screen is on top
        self.splash.lift()
        self.splash.focus_force()

    def update_status(self, text):
        """Update the loading text"""
        self.loading_label.config(text=text)

    def destroy(self):
        """Destroy the splash screen"""
        self.splash.destroy()

if __name__ == "__main__":
    root = tk.Tk()

    png_path = r"C:\Users\dcaoili\OneDrive - Samuel Engineering\Pictures\logo\LOGO.png"
    set_window_logo(root, png_path)

    app = PIDVisionApp(root)
    root.mainloop()
